import glob
import os
import sys

YOL = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, YOL)

from cetvel import cetvel_parse
from efatura import efatura_parse
from matcher import capraz_kontrol
from utils import tl_format

TEST_KLASORU = os.path.join(YOL, "test_veri")

BASARILI = True


def kontrol(ad, kosul, detay=""):
    global BASARILI
    durum = "TAMAM" if kosul else "HATA"
    if not kosul:
        BASARILI = False
    print(f"  [{durum}] {ad} {detay}")


if __name__ == "__main__":
    fatura_dosyalari = sorted(glob.glob(os.path.join(TEST_KLASORU, "fatura_*.pdf")))
    cetvel_dosyalari = glob.glob(os.path.join(TEST_KLASORU, "kontrol_cetveli.pdf"))

    print("== TOPLU FATURA (çok sayfalı) ==")
    toplu_faturalar = efatura_parse(os.path.join(TEST_KLASORU, "Toplu Fatura Yazdırma.pdf"))
    for f in toplu_faturalar:
        print(f"  sayfa={f['sayfa']} belge={f['belge_no']} matrah={tl_format(f['matrah'])} kdv={tl_format(f['kdv'])} toplam={tl_format(f['toplam'])}")
    kontrol("toplu 2 fatura", len(toplu_faturalar) == 2, f"-> {len(toplu_faturalar)}")
    kontrol("toplu sayfa 1", toplu_faturalar[0]["belge_no"] == "GFE202400000011")
    kontrol("toplu sayfa 2", toplu_faturalar[1]["belge_no"] == "GFE202400000012")
    kontrol("toplu kdv 1", toplu_faturalar[0]["kdv"] == 200)
    kontrol("toplu tutar tutarli", not any(n for n in toplu_faturalar[0]["notlar"] if "≠" in n))

    print("\n== TARANMIŞ CETVEL (OCR) ==")
    from cetvel import cetvel_parse
    tarama_sonuc = cetvel_parse(os.path.join(TEST_KLASORU, "taranmis_cetvel.pdf"))
    print("  notlar:", tarama_sonuc["notlar"])
    for c in tarama_sonuc["kayitlar"][:8]:
        print(f"  vkn={c['vkn']} belge={c['belge_no']} tarih={c['tarih']} matrah={tl_format(c['matrah'])} kdv={tl_format(c['kdv'])}")
    kontrol("tarama kayit sayisi", len(tarama_sonuc["kayitlar"]) >= 5, f"-> {len(tarama_sonuc['kayitlar'])}")
    ilk = tarama_sonuc["kayitlar"][0]
    kontrol("tarama ilk vkn", ilk["vkn"] == "12345678901", f"-> {ilk['vkn']}")
    kontrol("tarama ilk kdv", abs(ilk["kdv"] - 200) < 1, f"-> {ilk['kdv']}")

    print("\n== EXCEL FATURA LİSTESİ ==")
    from dosya import cetvel_dosya_parse, fatura_dosya_parse
    excel_faturalar = fatura_dosya_parse(os.path.join(TEST_KLASORU, "fatura_listesi.xlsx"))
    for f in excel_faturalar:
        print(f"  satir={f['satir']} belge={f['belge_no']} vkn={f['satici_vkn']} matrah={tl_format(f['matrah'])} kdv={tl_format(f['kdv'])} toplam={tl_format(f['toplam'])}")
    kontrol("excel fatura 4 kayit", len(excel_faturalar) == 4, f"-> {len(excel_faturalar)}")
    kontrol("excel fatura 1", excel_faturalar[0]["belge_no"] == "GFE202400000001" and excel_faturalar[0]["matrah"] == 1000)
    kontrol("excel fatura kdv", excel_faturalar[2]["kdv"] == 50)

    print("\n== GELEN FATURALAR (ornek veri) ==")
    gelen_faturalar = fatura_dosya_parse(os.path.join(TEST_KLASORU, "gelen_faturalar.xlsx"))
    for f in gelen_faturalar:
        print(f"  satir={f['satir']} belge={f['belge_no']} vkn={f['satici_vkn']} matrah={tl_format(f['matrah'])} kdv={tl_format(f['kdv'])} toplam={tl_format(f['toplam'])} oran={f['oranlar']}")
    kontrol("gelen fatura 3 kayit", len(gelen_faturalar) == 3, f"-> {len(gelen_faturalar)}")
    kontrol("gelen fatura 1 (%10)", gelen_faturalar[0]["belge_no"] == "ABC202600000101"
            and round(float(gelen_faturalar[0]["matrah"]), 2) == 5786.12
            and round(float(gelen_faturalar[0]["kdv"]), 2) == 578.61
            and gelen_faturalar[0]["oranlar"] == [10])
    kontrol("gelen fatura 2 (%20)", gelen_faturalar[1]["belge_no"] == "ABD202600000102"
            and round(float(gelen_faturalar[1]["matrah"]), 2) == 1179.86
            and round(float(gelen_faturalar[1]["kdv"]), 2) == 235.97
            and gelen_faturalar[1]["oranlar"] == [20])
    kontrol("gelen fatura 3 (%1+%20)", round(float(gelen_faturalar[2]["matrah"]), 2) == round(643.57 + 0.83, 2)
            and round(float(gelen_faturalar[2]["kdv"]), 2) == round(6.44 + 0.17, 2)
            and round(float(gelen_faturalar[2]["toplam"]), 2) == 651.01)

    print("\n== EXCEL CETVEL LİSTESİ ==")
    excel_cetvel = cetvel_dosya_parse(os.path.join(TEST_KLASORU, "kontrol_cetveli.xlsx"))
    print("  notlar:", excel_cetvel["notlar"])
    for c in excel_cetvel["kayitlar"]:
        print(f"  vkn={c['vkn']} belge={c['belge_no']} tarih={c['tarih']} matrah={tl_format(c['matrah'])} kdv={tl_format(c['kdv'])}")
    kontrol("excel cetvel 4 kayit", len(excel_cetvel["kayitlar"]) == 4, f"-> {len(excel_cetvel['kayitlar'])}")
    kontrol("excel cetvel kdv", excel_cetvel["kayitlar"][2]["kdv"] == 45)
    kontrol("excel cetvel toplam satiri atlandi", not any(c["belge_no"] == "GENELTOPLAM" for c in excel_cetvel["kayitlar"]))

    print("\n== EXCEL 391 HESAP DEFTERİ (satış KDV) ==")
    m391 = cetvel_dosya_parse(os.path.join(TEST_KLASORU, "391_hesap.xlsx"))
    print("  notlar:", m391["notlar"])
    for c in m391["kayitlar"]:
        print(f"  {c['belge_no']} tarih={c['tarih']} kdv={tl_format(c['kdv'])} {' '.join(c['notlar'])}")
    kontrol("391 parse 5 kayit", len(m391["kayitlar"]) == 5, f"-> {len(m391['kayitlar'])}")
    kontrol("391 belge eklendi", m391["kayitlar"][0]["belge_no"] == "DNM2026000000201")
    kontrol("391 kdv alacak kolonu", round(float(m391["kayitlar"][2]["kdv"]), 2) == 3150.96)
    grub3 = sum(float(k["kdv"]) for k in m391["kayitlar"] if "391-03-001" in " ".join(k["notlar"]))
    kontrol("391-03-001 toplam (Genel Toplam'e eş)", round(grub3, 2) == 10137.56, f"-> {round(grub3,2)}")

    print("\n== EXCEL 191 HESAP DEFTERİ (alım KDV) ==")
    m191 = cetvel_dosya_parse(os.path.join(TEST_KLASORU, "191_hesap.xlsx"))
    print("  notlar:", m191["notlar"])
    for c in m191["kayitlar"]:
        print(f"  {c['belge_no']} tarih={c['tarih']} kdv={tl_format(c['kdv'])} {' '.join(c['notlar'])}")
    kontrol("191 parse 4 kayit", len(m191["kayitlar"]) == 4, f"-> {len(m191['kayitlar'])}")
    kontrol("191 belge FT.NIZ NO", m191["kayitlar"][0]["belge_no"] == "ORT2026000000301")
    kontrol("191 kdv borc kolonu", round(float(m191["kayitlar"][1]["kdv"]), 2) == 1189.99)
    kontrol("191 TTNET GN belge", any(k["belge_no"] == "A112026000000301" for k in m191["kayitlar"]))
    kontrol("191 TTNET GN kdv", any(k["belge_no"] == "A112026000000301" and round(float(k["kdv"]), 2) == 199.98 for k in m191["kayitlar"]))

    print("\n== XML FATURA (UBL e-fatura) ==")
    from dosya import fatura_dosya_parse as fatura_dosya_parse_fn
    from xml_oku import fatura_xml_parse
    xml_faturalar = []
    for d in sorted(glob.glob(os.path.join(TEST_KLASORU, "fatura_*.xml"))):
        xml_faturalar.extend(fatura_xml_parse(d))
    for f in xml_faturalar:
        print(f"  {os.path.basename(f['dosya'])}: belge={f['belge_no']} tarih={f['tarih']} "
              f"satici={f['satici_vkn']} matrah={tl_format(f['matrah'])} kdv={tl_format(f['kdv'])} toplam={tl_format(f['toplam'])} oran={f['oranlar']}")
    kontrol("xml 5 fatura", len(xml_faturalar) == 5, f"-> {len(xml_faturalar)}")
    kontrol("xml 1 belge", xml_faturalar[0]["belge_no"] == "GFE202400000001")
    kontrol("xml 1 vkn", xml_faturalar[0]["satici_vkn"] == "12345678901")
    kontrol("xml 1 matrah", xml_faturalar[0]["matrah"] == 1000)
    kontrol("xml 1 kdv", xml_faturalar[0]["kdv"] == 200)
    kontrol("xml 1 toplam", xml_faturalar[0]["toplam"] == 1200)
    kontrol("xml 3 kdv", xml_faturalar[2]["kdv"] == 50)
    kontrol("xml 3 oran", 10 in xml_faturalar[2]["oranlar"])
    kontrol("xml 5 vkn", xml_faturalar[4]["satici_vkn"] == "00099988877")
    gz = fatura_xml_parse(os.path.join(TEST_KLASORU, "sikistirilmis_fatura.xml"))
    kontrol("xml gzip acildi", len(gz) == 1 and gz[0]["belge_no"] == "GFE202400000099" and gz[0]["kdv"] == 90, f"-> {gz[0]['belge_no'] if gz else None}")
    xml_dosya_parse = fatura_dosya_parse_fn(os.path.join(TEST_KLASORU, "fatura_1.xml"))
    kontrol("xml dosya yonlendirme", len(xml_dosya_parse) == 1 and xml_dosya_parse[0]["belge_no"] == "GFE202400000001")

    print("\n== XML TİP / VERGİ DETAYI / ORAN KONTROL ==")
    kontrol("xml 1 tip satis", xml_faturalar[0]["fatura_tipi"] == "SATIS", f"-> {xml_faturalar[0]['fatura_tipi']}")
    kontrol("xml 1 vergi detayi var", len(xml_faturalar[0]["vergi_detay"]) >= 1, f"-> {len(xml_faturalar[0]['vergi_detay'])}")
    ilk_detay = xml_faturalar[0]["vergi_detay"][0]
    kontrol("xml 1 detay oran", ilk_detay["oran"] == 20, f"-> {ilk_detay.get('oran')}")
    kontrol("xml 1 detay matrah", ilk_detay["matrah"] == 1000, f"-> {ilk_detay.get('matrah')}")
    kontrol("xml 1 detay kdv", ilk_detay["kdv"] == 200, f"-> {ilk_detay.get('kdv')}")
    kontrol("xml 1 oran kontrol ok", xml_faturalar[0]["oran_kontrol"] == "OK", f"-> {xml_faturalar[0]['oran_kontrol']}")
    kontrol("xml 1 oran notu yok", not any("Matrah×Oran" in n for n in xml_faturalar[0]["notlar"]))
    kontrol("xml 3 oran kontrol", xml_faturalar[2]["oran_kontrol"] in ("OK", ""), f"-> {xml_faturalar[2]['oran_kontrol']}")

    print("\n== GELEN ZIP (telekom + elektrik XML) ==")
    gz_faturalar = fatura_dosya_parse_fn(os.path.join(TEST_KLASORU, "gelen_zip_faturalar.zip"))
    kontrol("gzip 2 kayit", len(gz_faturalar) == 2, f"-> {len(gz_faturalar)}")
    gediz = next((f for f in gz_faturalar if f["belge_no"] == "ORN2026000000001"), None)
    turkcell = next((f for f in gz_faturalar if f["belge_no"] == "TEL2026000000002"), None)
    kontrol("zip gediz elektrik kdv", gediz is not None and round(float(gediz["kdv"]), 2) == 227.27
            and round(float(gediz["kdv_ayrik"]), 2) == 227.27)
    kontrol("zip turkcell sektor", turkcell is not None and turkcell["sektor"] == "TELECOM"
            and round(float(turkcell["kdv"]), 2) == 76.92
            and round(float(turkcell["kdv_ayrik"]), 2) == 76.92
            and round(float(turkcell["diger_vergi_toplam"]), 2) == 65.44)
    kontrol("zip turkcell vergi detay", turkcell is not None and any(
        d.get("ad") == "Özel İletişim Vergisi" and d.get("kod") == "4081" for d in turkcell["vergi_detay"]))

    print("\n== XML ORAN FARKI (NameError regresyon) ==")
    import re as _re
    import tempfile as _tmp
    bozuk_xml = open(os.path.join(TEST_KLASORU, "fatura_1.xml"), "rb").read().decode("utf-8")
    bozuk_xml = _re.sub(
        r"(<cbc:TaxAmount[^>]*>)200\.00(</cbc:TaxAmount>)",
        lambda m: m.group(1) + "250.00" + m.group(2),
        bozuk_xml,
    )
    bozuk_yol = os.path.join(_tmp.gettempdir(), "kdv_bozuk_oran.xml")
    with open(bozuk_yol, "w", encoding="utf-8") as f:
        f.write(bozuk_xml)
    try:
        bozuk_sonuc = fatura_xml_parse(bozuk_yol)
        b_f = bozuk_sonuc[0]
        kontrol("xml oran farki kdv 250", b_f["kdv"] == 250, f"-> {b_f['kdv']}")
        kontrol("xml oran kontrol FARK", b_f["oran_kontrol"] == "FARK", f"-> {b_f['oran_kontrol']}")
        kontrol("xml oran fark notu", any("Matrah×Oran" in n for n in b_f["notlar"]), f"-> {b_f['notlar']}")
    finally:
        if os.path.exists(bozuk_yol):
            os.remove(bozuk_yol)

    print("\n== İADE FATURA EŞLEŞMESİ ==")
    from decimal import Decimal
    iade_f = [{
        "belge_no": "IADE202400000001", "tarih": "2024-02-10",
        "satici_vkn": "12345678901", "satici_unvan": "Test Satıcı",
        "matrah": Decimal("-1000.00"), "kdv": Decimal("-200.00"),
        "toplam": Decimal("-1200.00"), "oranlar": [20], "fatura_tipi": "IADE",
        "oran_kontrol": "OK", "notlar": [], "dosya": "iade.xml", "tip": "xml",
    }]
    iade_c = [{
        "belge_no": "IADE202400000001", "tarih": "2024-02-10",
        "vkn": "12345678901", "unvan": "Test Satıcı",
        "matrah": None, "kdv": Decimal("200.00"), "notlar": [],
    }]
    from matcher import capraz_kontrol_iade_destekli
    i_sonuc, i_ozet = capraz_kontrol_iade_destekli(iade_f, iade_c)
    kontrol("iade eslesen durum", i_sonuc[0]["durum"] == "İADE EŞLEŞTİ", f"-> {i_sonuc[0]['durum']}")
    kontrol("iade kdv abs", i_sonuc[0]["kdv"] == 200, f"-> {i_sonuc[0]['kdv']}")
    kontrol("iade tip tasindi", i_sonuc[0]["tip"] == "IADE", f"-> {i_sonuc[0]['tip']}")
    kontrol("iade oran kontrol", i_sonuc[0]["oran_kontrol"] == "OK", f"-> {i_sonuc[0]['oran_kontrol']}")
    kontrol("iade ozet", i_ozet["iade_adet"] == 1 and i_ozet["fatura_adet"] == 1, f"-> {i_ozet}")

    print("\n== FATURA PARSE ==")
    faturalar = []
    for d in fatura_dosyalari:
        faturalar.extend(efatura_parse(d))
    for f in faturalar:
        print(f"  {os.path.basename(f['dosya'])}: belge={f['belge_no']} tarih={f['tarih']} "
              f"satici={f['satici_vkn']} matrah={tl_format(f['matrah'])} kdv={tl_format(f['kdv'])} toplam={tl_format(f['toplam'])}")

    kontrol("1 no okundu", faturalar[0]["belge_no"] == "GFE202400000001")
    kontrol("1 tarih", faturalar[0]["tarih"] == "2024-02-05")
    kontrol("1 satıcı VKN", faturalar[0]["satici_vkn"] == "12345678901")
    kontrol("1 matrah", faturalar[0]["matrah"] == 1000)
    kontrol("1 kdv", faturalar[0]["kdv"] == 200)
    kontrol("1 toplam", faturalar[0]["toplam"] == 1200)
    kontrol("3 kdv %10", faturalar[2]["kdv"] == 50)
    kontrol("5 vkn", faturalar[4]["satici_vkn"] == "00099988877")

    print("\n== CETVEL PARSE ==")
    cetvel_sonuc = cetvel_parse(cetvel_dosyalari[0])
    for c in cetvel_sonuc["kayitlar"]:
        print(f"  vkn={c['vkn']} belge={c['belge_no']} tarih={c['tarih']} "
              f"matrah={tl_format(c['matrah'])} kdv={tl_format(c['kdv'])} unvan={c['unvan']}")
    kontrol("cetvel 6 kayit", len(cetvel_sonuc["kayitlar"]) == 6, f"-> {len(cetvel_sonuc['kayitlar'])}")
    kontrol("cetvel 1 belge no", cetvel_sonuc["kayitlar"][0]["belge_no"] == "GFE202400000001")
    kontrol("cetvel 1 kdv", cetvel_sonuc["kayitlar"][0]["kdv"] == 200)
    kontrol("cetvel 3 kdv 45", cetvel_sonuc["kayitlar"][2]["kdv"] == 45)
    kontrol("cetvel 4 belge", cetvel_sonuc["kayitlar"][3]["belge_no"] == "GFE202400000009")
    kontrol("cetvel 6 vkn", cetvel_sonuc["kayitlar"][5]["vkn"] == "11122233344")

    print("\n== ÇAPRAZ KONTROL ==")
    sonuclar, ozet = capraz_kontrol(faturalar, cetvel_sonuc["kayitlar"])
    for s in sonuclar:
        print(f"  [{s['durum']}] {s['belge_no'] or ''} vkn={s['vkn'] or ''} "
              f"matrah={tl_format(s['matrah'])} kdv={tl_format(s['kdv'])} {s['detay']}")
    print("  OZET:", ozet)

    kontrol("eslesen 2", ozet["eslesen"] == 2, f"-> {ozet['eslesen']}")
    kontrol("tutar farki 1", ozet["tutar_farki"] == 1, f"-> {ozet['tutar_farki']}")
    kontrol("vkn farki 1", ozet["vkn_farki"] == 1, f"-> {ozet['vkn_farki']}")
    kontrol("cetvelde yok 1", ozet["cetvelde_yok"] == 1, f"-> {ozet['cetvelde_yok']}")
    kontrol("faturada yok 1", ozet["faturada_yok"] == 1, f"-> {ozet['faturada_yok']}")
    kontrol("mukerrer 1", ozet["mukerrer"] == 1, f"-> {ozet['mukerrer']}")

    print("\n== İADE DURUMLU EXCEL RAPORU (StopIteration regresyon) ==")
    import tempfile as _tmp2
    from report import rapor_olustur as _rapor_olustur
    iade_rap_satirlar = sonuclar + i_sonuc
    iade_rap_ozet = dict(ozet)
    iade_rap_ozet["fatura_adet"] = iade_rap_ozet.get("fatura_adet", 0) + 1
    iade_rap_yol = os.path.join(_tmp2.gettempdir(), "kdv_iade_rapor.xlsx")
    try:
        _rapor_olustur(iade_rap_satirlar, iade_rap_ozet, faturalar + iade_f,
                       cetvel_sonuc["kayitlar"] + iade_c, iade_rap_yol)
        kontrol("iade excel raporu olustu", os.path.exists(iade_rap_yol) and os.path.getsize(iade_rap_yol) > 0)
    finally:
        if os.path.exists(iade_rap_yol):
            os.remove(iade_rap_yol)

    print("\n== MAHSUP FİŞİ (fiş listesi PDF) ==")
    fis_faturalar = fatura_dosya_parse(os.path.join(TEST_KLASORU, "fis_listesi_ornek.pdf"))
    for f in fis_faturalar:
        print(f"  belge={f['belge_no']} tarih={f['tarih']} tip={f['fatura_tipi']} unvan={f['satici_unvan']!r} "
              f"matrah={tl_format(f['matrah'])} kdv={tl_format(f['kdv'])} toplam={tl_format(f['toplam'])} oran={f['oranlar']}")
    kontrol("fis 3 kayit", len(fis_faturalar) == 3, f"-> {len(fis_faturalar)}")
    z = next(f for f in fis_faturalar if f["belge_no"] == "Z2127")
    kontrol("fis Z tip", z["fatura_tipi"] == "Z RAPORU", f"-> {z['fatura_tipi']}")
    kontrol("fis Z kdv", z["kdv"] == Decimal("267.71"), f"-> {z['kdv']}")
    kontrol("fis Z matrah", z["matrah"] == Decimal("15133.50"), f"-> {z['matrah']}")
    kontrol("fis Z oranlar", z["oranlar"] == [1, 20], f"-> {z['oranlar']}")
    kontrol("fis Z tarih", z["tarih"] == "2026-07-02", f"-> {z['tarih']}")
    kontrol("fis Z vkn bos", z["satici_vkn"] == "", f"-> {z['satici_vkn']!r}")
    kontrol("fis Z not fis", any("MAHSUP fişi 001851 (Z RAPORU)" in n for n in z["notlar"]), f"-> {z['notlar']}")
    kontrol("fis Z not matrah", any("oranlarından hesaplandı" in n for n in z["notlar"]))
    ear = next(f for f in fis_faturalar if f["belge_no"] == "EAR2026000000001")
    kontrol("fis EAR tip", ear["fatura_tipi"] == "E-ARSIV", f"-> {ear['fatura_tipi']}")
    kontrol("fis EAR kdv", ear["kdv"] == Decimal("1030.83"), f"-> {ear['kdv']}")
    kontrol("fis EAR matrah", ear["matrah"] == Decimal("10193.55"), f"-> {ear['matrah']}")
    kontrol("fis EAR oranlar", ear["oranlar"] == [1, 10, 20], f"-> {ear['oranlar']}")
    kontrol("fis EAR unvan", ear["satici_unvan"] == "ÖRNEK MÜŞTERİ A.Ş.", f"-> {ear['satici_unvan']!r}")
    efa = next(f for f in fis_faturalar if f["belge_no"] == "EFA2026000000002")
    kontrol("fis EFA tip", efa["fatura_tipi"] == "E-FATURA", f"-> {efa['fatura_tipi']}")
    kontrol("fis EFA kdv", efa["kdv"] == Decimal("29.70"), f"-> {efa['kdv']}")
    kontrol("fis EFA matrah", efa["matrah"] == Decimal("2970.00"), f"-> {efa['matrah']}")
    kontrol("fis EFA oranlar", efa["oranlar"] == [1], f"-> {efa['oranlar']}")
    kontrol("fis EFA unvan", efa["satici_unvan"] == "ÖRNEK TEDARİK LTD.ŞTİ.", f"-> {efa['satici_unvan']!r}")

    print("\n== MAHSUP FİŞİ CETVEL TARAFI + ÇAPRAZ KONTROL ==")
    fis_cetvel = cetvel_dosya_parse(os.path.join(TEST_KLASORU, "fis_listesi_ornek.pdf"))
    print("  notlar:", fis_cetvel["notlar"])
    for c in fis_cetvel["kayitlar"]:
        print(f"  belge={c['belge_no']} tarih={c['tarih']} matrah={tl_format(c['matrah'])} kdv={tl_format(c['kdv'])}")
    kontrol("fis cetvel 3 kayit", len(fis_cetvel["kayitlar"]) == 3, f"-> {len(fis_cetvel['kayitlar'])}")
    kontrol("fis cetvel vkn bos", all(c["vkn"] == "" for c in fis_cetvel["kayitlar"]))
    fis_sonuc, fis_ozet = capraz_kontrol(fis_faturalar, fis_cetvel["kayitlar"])
    for s in fis_sonuc:
        print(f"  [{s['durum']}] {s['belge_no']}")
    print("  OZET:", fis_ozet)
    kontrol("fis capraz 3 eslesen", fis_ozet["eslesen"] == 3, f"-> {fis_ozet['eslesen']}")
    kontrol("fis capraz sorun yok", fis_ozet["cetvelde_yok"] == 0 and fis_ozet["faturada_yok"] == 0,
            f"-> cetvelde_yok={fis_ozet['cetvelde_yok']} faturada_yok={fis_ozet['faturada_yok']}")

    print("\n== MAHSUP FİŞİ ↔ MUAVİN (hesap bazlı çapraz kontrol) ==")
    from fis_listesi import fis_listesi_hesap_parse
    from excel_oku import muavin_satis_parse
    from matcher import z_raporu_hesap_kontrol
    fis_hesap = fis_listesi_hesap_parse(os.path.join(TEST_KLASORU, "fis_listesi_ornek.pdf"))
    muavin_sonuc = muavin_satis_parse(os.path.join(TEST_KLASORU, "muavin_satis.xlsx"))
    muavin_hesap = muavin_sonuc["kayitlar"]
    print("  fis_hesap kayit sayisi:", len(fis_hesap))
    print("  muavin kayit sayisi:", len(muavin_hesap), "notlar:", muavin_sonuc["notlar"])
    for k in fis_hesap[:4]:
        print(f"  fis {k['belge']} hesap={k['hesap']} alacak={tl_format(k['alacak'])}")
    kontrol("muavin not var", bool(muavin_sonuc["notlar"]))
    kontrol("fis hesap 16 kayit", len(fis_hesap) == 16, f"-> {len(fis_hesap)}")
    kontrol("muavin 6 kayit", len(muavin_hesap) == 6, f"-> {len(muavin_hesap)}")
    kontrol("muavin Z2127", any(k["belge"] == "Z2127" and k["hesap"] == "600.01.002" and k["alacak"] == Decimal("13909.90") for k in muavin_hesap))

    mh_sonuc, mh_ozet = z_raporu_hesap_kontrol(fis_hesap, muavin_hesap)
    for s in mh_sonuc:
        print(f"  [{s['durum']}] {s['belge_no'] or ''} hesap={s['unvan']} tutar={tl_format(s['matrah'])} {s['detay']}")
    print("  OZET:", mh_ozet)
    kontrol("muavin hesap eslesen 2", mh_ozet["eslesen"] == 2, f"-> {mh_ozet['eslesen']}")
    kontrol("muavin hesap tutar farki 0", mh_ozet["tutar_farki"] == 0, f"-> {mh_ozet['tutar_farki']}")
    kontrol("muavin hesap cetvelde yok 14", mh_ozet["cetvelde_yok"] == 14, f"-> {mh_ozet['cetvelde_yok']}")
    kontrol("muavin hesap faturada yok 4", mh_ozet["faturada_yok"] == 4, f"-> {mh_ozet['faturada_yok']}")

    print("\n== KARIŞIK ÇAPRAZ KONTROL (PDF + Excel) ==")
    karisik_faturalar = list(faturalar) + excel_faturalar
    karisik_cetvel = cetvel_sonuc["kayitlar"] + excel_cetvel["kayitlar"]
    k_sonuc, k_ozet = capraz_kontrol(karisik_faturalar, karisik_cetvel)
    for s in k_sonuc:
        print(f"  [{s['durum']}] {s['belge_no'] or ''} vkn={s['vkn'] or ''} {s['detay']}")
    print("  OZET:", k_ozet)
    kontrol("karisik eslesen 4", k_ozet["eslesen"] == 4, f"-> {k_ozet['eslesen']}")
    kontrol("karisik tutar farki 2", k_ozet["tutar_farki"] == 2, f"-> {k_ozet['tutar_farki']}")

    print("\n== ÖZETLER (KDV DAĞILIMI / BA FORMU / EKSİK BELGELER) ==")
    from ozetler import ba_formu, eksik_belgeler, kdv_dagilim_fatura, kdv_dagilim_muavin
    fatura_dag = kdv_dagilim_fatura(xml_faturalar)
    kontrol("fatura dagilim %20 var", 20 in fatura_dag, f"-> {fatura_dag.get(20)}")
    kontrol("fatura dagilim %20 adet", fatura_dag[20]["adet"] == 3 and fatura_dag[20]["kdv"] == 860,
            f"-> {fatura_dag.get(20)}")
    kontrol("fatura dagilim %10 var", 10 in fatura_dag, f"-> {fatura_dag.get(10)}")
    muavin_dag = kdv_dagilim_muavin(cetvel_sonuc["kayitlar"])
    kontrol("muavin dagilim kayitli", sum(g["adet"] for g in muavin_dag.values()) == len(cetvel_sonuc["kayitlar"]),
            f"-> {len(cetvel_sonuc['kayitlar'])}")
    ba = ba_formu(xml_faturalar)
    kontrol("ba formu saticilar", len(ba) >= 3, f"-> {len(ba)}")
    kontrol("ba formu adet toplami", sum(s["adet"] for s in ba) == len(xml_faturalar))
    eksikler = eksik_belgeler(sonuclar)
    kontrol("eksik belgeler", eksikler == ["GFE202400000004"], f"-> {eksikler}")

    print("\n== GEÇMİŞ KARŞILAŞTIRMA (config) ==")
    import config
    import tempfile
    config.GECMIS_YOLU = os.path.join(tempfile.gettempdir(), "kdv_test_gecmis.json")
    if os.path.exists(config.GECMIS_YOLU):
        os.remove(config.GECMIS_YOLU)
    kontrol("gecmis bos", config.gecmis_karsilastir(eksikler) is None)
    config.gecmis_ekle(ozet, ["GFE202400000005", "GFE202400000006"])
    gecmis = config.gecmis_karsilastir(["GFE202400000005"])
    kontrol("gecmis kapanan", gecmis is not None and gecmis["kapanan"] == ["GFE202400000006"], f"-> {gecmis}")
    kontrol("gecmis yeni yok", gecmis["yeni"] == [])
    gecmis2 = config.gecmis_karsilastir(["GFE202400000005", "GFE202400000007"])
    kontrol("gecmis yeni", gecmis2["yeni"] == ["GFE202400000007"], f"-> {gecmis2['yeni']}")
    os.remove(config.GECMIS_YOLU)

    print("\n== EXCEL RAPORU ==")
    from report import rapor_olustur
    hedef = os.path.join(YOL, "test_rapor.xlsx")
    gecmis_bilgi = {"kapanan": [], "yeni": ["GFE202400000006"], "onceki_eslesen": 1, "zaman": "01.01.2024 10:00"}
    rapor_olustur(sonuclar, ozet, faturalar, cetvel_sonuc["kayitlar"], hedef, gecmis_bilgi=gecmis_bilgi)
    kontrol("excel olusturuldu", os.path.exists(hedef) and os.path.getsize(hedef) > 0)
    from openpyxl import load_workbook
    wb = load_workbook(hedef)
    kontrol("excel yeni sayfalar", {"KDVDagilimi", "BaFormu", "Grafik"} <= set(wb.sheetnames), f"-> {wb.sheetnames}")
    kontrol("sonuclar tip kolonu", wb["Sonuclar"].cell(row=1, column=6).value == "Tip")
    kontrol("eksik fatura tip kolonu", wb["EksikFaturalar"].cell(row=1, column=3).value == "Tip")
    kontrol("ozet gecmis satiri", any("SON KONTROLE" in str(c.value) for c in wb["Ozet"]["A"]),
            "-> ozet degisim bolumu")

    print("\n== PDF RAPORU ==")
    from report_pdf import rapor_pdf_olustur
    hedef_pdf = os.path.join(YOL, "test_rapor.pdf")
    rapor_pdf_olustur(sonuclar, ozet, faturalar, cetvel_sonuc["kayitlar"], hedef_pdf, gecmis_bilgi=gecmis_bilgi)
    kontrol("pdf olusturuldu", os.path.exists(hedef_pdf) and os.path.getsize(hedef_pdf) > 0)
    from pypdf import PdfReader
    pdf_metin = "\n".join((sayfa.extract_text() or "") for sayfa in PdfReader(hedef_pdf).pages)
    kontrol("pdf sonuclar tip", "Tip" in pdf_metin)
    kontrol("pdf kdv dagilim", "KDV DAĞILIMI" in pdf_metin.replace("İ", "I").upper())
    kontrol("pdf ba formu", "BA FORMU" in pdf_metin.replace("İ", "I").upper())
    kontrol("pdf gecmis degisim", "DEĞİŞİM" in pdf_metin)

    print("\n== İADE FATURA KONTROLÜ ==")
    from matcher import capraz_kontrol_iade_destekli
    iade_fats = [
        {'belge_no': 'EFA2026000000002', 'satici_vkn': '12345678901', 'tarih': '2026-07-30',
         'matrah': Decimal('-2970.00'), 'kdv': Decimal('-29.70'), 'toplam': Decimal('-3267.00'),
         'fatura_tipi': 'IADE', 'oranlar': [1]},
    ]
    iade_cetvel = [
        {'belge_no': 'EFA2026000000002', 'vkn': '12345678901', 'tarih': '2026-07-30',
         'matrah': None, 'kdv': Decimal('29.70'), 'unvan': 'DENEME TİCARET'},
    ]
    iade_sonuc, iade_ozet = capraz_kontrol_iade_destekli(iade_fats, iade_cetvel)
    iade_durum = iade_sonuc[0]["durum"] if iade_sonuc else "?"
    kontrol("iade eslesti (negatif kdv, pozitif muavin)", iade_durum == "İADE EŞLEŞTİ", f"-> {iade_durum}")
    kontrol("iade muavinde yok sayaci 0", iade_ozet.get("iade_muavinde_yok", -1) == 0)

    print("\n== BELGE NO ÇAKIŞMASI (aynı belge farklı faturalar) ==")
    from dosya import fatura_birlestir
    from excel_oku import _muavin_birlestir

    def _f(belge, tarih, vkn, matrah, kdv):
        return {"belge_no": belge, "tarih": tarih, "satici_vkn": vkn,
                "satici_unvan": "", "matrah": Decimal(matrah), "kdv": Decimal(kdv),
                "toplam": Decimal(matrah) + Decimal(kdv), "oranlar": [20],
                "fatura_tipi": "", "oran_kontrol": "OK", "notlar": [], "dosya": "t.xml"}

    def _c(belge, tarih, vkn, kdv):
        return {"belge_no": belge, "vkn": vkn, "tarih": tarih,
                "matrah": None, "kdv": Decimal(kdv), "unvan": "X", "notlar": []}

    # 1) Gerçek kopya tekilleşir (iki kez indirme)
    a = _f("GIB2026000000099", "2026-07-01", "11111111111", "13000.00", "2600.00")
    kontrol("kopya tekillşti", len(fatura_birlestir([a, dict(a)])) == 1)
    # 2) Aynı belge farklı tarihli İKİ FARKLI fatura ayrı kalır
    b2 = _f("GIB2026000000099", "2026-07-10", "22222222222", "6500.00", "1300.00")
    kontrol("farklı içerik ayrı kaldı", len(fatura_birlestir([a, dict(a), b2])) == 2)

    # 3) Muavin parça satırları (191-01 + 191-03) birleşir; farklı tarih ayrı kalır
    parcalar = [
        _c("GIB2026000000011", "2026-07-01", "33333333333", "2080.00"),
        _c("GIB2026000000011", "2026-07-01", "33333333333", "520.00"),
        _c("GIB2026000000011", "2026-07-10", "44444444444", "2080.00"),
        _c("GIB2026000000011", "2026-07-10", "44444444444", "520.00"),
    ]
    birlesik = _muavin_birlestir(parcalar)
    kontrol("muavin parça satır birleşti",
            len(birlesik) == 2 and all(float(r["kdv"]) == 2600 for r in birlesik),
            f"-> {[(r['tarih'], str(r['kdv'])) for r in birlesik]}")

    # 4) GIB0013 senaryosu: 2600 eşleşir + 2166.67 cetvelde yok (TUTAR FARKI DEĞİL)
    cak_f = [
        _f("GIB2026000000013", "2026-07-03", "55555555555", "13000.00", "2600.00"),
        _f("GIB2026000000013", "2026-08-10", "66666666666", "10833.33", "2166.67"),
    ]
    cak_c = [_c("GIB2026000000013", "2026-07-03", "55555555555", "2600.00")]
    cak_sonuc, cak_ozet = capraz_kontrol_iade_destekli(cak_f, cak_c)
    kontrol("çakışma: 2600 eşleşti", cak_ozet["eslesen"] == 1, f"-> {cak_ozet['eslesen']}")
    kontrol("çakışma: tutar farkı 0", cak_ozet["tutar_farki"] == 0, f"-> {cak_ozet['tutar_farki']}")
    kontrol("çakışma: fazlalık CETVELDE YOK", cak_ozet["cetvelde_yok"] == 1, f"-> {cak_ozet['cetvelde_yok']}")

    # 5) Aynı belgede iki tarihli iki fatura ↔ muavin iki kayıt (GIB0011 senaryosu)
    cift_f = [
        _f("GIB2026000000011", "2026-07-01", "33333333333", "13000.00", "2600.00"),
        _f("GIB2026000000011", "2026-07-10", "44444444444", "13000.00", "2600.00"),
    ]
    cift_sonuc, cift_ozet = capraz_kontrol_iade_destekli(
        fatura_birlestir(cift_f), _muavin_birlestir(parcalar))
    kontrol("çift tarih: ikisi de eşleşti", cift_ozet["eslesen"] == 2, f"-> {cift_ozet['eslesen']}")
    kontrol("çift tarih: fark yok", cift_ozet["tutar_farki"] == 0 and cift_ozet["mukerrer"] == 0,
            f"-> {cift_ozet}")

    print("\nSONUÇ:", "TÜM TESTLER TAMAM" if BASARILI else "HATALAR VAR")
    sys.exit(0 if BASARILI else 1)
