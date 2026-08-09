"""KDV Beyannamesi için Ba/Bs formu çıktısı hazırla (GİB formatı)."""
import os
from datetime import datetime
from decimal import Decimal
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def ba_formu_olustur(
    faturalar: List[Dict],
    cetvel_kayitlari: List[Dict],
    hedef_yol: str,
    donem: str = "",
) -> str:
    """KDV Beyannamesi için Ba/Bs formu verisi hazırla.

    - Satıcı bazında: VKN, ünvan, fatura adedi, matrah, KDV
    - Oran bazında dağılım (1, 10, 20)
    - Hem alış hem satış için ayrı sayfa
    """
    wb = Workbook()

    INCE = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    BASLIK_FONT = Font(bold=True, color="FFFFFF", size=11)
    BASLIK_DOLGU = PatternFill("solid", fgColor="4472C4")
    TOPLAM_DOLGU = PatternFill("solid", fgColor="D9E1F2")

    # ----- 1) SATIŞ FATURALARI (Alışlar için) -----
    ws1 = wb.active
    ws1.title = "Alış Faturaları"
    satislar = {}
    for f in faturalar:
        vkn = f.get("satici_vkn") or "BILINMIYOR"
        unvan = f.get("satici_unvan") or "Bilinmeyen"
        anahtar = (vkn, unvan)
        g = satislar.setdefault(anahtar, {"adet": 0, "matrah": Decimal("0"), "kdv": Decimal("0")})
        g["adet"] += 1
        g["matrah"] += abs(f.get("matrah") or 0)
        g["kdv"] += abs(f.get("kdv") or 0)

    baslik = ws1.cell(row=1, column=1, value=f"KDV ALIŞ (BA) FORMU - {donem or datetime.now().strftime('%Y-%m')}")
    baslik.font = Font(bold=True, size=14)
    ws1.cell(row=2, column=1, value=f"Oluşturma: {datetime.now().strftime('%d.%m.%Y %H:%M')}")

    basliklar = ["VKN", "Ünvan", "Belge Sayısı", "Matrah (TL)", "KDV (TL)"]
    for j, b in enumerate(basliklar, 1):
        h = ws1.cell(row=4, column=j, value=b)
        h.font = BASLIK_FONT
        h.fill = BASLIK_DOLGU
        h.alignment = Alignment(horizontal="center")
        h.border = INCE

    satir_no = 5
    toplam_matrah = Decimal("0")
    toplam_kdv = Decimal("0")
    toplam_adet = 0
    for (vkn, unvan), g in sorted(satislar.items(), key=lambda x: -x[1]["kdv"]):
        ws1.cell(row=satir_no, column=1, value=vkn)
        ws1.cell(row=satir_no, column=2, value=unvan)
        ws1.cell(row=satir_no, column=3, value=g["adet"])
        ws1.cell(row=satir_no, column=4, value=float(g["matrah"]))
        ws1.cell(row=satir_no, column=5, value=float(g["kdv"]))
        for j in range(1, 6):
            ws1.cell(row=satir_no, column=j).border = INCE
            if j in (4, 5):
                ws1.cell(row=satir_no, column=j).number_format = "#,##0.00"
                ws1.cell(row=satir_no, column=j).alignment = Alignment(horizontal="right")
        toplam_matrah += g["matrah"]
        toplam_kdv += g["kdv"]
        toplam_adet += g["adet"]
        satir_no += 1

    # Toplam
    ws1.cell(row=satir_no, column=1, value="TOPLAM").font = Font(bold=True)
    ws1.cell(row=satir_no, column=3, value=toplam_adet)
    ws1.cell(row=satir_no, column=4, value=float(toplam_matrah))
    ws1.cell(row=satir_no, column=5, value=float(toplam_kdv))
    for j in range(1, 6):
        h = ws1.cell(row=satir_no, column=j)
        h.fill = TOPLAM_DOLGU
        h.border = INCE
        if j in (4, 5):
            h.number_format = "#,##0.00"

    for j, w in enumerate([14, 40, 14, 16, 16], 1):
        ws1.column_dimensions[chr(64 + j)].width = w

    # ----- 2) ORAN BAZLI DAĞILIM -----
    ws2 = wb.create_sheet("Oran Dağılımı")
    ws2.cell(row=1, column=1, value="KDV ORAN BAZLI DAĞILIM").font = Font(bold=True, size=14)

    basliklar2 = ["KDV Oranı", "Fatura Adedi", "Toplam Matrah (TL)", "Toplam KDV (TL)"]
    for j, b in enumerate(basliklar2, 1):
        h = ws2.cell(row=3, column=j, value=b)
        h.font = BASLIK_FONT
        h.fill = BASLIK_DOLGU
        h.alignment = Alignment(horizontal="center")
        h.border = INCE

    oran_dagilim = {}
    for f in faturalar:
        for o in f.get("oranlar") or [0]:
            g = oran_dagilim.setdefault(o, {"adet": 0, "matrah": Decimal("0"), "kdv": Decimal("0")})
            g["adet"] += 1
            g["matrah"] += abs(f.get("matrah") or 0)
            g["kdv"] += abs(f.get("kdv") or 0)

    satir = 4
    toplam_m = Decimal("0")
    toplam_k = Decimal("0")
    toplam_a = 0
    for oran in sorted(oran_dagilim.keys()):
        g = oran_dagilim[oran]
        ws2.cell(row=satir, column=1, value=f"%{oran}" if oran else "Belirsiz")
        ws2.cell(row=satir, column=2, value=g["adet"])
        ws2.cell(row=satir, column=3, value=float(g["matrah"]))
        ws2.cell(row=satir, column=4, value=float(g["kdv"]))
        for j in range(1, 5):
            ws2.cell(row=satir, column=j).border = INCE
            if j in (3, 4):
                ws2.cell(row=satir, column=j).number_format = "#,##0.00"
        toplam_m += g["matrah"]
        toplam_k += g["kdv"]
        toplam_a += g["adet"]
        satir += 1

    ws2.cell(row=satir, column=1, value="TOPLAM").font = Font(bold=True)
    ws2.cell(row=satir, column=2, value=toplam_a)
    ws2.cell(row=satir, column=3, value=float(toplam_m))
    ws2.cell(row=satir, column=4, value=float(toplam_k))
    for j in range(1, 5):
        ws2.cell(row=satir, column=j).fill = TOPLAM_DOLGU
        ws2.cell(row=satir, column=j).border = INCE

    for j, w in enumerate([14, 14, 22, 22], 1):
        ws2.column_dimensions[chr(64 + j)].width = w

    # ----- 3) İADE FATURALARI -----
    iadeler = [f for f in faturalar if (f.get("tip") or "").upper() in {"IADE", "CREDIT_NOTE"}]
    if iadeler:
        ws3 = wb.create_sheet("İade Faturaları")
        basliklar3 = ["Tarih", "Belge No", "VKN", "Ünvan", "Matrah", "KDV"]
        for j, b in enumerate(basliklar3, 1):
            h = ws3.cell(row=1, column=j, value=b)
            h.font = BASLIK_FONT
            h.fill = BASLIK_DOLGU
            h.alignment = Alignment(horizontal="center")
            h.border = INCE

        toplam_m_i = Decimal("0")
        toplam_k_i = Decimal("0")
        for i, f in enumerate(iadeler, 2):
            ws3.cell(row=i, column=1, value=f.get("tarih") or "")
            ws3.cell(row=i, column=2, value=f.get("belge_no") or "")
            ws3.cell(row=i, column=3, value=f.get("satici_vkn") or "")
            ws3.cell(row=i, column=4, value=f.get("satici_unvan") or "")
            ws3.cell(row=i, column=5, value=float(abs(f.get("matrah") or 0)))
            ws3.cell(row=i, column=6, value=float(abs(f.get("kdv") or 0)))
            for j in range(1, 7):
                ws3.cell(row=i, column=j).border = INCE
                if j in (5, 6):
                    ws3.cell(row=i, column=j).number_format = "#,##0.00"
            toplam_m_i += abs(f.get("matrah") or 0)
            toplam_k_i += abs(f.get("kdv") or 0)

        toplam_row = 2 + len(iadeler)
        ws3.cell(row=toplam_row, column=1, value="TOPLAM").font = Font(bold=True)
        ws3.cell(row=toplam_row, column=5, value=float(toplam_m_i))
        ws3.cell(row=toplam_row, column=6, value=float(toplam_k_i))
        for j in range(1, 7):
            ws3.cell(row=toplam_row, column=j).fill = TOPLAM_DOLGU
            ws3.cell(row=toplam_row, column=j).border = INCE

        for j, w in enumerate([12, 22, 14, 35, 16, 16], 1):
            ws3.column_dimensions[chr(64 + j)].width = w

    wb.save(hedef_yol)
    return hedef_yol
