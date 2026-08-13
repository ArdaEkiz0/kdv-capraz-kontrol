import os

from fpdf import FPDF

YOL = os.path.dirname(os.path.abspath(__file__))
TEST_KLASORU = os.path.join(YOL, "test_veri")
FONT_YOLU = r"C:\Windows\Fonts\arial.ttf"


def yeni_pdf():
    pdf = FPDF()
    pdf.add_font("Arial", "", FONT_YOLU, uni=True)
    pdf.add_font("Arial", "B", r"C:\Windows\Fonts\arialbd.ttf", uni=True)
    pdf.set_font("Arial", "", 10)
    pdf.add_page()
    return pdf


def fatura_pdf(ad, belge_no, tarih, satici_vkn, alici_vkn, matrah, kdv, toplam, oran):
    pdf = yeni_pdf()
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 10, "E-FATURA", ln=1, align="C")
    pdf.set_font("Arial", "", 10)
    pdf.ln(4)
    pdf.cell(0, 6, f"Belge No : {belge_no}", ln=1)
    pdf.cell(0, 6, f"Düzenlenme Tarihi : {tarih}", ln=1)
    pdf.ln(4)
    pdf.cell(0, 6, "SATAN", ln=1)
    pdf.cell(0, 6, f"Ünvan : ÖRNEK SANAYİ VE TİCARET LTD. ŞTİ.", ln=1)
    pdf.cell(0, 6, f"Vergi Kimlik No : {satici_vkn}", ln=1)
    pdf.ln(4)
    pdf.cell(0, 6, "ALICI", ln=1)
    pdf.cell(0, 6, f"Ünvan : MÜŞTERİ TİCARET A.Ş.", ln=1)
    pdf.cell(0, 6, f"Vergi Kimlik No : {alici_vkn}", ln=1)
    pdf.ln(6)
    pdf.cell(0, 6, "Mal Hizmet Toplam Tutarı : " + tl(matrah), ln=1)
    pdf.cell(0, 6, f"Hesaplanan KDV(%{oran}) : " + tl(kdv), ln=1)
    pdf.cell(0, 6, "Toplam : " + tl(toplam), ln=1)
    pdf.output(os.path.join(TEST_KLASORU, ad))
    print("olusturuldu:", ad)


def cetvel_pdf(ad, satirlar):
    pdf = yeni_pdf()
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 8, "KDV KONTROL CETVELİ", ln=1, align="C")
    pdf.set_font("Arial", "B", 8)
    pdf.cell(0, 6, "Dönem : 2024/02", ln=1, align="C")
    pdf.ln(3)
    pdf.set_font("Arial", "B", 8)
    pdf.cell(8, 6, "Sıra", border=1, align="C")
    pdf.cell(32, 6, "VKN", border=1, align="C")
    pdf.cell(55, 6, "Ünvan", border=1, align="C")
    pdf.cell(40, 6, "Belge No", border=1, align="C")
    pdf.cell(24, 6, "Tarih", border=1, align="C")
    pdf.cell(25, 6, "Matrah", border=1, align="C")
    pdf.cell(25, 6, "KDV", border=1, align="C")
    pdf.ln()
    pdf.set_font("Arial", "", 8)
    for i, s in enumerate(satirlar, start=1):
        pdf.cell(8, 6, str(i), border=1, align="C")
        pdf.cell(32, 6, s["vkn"], border=1, align="C")
        pdf.cell(55, 6, s["unvan"], border=1, align="C")
        pdf.cell(40, 6, s["belge_no"], border=1, align="C")
        pdf.cell(24, 6, s["tarih"], border=1, align="C")
        pdf.cell(25, 6, s["matrah"], border=1, align="C")
        pdf.cell(25, 6, s["kdv"], border=1, align="C")
        pdf.ln()
    pdf.output(os.path.join(TEST_KLASORU, ad))
    print("olusturuldu:", ad)


def toplu_fatura_pdf(ad, faturalar):
    pdf = yeni_pdf()
    for i, (belge_no, tarih, satici_vkn, alici_vkn, matrah, kdv, toplam, oran) in enumerate(faturalar):
        if i > 0:
            pdf.add_page()
        pdf.set_font("Arial", "B", 14)
        pdf.cell(0, 10, "E-FATURA", ln=1, align="C")
        pdf.set_font("Arial", "", 10)
        pdf.ln(4)
        pdf.cell(0, 6, f"Belge No : {belge_no}", ln=1)
        pdf.cell(0, 6, f"Düzenlenme Tarihi : {tarih}", ln=1)
        pdf.ln(4)
        pdf.cell(0, 6, "SATAN", ln=1)
        pdf.cell(0, 6, f"Ünvan : ÖRNEK SANAYİ VE TİCARET LTD. ŞTİ.", ln=1)
        pdf.cell(0, 6, f"Vergi Kimlik No : {satici_vkn}", ln=1)
        pdf.ln(4)
        pdf.cell(0, 6, "ALICI", ln=1)
        pdf.cell(0, 6, f"Ünvan : MÜŞTERİ TİCARET A.Ş.", ln=1)
        pdf.cell(0, 6, f"Vergi Kimlik No : {alici_vkn}", ln=1)
        pdf.ln(6)
        pdf.cell(0, 6, "Mal Hizmet Toplam Tutarı : " + tl(matrah), ln=1)
        pdf.cell(0, 6, f"Hesaplanan KDV(%{oran}) : " + tl(kdv), ln=1)
        pdf.cell(0, 6, "Toplam : " + tl(toplam), ln=1)
    pdf.output(os.path.join(TEST_KLASORU, ad))
    print("olusturuldu:", ad)


def taranmis_pdf(ad, kaynak):
    import pymupdf
    doc = pymupdf.open(os.path.join(TEST_KLASORU, kaynak))
    yeni = pymupdf.open()
    for sayfa in doc:
        pix = sayfa.get_pixmap(matrix=pymupdf.Matrix(2, 2), alpha=False)
        gorsel = pix.tobytes("png")
        yeni.new_page(width=sayfa.rect.width, height=sayfa.rect.height)
        yeni[-1].insert_image(sayfa.rect, stream=gorsel)
    doc.close()
    yeni.save(os.path.join(TEST_KLASORU, ad))
    yeni.close()
    print("olusturuldu:", ad)


def excel_fatura_listesi():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["Fatura No", "Tarih", "VKN", "Ünvan", "Matrah", "KDV", "Toplam"])
    ws.append(["GFE202400000001", "05.02.2024", "12345678901", "ABC TİCARET", 1000.00, 200.00, 1200.00])
    ws.append(["GFE202400000002", "10.02.2024", "98765432109", "XYZ İTHALAT", 2500.00, 500.00, 3000.00])
    ws.append(["GFE202400000003", "11.02.2024", "55544433322", "DENEME ELEKTRONİK", 500.00, 50.00, 550.00])
    ws.append(["GFE202400000004", "15.02.2024", "11122233344", "BAŞKA GIDA", 300.00, 30.00, 330.00])
    ws.append(["TOPLAM", "", "", "", 4300.00, 780.00, 5080.00])
    wb.save(os.path.join(TEST_KLASORU, "fatura_listesi.xlsx"))
    print("olusturuldu: fatura_listesi.xlsx")


def muavin_satis_xlsx():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["MUAVİN DEFTER"])
    ws.append(["DENEME TİCARET ORTAKLIĞI"])
    ws.append(["Dönem :", "01/01/2026-31/12/2026"])
    ws.append(["Tarih Aralığı :", "01/07/2026-31/07/2026"])
    ws.append(["600.00.001 SİGARA SATIŞ", "", "", "", "TL"])
    ws.append(["TARİH", "TİP", "FİŞ NO", "AÇIKLAMA", "BORÇ", "ALACAK", "BAKİYE", "B/A"])
    ws.append(["", "", "", "Nakli Yekün:", 1338242.0, 1338242.0])
    ws.append(["2026-07-01", "Mahsup", "001850", "01/07/2026-2126-Z RAPORU", "", 9080.00])
    ws.append(["2026-07-02", "Mahsup", "001851", "02/07/2026-2127-Z RAPORU", "", 15221.00])
    ws.append(["2026-07-08", "Mahsup", "001857", "08/07/2026-2133-Z RAPORU", "", 4455.00])
    ws.append(["Nakli Yekün Hariç :"])
    ws.append(["Genel Toplam :"])
    ws.append([""])
    ws.append(["600.01.002 1 Lİ TİCARİ MALLAR SATIŞI", "", "", "", "TL"])
    ws.append(["TARİH", "TİP", "FİŞ NO", "AÇIKLAMA", "BORÇ", "ALACAK", "BAKİYE", "B/A"])
    ws.append(["", "", "", "Nakli Yekün:", 530924.89, 530924.89])
    ws.append(["2026-07-01", "Mahsup", "001850", "01/07/2026-2126-Z RAPORU", "", 10623.76])
    ws.append(["2026-07-02", "Mahsup", "001851", "02/07/2026-2127-Z RAPORU", "", 13909.90])
    ws.append(["Nakli Yekün Hariç :"])
    ws.append(["Genel Toplam :"])
    ws.append([""])
    ws.append(["600.20.020 20 Lİ TİCARİ MALLAR SATIŞI", "", "", "", "TL"])
    ws.append(["TARİH", "TİP", "FİŞ NO", "AÇIKLAMA", "BORÇ", "ALACAK", "BAKİYE", "B/A"])
    ws.append(["", "", "", "Nakli Yekün:", 743610.33, 743610.33])
    ws.append(["2026-07-02", "Mahsup", "001851", "02/07/2026-2127-Z RAPORU", "", 612.50])
    ws.append(["Nakli Yekün Hariç :"])
    ws.append(["Genel Toplam :"])
    wb.save(os.path.join(TEST_KLASORU, "muavin_satis.xlsx"))
    print("olusturuldu: muavin_satis.xlsx")


def excel_cetvel_listesi():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["Sıra No", "VKN", "Ünvan", "Belge No", "Tarih", "Matrah", "KDV"])
    ws.append([1, 12345678901, "ABC TİCARET LTD. ŞTİ.", "GFE202400000001", "05.02.2024", 1000.00, 200.00])
    ws.append([2, 98765432109, "XYZ İTHALAT A.Ş.", "GFE202400000002", "10.02.2024", 2500.00, 500.00])
    ws.append([3, 55544433322, "DENEME ELEKTRONİK", "GFE202400000003", "11.02.2024", 500.00, 45.00])
    ws.append([4, 99988877766, "FAZLA KAYIT SANAYİ", "GFE202400000009", "12.02.2024", 700.00, 140.00])
    ws.append(["GENEL TOPLAM", "", "", "", "", 4700.00, 885.00])
    wb.save(os.path.join(TEST_KLASORU, "kontrol_cetveli.xlsx"))
    print("olusturuldu: kontrol_cetveli.xlsx")


def fatura_xml(ad, belge_no, tarih, satici_vkn, alici_vkn, matrah, kdv, toplam, oran):
    icerik = f'''<?xml version="1.0" encoding="UTF-8"?>
<Invoice xmlns="urn:oasis:names:specification:ubl:schema:xsd:Invoice-2"
         xmlns:cac="urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2"
         xmlns:cbc="urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2">
  <cbc:CustomizationID>TR1.2</cbc:CustomizationID>
  <cbc:ProfileID>TEMELFATURA</cbc:ProfileID>
  <cbc:ID>{belge_no}</cbc:ID>
  <cbc:IssueDate>{tarih}</cbc:IssueDate>
  <cbc:InvoiceTypeCode>SATIS</cbc:InvoiceTypeCode>
  <cac:AccountingSupplierParty>
    <cac:Party>
      <cac:PartyIdentification><cbc:ID schemeID="VKN">{satici_vkn}</cbc:ID></cac:PartyIdentification>
      <cac:PartyName><cbc:Name>ÖRNEK SANAYİ VE TİCARET LTD. ŞTİ.</cbc:Name></cac:PartyName>
    </cac:Party>
  </cac:AccountingSupplierParty>
  <cac:AccountingCustomerParty>
    <cac:Party>
      <cac:PartyIdentification><cbc:ID schemeID="VKN">{alici_vkn}</cbc:ID></cac:PartyIdentification>
      <cac:PartyName><cbc:Name>MÜŞTERİ TİCARET A.Ş.</cbc:Name></cac:PartyName>
    </cac:Party>
  </cac:AccountingCustomerParty>
  <cac:TaxTotal>
    <cbc:TaxAmount currencyID="TRY">{kdv:.2f}</cbc:TaxAmount>
    <cac:TaxSubtotal>
      <cbc:TaxableAmount currencyID="TRY">{matrah:.2f}</cbc:TaxableAmount>
      <cbc:TaxAmount currencyID="TRY">{kdv:.2f}</cbc:TaxAmount>
      <cac:TaxCategory>
        <cbc:ID>K</cbc:ID>
        <cbc:Percent>{oran}</cbc:Percent>
        <cac:TaxScheme><cbc:Name>KDV</cbc:Name></cac:TaxScheme>
      </cac:TaxCategory>
    </cac:TaxSubtotal>
  </cac:TaxTotal>
  <cac:LegalMonetaryTotal>
    <cbc:LineExtensionAmount currencyID="TRY">{matrah:.2f}</cbc:LineExtensionAmount>
    <cbc:TaxExclusiveAmount currencyID="TRY">{matrah:.2f}</cbc:TaxExclusiveAmount>
    <cbc:TaxInclusiveAmount currencyID="TRY">{toplam:.2f}</cbc:TaxInclusiveAmount>
    <cbc:PayableAmount currencyID="TRY">{toplam:.2f}</cbc:PayableAmount>
  </cac:LegalMonetaryTotal>
</Invoice>
'''
    with open(os.path.join(TEST_KLASORU, ad), "w", encoding="utf-8") as f:
        f.write(icerik)
    print("olusturuldu:", ad)


def fatura_xml_gzip(ad, icerik):
    import gzip
    with gzip.open(os.path.join(TEST_KLASORU, ad), "wb") as f:
        f.write(icerik.encode("utf-8"))
    print("olusturuldu:", ad)


def mahsup_fis_pdf(ad, fisler):
    pdf = yeni_pdf()
    for i, fis in enumerate(fisler):
        if i > 0:
            pdf.add_page()
        pdf.set_font("Arial", "", 10)
        pdf.cell(0, 6, "DENEME TİCARET ORTAKLIĞI", ln=1)
        pdf.cell(0, 6, "Şirket", ln=1)
        pdf.cell(0, 6, "Dönem", ln=1)
        pdf.cell(0, 6, "2026", ln=1)
        pdf.cell(0, 6, ":", ln=1)
        pdf.cell(0, 6, ":", ln=1)
        pdf.set_font("Arial", "B", 12)
        pdf.cell(0, 6, "MAHSUP  FİŞİ", ln=1)
        pdf.set_font("Arial", "", 10)
        pdf.cell(0, 6, "Tarih", ln=1)
        pdf.cell(0, 6, ": " + fis["tarih"], ln=1)
        pdf.cell(0, 6, "Fiş No", ln=1)
        pdf.cell(0, 6, ": " + fis["fis_no"], ln=1)
        pdf.cell(0, 6, "Yevmiye Madde No", ln=1)
        pdf.cell(0, 6, ": 00000000", ln=1)
        for baslik in ("HESAP KODU", "HESAP ADI", "AÇIKLAMA", "BORÇ", "ALACAK"):
            pdf.cell(0, 6, baslik, ln=1)
        pdf.cell(0, 6, "Belge Düzenleme Nedeni", ln=1)
        pdf.cell(0, 6, ": " + fis["neden"], ln=1)
        for satir in fis["satirlar"]:
            for hucre in (satir["hesap_kodu"], satir["hesap_adi"], satir["aciklama"], tl(satir["borc"]), tl(satir["alacak"])):
                pdf.cell(0, 6, hucre, ln=1)
        pdf.cell(0, 6, "FİŞ TOPLAM :", ln=1)
        pdf.cell(0, 6, tl(fis["toplam"]), ln=1)
        pdf.cell(0, 6, tl(fis["toplam"]), ln=1)
    pdf.output(os.path.join(TEST_KLASORU, ad))
    print("olusturuldu:", ad)


def tl(sayi):
    return f"{sayi:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


if __name__ == "__main__":
    os.makedirs(TEST_KLASORU, exist_ok=True)
    fatura_pdf("fatura_1_ok.pdf", "GFE202400000001", "05.02.2024", "12345678901", "99900011122", 1000.00, 200.00, 1200.00, 20)
    fatura_pdf("fatura_2_ok.pdf", "GFE202400000002", "10.02.2024", "98765432109", "99900011122", 2500.00, 500.00, 3000.00, 20)
    fatura_pdf("fatura_3_kdv_fark.pdf", "GFE202400000003", "11.02.2024", "55544433322", "99900011122", 500.00, 50.00, 550.00, 10)
    fatura_pdf("fatura_4_cetvelde_yok.pdf", "GFE202400000004", "15.02.2024", "11122233344", "99900011122", 300.00, 30.00, 330.00, 10)
    fatura_pdf("fatura_5_vkn_fark.pdf", "GFE202400000010", "13.02.2024", "00099988877", "99900011122", 800.00, 160.00, 960.00, 20)

    cetvel_pdf("kontrol_cetveli.pdf", [
        {"vkn": "12345678901", "unvan": "ABC TİCARET LTD. ŞTİ.", "belge_no": "GFE202400000001", "tarih": "05.02.2024", "matrah": tl(1000.00), "kdv": tl(200.00)},
        {"vkn": "98765432109", "unvan": "XYZ İTHALAT A.Ş.", "belge_no": "GFE202400000002", "tarih": "10.02.2024", "matrah": tl(2500.00), "kdv": tl(500.00)},
        {"vkn": "55544433322", "unvan": "DENEME ELEKTRONİK", "belge_no": "GFE202400000003", "tarih": "11.02.2024", "matrah": tl(500.00), "kdv": tl(45.00)},
        {"vkn": "99988877766", "unvan": "FAZLA KAYIT SANAYİ", "belge_no": "GFE202400000009", "tarih": "12.02.2024", "matrah": tl(700.00), "kdv": tl(140.00)},
        {"vkn": "12345678901", "unvan": "ABC TİCARET LTD. ŞTİ.", "belge_no": "GFE202400000001", "tarih": "05.02.2024", "matrah": tl(1000.00), "kdv": tl(200.00)},
        {"vkn": "11122233344", "unvan": "BAŞKA GIDA LTD. ŞTİ.", "belge_no": "GFE202400000010", "tarih": "13.02.2024", "matrah": tl(800.00), "kdv": tl(160.00)},
    ])

    toplu_fatura_pdf("Toplu Fatura Yazdırma.pdf", [
        ("GFE202400000011", "05.02.2024", "12345678901", "99900011122", 1000.00, 200.00, 1200.00, 20),
        ("GFE202400000012", "10.02.2024", "98765432109", "99900011122", 2500.00, 500.00, 3000.00, 20),
    ])

    taranmis_pdf("taranmis_cetvel.pdf", "kontrol_cetveli.pdf")

    fatura_xml("fatura_1.xml", "GFE202400000001", "2024-02-05", "12345678901", "99900011122", 1000.00, 200.00, 1200.00, 20)
    fatura_xml("fatura_2.xml", "GFE202400000002", "2024-02-10", "98765432109", "99900011122", 2500.00, 500.00, 3000.00, 20)
    fatura_xml("fatura_3.xml", "GFE202400000003", "2024-02-11", "55544433322", "99900011122", 500.00, 50.00, 550.00, 10)
    fatura_xml("fatura_4.xml", "GFE202400000004", "2024-02-15", "11122233344", "99900011122", 300.00, 30.00, 330.00, 10)
    fatura_xml("fatura_5.xml", "GFE202400000010", "2024-02-13", "00099988877", "99900011122", 800.00, 160.00, 960.00, 20)

    import io
    gz_icerik = io.StringIO()
    fatura_xml_gzip("sikistirilmis_fatura.xml",
                    f'''<?xml version="1.0" encoding="UTF-8"?>
<Invoice xmlns="urn:oasis:names:specification:ubl:schema:xsd:Invoice-2"
         xmlns:cac="urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2"
         xmlns:cbc="urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2">
  <cbc:ID>GFE202400000099</cbc:ID>
  <cbc:IssueDate>2024-02-20</cbc:IssueDate>
  <cac:AccountingSupplierParty>
    <cac:Party><cac:PartyIdentification><cbc:ID>12345678901</cbc:ID></cac:PartyIdentification></cac:Party>
  </cac:AccountingSupplierParty>
  <cac:AccountingCustomerParty>
    <cac:Party><cac:PartyIdentification><cbc:ID>99900011122</cbc:ID></cac:PartyIdentification></cac:Party>
  </cac:AccountingCustomerParty>
  <cac:TaxTotal><cbc:TaxAmount>90.00</cbc:TaxAmount></cac:TaxTotal>
  <cac:LegalMonetaryTotal>
    <cbc:TaxExclusiveAmount>450.00</cbc:TaxExclusiveAmount>
    <cbc:TaxInclusiveAmount>540.00</cbc:TaxInclusiveAmount>
  </cac:LegalMonetaryTotal>
</Invoice>
''')

    excel_fatura_listesi()
    excel_cetvel_listesi()
    muavin_satis_xlsx()

    mahsup_fis_pdf("fis_listesi_ornek.pdf", [
        {
            "tarih": "02/07/2026", "fis_no": "001851", "neden": "Z RAPORU",
            "satirlar": [
                {"hesap_kodu": "120.01.037", "hesap_adi": "PRATİK İŞLEM ÖDEME", "aciklama": "02/07/2026 2127 Z RAPORU", "borc": 640.00, "alacak": 0.00},
                {"hesap_kodu": "108.01.001", "hesap_adi": "DİĞER HAZIR DEĞERLER", "aciklama": "02/07/2026 2127 Z RAPORU", "borc": 14150.11, "alacak": 0.00},
                {"hesap_kodu": "391.01.001", "hesap_adi": "1Lİ HESAPLANAN KDV", "aciklama": "02/07/2026 2127 Z RAPORU", "borc": 0.00, "alacak": 145.21},
                {"hesap_kodu": "391.01.020", "hesap_adi": "20Lİ HESAPLANAN KDV", "aciklama": "02/07/2026 2127 Z RAPORU", "borc": 0.00, "alacak": 122.50},
                {"hesap_kodu": "600.01.002", "hesap_adi": "1 Lİ TİCARİ MALLAR SATIŞI", "aciklama": "02/07/2026 2127 Z RAPORU", "borc": 0.00, "alacak": 13909.90},
                {"hesap_kodu": "600.20.020", "hesap_adi": "20 Lİ TİCARİ MALLAR SATIŞI", "aciklama": "02/07/2026 2127 Z RAPORU", "borc": 0.00, "alacak": 612.50},
            ],
            "toplam": 14790.11,
        },
        {
            "tarih": "16/07/2026", "fis_no": "002063", "neden": "ÖRNEK MÜŞTERİ A.Ş.",
            "satirlar": [
                {"hesap_kodu": "600.01.002", "hesap_adi": "1 Lİ TİCARİ MALLAR SATIŞI", "aciklama": "16/07/2026 EAR2026000000001 ÖRNEK MÜŞTERİ A.Ş.", "borc": 0.00, "alacak": 3940.59},
                {"hesap_kodu": "600.10.001", "hesap_adi": "10 LU TİCARİ MALLAR SATIŞI", "aciklama": "16/07/2026 EAR2026000000001 ÖRNEK MÜŞTERİ A.Ş.", "borc": 0.00, "alacak": 2590.92},
                {"hesap_kodu": "600.20.020", "hesap_adi": "20 Lİ TİCARİ MALLAR SATIŞI", "aciklama": "16/07/2026 EAR2026000000001 ÖRNEK MÜŞTERİ A.Ş.", "borc": 0.00, "alacak": 3661.63},
                {"hesap_kodu": "391.01.001", "hesap_adi": "1Lİ HESAPLANAN KDV", "aciklama": "16/07/2026 EAR2026000000001 ÖRNEK MÜŞTERİ A.Ş.", "borc": 0.00, "alacak": 39.41},
                {"hesap_kodu": "391.01.010", "hesap_adi": "10LU HESAPLANAN KDV", "aciklama": "16/07/2026 EAR2026000000001 ÖRNEK MÜŞTERİ A.Ş.", "borc": 0.00, "alacak": 259.09},
                {"hesap_kodu": "391.01.020", "hesap_adi": "20Lİ HESAPLANAN KDV", "aciklama": "16/07/2026 EAR2026000000001 ÖRNEK MÜŞTERİ A.Ş.", "borc": 0.00, "alacak": 732.33},
                {"hesap_kodu": "100.01.001", "hesap_adi": "KASA", "aciklama": "16/07/2026 EAR2026000000001 ÖRNEK MÜŞTERİ A.Ş.", "borc": 11223.97, "alacak": 0.00},
            ],
            "toplam": 11223.97,
        },
        {
            "tarih": "30/07/2026", "fis_no": "002110", "neden": "ÖRNEK TEDARİK LTD.ŞTİ.",
            "satirlar": [
                {"hesap_kodu": "153.01.001", "hesap_adi": "1 Lİ EKMEK ALIŞLARI", "aciklama": "30/07/2026 EFA2026000000002 ÖRNEK TEDARİK LTD.ŞTİ.", "borc": 2970.30, "alacak": 0.00},
                {"hesap_kodu": "191.01.001", "hesap_adi": "1Lİ İNDİRİLECEK KDV", "aciklama": "30/07/2026 EFA2026000000002 ÖRNEK TEDARİK LTD.ŞTİ.", "borc": 29.70, "alacak": 0.00},
                {"hesap_kodu": "100.01.001", "hesap_adi": "KASA", "aciklama": "30/07/2026 EFA2026000000002 ÖRNEK TEDARİK LTD.ŞTİ.", "borc": 0.00, "alacak": 3000.00},
            ],
            "toplam": 3000.00,
        },
    ])

    print("Test verileri hazır:", TEST_KLASORU)
