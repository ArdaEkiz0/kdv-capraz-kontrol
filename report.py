import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from matcher import (DURUM_CETVELDE_YOK, DURUM_FATURADA_YOK, DURUM_MUKERRER,
                     DURUM_OK, DURUM_PARSE_SORUNU, DURUM_TUTAR_FARKI,
                     DURUM_VKN_FARKI, SORUNLU_DURUMLAR)
from ozetler import ba_formu, kdv_dagilim_fatura, kdv_dagilim_muavin
from utils import tl_format

BASLIK_FONT = Font(bold=True, color="FFFFFF", size=11)
BASLIK_DOLGU = PatternFill("solid", fgColor="4472C4")
OK_DOLGU = PatternFill("solid", fgColor="C6EFCE")
SORUN_DOLGU = PatternFill("solid", fgColor="FFC7CE")
UYARI_DOLGU = PatternFill("solid", fgColor="FFEB9C")
TOPLAM_DOLGU = PatternFill("solid", fgColor="D9E1F2")
INCE_KENAR = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

DURUM_RENK = {
    DURUM_OK: OK_DOLGU,
    DURUM_TUTAR_FARKI: SORUN_DOLGU,
    DURUM_VKN_FARKI: UYARI_DOLGU,
    DURUM_MUKERRER: UYARI_DOLGU,
    DURUM_CETVELDE_YOK: SORUN_DOLGU,
    DURUM_FATURADA_YOK: SORUN_DOLGU,
    DURUM_PARSE_SORUNU: SORUN_DOLGU,
}

DURUM_ADLARI = {
    DURUM_OK: "Eşleşti",
    DURUM_TUTAR_FARKI: "Tutar Farkı",
    DURUM_VKN_FARKI: "VKN Farkı",
    DURUM_CETVELDE_YOK: "Muavinde Yok",
    DURUM_FATURADA_YOK: "Faturalarda Yok",
    DURUM_MUKERRER: "Mükerrer",
    DURUM_PARSE_SORUNU: "Okunamadı",
}

# Durum görünen adı -> renk (iade gibi harici durumlar için güvenli eşleme)
DURUM_AD_RENK = {
    DURUM_ADLARI[k]: renk
    for k, renk in DURUM_RENK.items()
    if k in DURUM_ADLARI
}


def tablo_yaz(sayfa, basliklar, satirlar, genislikler=None, renk_kurali=None, sayi_kolonlari=None):
    for j, b in enumerate(basliklar, start=1):
        hucre = sayfa.cell(row=1, column=j, value=b)
        hucre.font = BASLIK_FONT
        hucre.fill = BASLIK_DOLGU
        hucre.alignment = Alignment(horizontal="center", vertical="center")
        hucre.border = INCE_KENAR
    for i, satir in enumerate(satirlar, start=2):
        for j, deger in enumerate(satir, start=1):
            hucre = sayfa.cell(row=i, column=j, value=deger)
            hucre.border = INCE_KENAR
            if sayi_kolonlari and j in sayi_kolonlari:
                hucre.number_format = "#,##0.00"
                hucre.alignment = Alignment(horizontal="right")
            if renk_kurali:
                dolgu = renk_kurali(satir)
                if dolgu:
                    hucre.fill = dolgu
    if genislikler:
        for j, g in enumerate(genislikler, start=1):
            sayfa.column_dimensions[get_column_letter(j)].width = g
    sayfa.freeze_panes = "A2"


def _tutar(deger):
    return deger if deger is not None else ""


def _oranlar_str(oranlar):
    if not oranlar:
        return ""
    return ", ".join(f"%{o}%" for o in oranlar)


def _tarih_araligi(faturalar, cetvel_kayitlari):
    tarihler = [f["tarih"] for f in faturalar if f.get("tarih")]
    tarihler += [c["tarih"] for c in cetvel_kayitlari if c.get("tarih")]
    if not tarihler:
        return "-"
    return f"{min(tarihler)} - {max(tarihler)}"


OZET_VARSAYILANLARI = {
    "fatura_adet": 0, "cetvel_adet": 0, "eslesen": 0, "tutar_farki": 0,
    "vkn_farki": 0, "kdv_sifir": 0, "cetvelde_yok": 0, "faturada_yok": 0,
    "mukerrer": 0, "parse_sorunu": 0, "tevkifatli": 0, "indirimli": 0,
    "fark_toplami": 0,
}


def _ozet_tamamla(ozet):
    """Eksik özet anahtarlarını varsayılanla doldurur (kısmi özetlerde çökmesin)."""
    tamam = dict(OZET_VARSAYILANLARI)
    if ozet:
        tamam.update({k: v for k, v in ozet.items() if v is not None})
    return tamam


def rapor_olustur(sonuc_satirlari, ozet, faturalar, cetvel_kayitlari, hedef_yol, gecmis_bilgi=None):
    ozet = _ozet_tamamla(ozet)
    wb = Workbook()

    def durum_toplamlar():
        toplamlar = {}
        for d in DURUM_ADLARI:
            toplamlar[d] = {"adet": 0, "matrah": 0, "kdv": 0}
        for r in sonuc_satirlari:
            d = r["durum"]
            toplamlar.setdefault(d, {"adet": 0, "matrah": 0, "kdv": 0})
            t = toplamlar[d]
            t["adet"] += 1
            if r["matrah"] is not None:
                t["matrah"] += r["matrah"]
            if r["kdv"] is not None:
                t["kdv"] += r["kdv"]
        return toplamlar

    # ---------- 1) Özet ----------
    ws = wb.active
    ws.title = "Ozet"
    toplamlar = durum_toplamlar()
    genel_kdv = sum((f["kdv"] or 0) for f in faturalar)
    eksik_kdv = toplamlar[DURUM_CETVELDE_YOK]["kdv"]
    fazla_kdv = toplamlar[DURUM_FATURADA_YOK]["kdv"]
    fark_kdv = toplamlar[DURUM_TUTAR_FARKI]["kdv"]
    durum_kdv = sum((r["kdv"] or 0) for r in sonuc_satirlari if r["durum"] == DURUM_OK)

    baslik = ws.cell(row=1, column=1, value="KDV ÇAPRAZ KONTROL RAPORU")
    baslik.font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Üretim zamanı: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    ws.cell(row=3, column=1, value=f"Belge dönemi: {_tarih_araligi(faturalar, cetvel_kayitlari)}")

    ozet_metinler = [
        ("", ""),
        ("KAPSAM", ""),
        ("XML/PDF Fatura Sayısı", ozet["fatura_adet"]),
        ("Muavin Kayıt Sayısı", ozet["cetvel_adet"]),
        ("", ""),
        ("EŞLEŞME SONUÇLARI", ""),
        ("Eşleşen (tutar uyumlu)", ozet["eslesen"]),
        ("Tutar Farkı Olan", ozet["tutar_farki"]),
        ("VKN Farkı Olan", ozet["vkn_farki"]),
        ("Muavinde Kaydı Olmayan Fatura", ozet["cetvelde_yok"]),
        ("Faturalarda Olmayan Muavin Kaydı", ozet["faturada_yok"]),
        ("Mükerrer Kayıt", ozet["mukerrer"]),
        ("Okunamayan Fatura", ozet["parse_sorunu"]),
        ("", ""),
        ("KDV TUTARLARI (TL)", ""),
        ("Tüm faturaların toplam KDV'si", genel_kdv),
        ("Eşleşen faturaların toplam KDV'si", durum_kdv),
        ("Muavinde olmayan faturaların KDV'si (EKSİK)", eksik_kdv),
        ("XML'de olmayan muavin KDV'si (FAZLA)", fazla_kdv),
        ("Tutar farkı olan kayıtların KDV'si", fark_kdv),
        ("", ""),
        ("SATICI BAZINDA MUAVİNDE OLMAYAN KDV (EKSİK)", ""),
    ]
    satir_ozet = len(ozet_metinler) + 2
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 18

    # Satıcı bazlı eksik özeti
    eksik_satici = {}
    for r in sonuc_satirlari:
        if r["durum"] == DURUM_CETVELDE_YOK and r["kdv"] is not None and r["kdv"]:
            anahtar = (r["vkn"], r["unvan"])
            g = eksik_satici.setdefault(anahtar, {"adet": 0, "kdv": 0})
            g["adet"] += 1
            g["kdv"] += r["kdv"]
    for (vkn, unvan), g in sorted(eksik_satici.items(), key=lambda x: -x[1]["kdv"]):
        if not vkn and not unvan:
            continue
        ad = unvan or vkn
        ozet_metinler.append((f"  {ad}  (VKN: {vkn or '-'})", f"{g['adet']} fatura / {tl_format(g['kdv'])} KDV"))

    for i, (a, v) in enumerate(ozet_metinler, start=1):
        ws.cell(row=i, column=1, value=a)
        if v != "":
            h = ws.cell(row=i, column=2, value=v)
            h.border = INCE_KENAR
            if a.startswith("  "):
                h.fill = UYARI_DOLGU
            elif a.startswith(("EŞLEŞME", "KDV TUTARLARI", "SATICI", "KAPSAM")):
                ws.cell(row=i, column=1).font = Font(bold=True)
                h.fill = TOPLAM_DOLGU
            else:
                if isinstance(v, int):
                    h.number_format = "#,##0.00"
                    h.alignment = Alignment(horizontal="right")
                    h.fill = OK_DOLGU if v == 0 else SORUN_DOLGU
    ws.cell(row=satir_ozet - 1, column=1).border = INCE_KENAR

    # Geçmiş kontrol karşılaştırması
    if gecmis_bilgi:
        baslangic = len(ozet_metinler) + 2
        ws.cell(row=baslangic + 1, column=1, value="SON KONTROLE GÖRE DEĞİŞİM").font = Font(bold=True)
        ozet_metinler.append(("SON KONTROLE GÖRE DEĞİŞİM", ""))
        if gecmis_bilgi.get("kapanan"):
            ozet_metinler.append(("  Bu ay muavine eklenen belgeler (çözüldü)", str(len(gecmis_bilgi["kapanan"])) + " adet"))
        if gecmis_bilgi.get("yeni"):
            ozet_metinler.append(("  Bu ay yeni ortaya çıkan eksikler", str(len(gecmis_bilgi["yeni"])) + " adet"))
        if gecmis_bilgi.get("onceki_eslesen") is not None:
            ozet_metinler.append(("  Önceki eşleşen adedi", gecmis_bilgi["onceki_eslesen"]))
        if gecmis_bilgi.get("zaman"):
            ozet_metinler.append(("  Önceki kontrol zamanı", gecmis_bilgi["zaman"]))
        for i, (a, v) in enumerate(ozet_metinler, start=1):
            ws.cell(row=i, column=1, value=a)
            if v != "":
                h = ws.cell(row=i, column=2, value=v)
                h.border = INCE_KENAR
                if a.startswith("  "):
                    h.fill = OK_DOLGU if "çözüldü" in a else UYARI_DOLGU
                elif a == "SON KONTROLE GÖRE DEĞİŞİM":
                    h.fill = TOPLAM_DOLGU

    # ---------- 2) Sonuçlar ----------
    ws2 = wb.create_sheet("Sonuclar")
    basliklar = ["Durum", "Belge No", "VKN", "Satıcı/Ünvan", "Tarih", "Tip", "Matrah", "KDV", "Toplam", "Oranlar", "Oran Kontrol", "Kaynak", "Detay"]
    satirlar = []
    for r in sonuc_satirlari:
        satirlar.append([
            DURUM_ADLARI.get(r["durum"], r["durum"]),
            r["belge_no"] or "",
            r["vkn"] or "",
            r["unvan"] or "",
            r["tarih"] or "",
            r.get("tip") or "",
            _tutar(r["matrah"]),
            _tutar(r["kdv"]),
            _tutar(r.get("toplam")),
            _oranlar_str(r.get("oranlar")),
            r.get("oran_kontrol") or "",
            r["kaynak"] or "",
            r["detay"] or "",
        ])
    tablo_yaz(
        ws2, basliklar, satirlar,
        genislikler=[18, 22, 13, 30, 12, 14, 14, 14, 14, 10, 14, 10, 60],
        renk_kurali=lambda s: DURUM_AD_RENK.get(s[0]),
        sayi_kolonlari={7, 8, 9},
    )

    # ---------- 3) Muavinde Olmayan Faturalar (EKSİK) ----------
    ws3 = wb.create_sheet("EksikFaturalar")
    basliklar = ["Tarih", "Belge No", "Tip", "Satıcı VKN", "Satıcı Ünvanı", "Matrah", "KDV", "Toplam", "Oranlar", "Detay"]
    satirlar = []
    for r in sonuc_satirlari:
        if r["durum"] != DURUM_CETVELDE_YOK:
            continue
        satirlar.append([
            r["tarih"] or "", r["belge_no"] or "", r.get("tip") or "",
            r["vkn"] or "", r["unvan"] or "",
            _tutar(r["matrah"]), _tutar(r["kdv"]), _tutar(r.get("toplam")),
            _oranlar_str(r.get("oranlar")), r["detay"] or "",
        ])
    toplam_matrah = sum((r["matrah"] or 0) for r in sonuc_satirlari if r["durum"] == DURUM_CETVELDE_YOK)
    toplam_kdv = sum((r["kdv"] or 0) for r in sonuc_satirlari if r["durum"] == DURUM_CETVELDE_YOK)
    if satirlar:
        satirlar.append(["", "TOPLAM", "", "", "", toplam_matrah, toplam_kdv, "", "", ""])
    tablo_yaz(
        ws3, basliklar, satirlar,
        genislikler=[12, 22, 14, 13, 30, 14, 14, 14, 10, 60],
        sayi_kolonlari={6, 7, 8},
    )
    if satirlar:
        for j in range(1, len(basliklar) + 1):
            ws3.cell(row=len(satirlar) + 1, column=j).fill = TOPLAM_DOLGU

    # ---------- 4) Faturalarda Olmayan Muavin Kayıtları (FAZLA) ----------
    ws4 = wb.create_sheet("FazlaMuavin")
    basliklar = ["Tarih", "Belge No", "Ünvan", "KDV", "Notlar"]
    satirlar = []
    for r in sonuc_satirlari:
        if r["durum"] != DURUM_FATURADA_YOK:
            continue
        satirlar.append([
            r["tarih"] or "", r["belge_no"] or "", r["unvan"] or "",
            _tutar(r["kdv"]), r["detay"] or "",
        ])
    toplam_kdv = sum((r["kdv"] or 0) for r in sonuc_satirlari if r["durum"] == DURUM_FATURADA_YOK)
    if satirlar:
        satirlar.append(["", "TOPLAM", "", toplam_kdv, ""])
    tablo_yaz(
        ws4, basliklar, satirlar,
        genislikler=[12, 22, 40, 14, 50],
        sayi_kolonlari={4},
    )
    if satirlar:
        for j in range(1, len(basliklar) + 1):
            ws4.cell(row=len(satirlar) + 1, column=j).fill = TOPLAM_DOLGU

    # ---------- 5) Tutar/VKN Farkları ----------
    ws5 = wb.create_sheet("TutarFarklari")
    basliklar = ["Durum", "Belge No", "VKN", "Ünvan", "Tarih", "Matrah", "KDV", "Detay"]
    satirlar = []
    for r in sonuc_satirlari:
        if r["durum"] not in (DURUM_TUTAR_FARKI, DURUM_VKN_FARKI, DURUM_MUKERRER):
            continue
        satirlar.append([
            DURUM_ADLARI.get(r["durum"], r["durum"]),
            r["belge_no"] or "", r["vkn"] or "", r["unvan"] or "",
            r["tarih"] or "", _tutar(r["matrah"]), _tutar(r["kdv"]), r["detay"] or "",
        ])
    tablo_yaz(
        ws5, basliklar, satirlar,
        genislikler=[14, 22, 13, 30, 12, 14, 14, 60],
        sayi_kolonlari={6, 7},
    )

    # ---------- 6) Satıcı Özeti (fatura + kontrol sonucu kırılımı) ----------
    satici_ozet_sayfasi_ekle(wb, sonuc_satirlari, faturalar)

    # ---------- 7) Faturalar (tümü) ----------
    ws7 = wb.create_sheet("Faturalar")
    satirlar = []
    for f in faturalar:
        konum = f.get("satir", 1) if f.get("tip") == "excel" else f.get("sayfa", 1)
        satirlar.append([
            os.path.basename(f["dosya"]),
            konum,
            f["belge_no"] or "",
            f["tarih"] or "",
            f.get("satici_unvan") or "",
            f.get("satici_vkn") or "",
            f.get("alici_vkn") or "",
            f.get("fatura_tipi") or "",
            _tutar(f.get("matrah")),
            _tutar(f.get("kdv")),
            _tutar(f.get("toplam")),
            _oranlar_str(f.get("oranlar")),
            " / ".join(f["notlar"]) if f.get("notlar") else "",
        ])
    tablo_yaz(
        ws7, ["Dosya", "Sayfa/Satır", "Belge No", "Tarih", "Satıcı Ünvanı", "Satıcı VKN", "Alıcı VKN", "Tip", "Matrah", "KDV", "Toplam", "Oranlar", "Notlar"],
        satirlar,
        genislikler=[30, 10, 22, 12, 30, 14, 14, 14, 14, 14, 14, 10, 50],
        sayi_kolonlari={9, 10, 11},
    )

    # ---------- 8) Kontrol Cetveli (muavin) ----------
    ws8 = wb.create_sheet("KontrolCetveli")
    satirlar = []
    for c in cetvel_kayitlari:
        satirlar.append([
            c["vkn"] or "",
            c["belge_no"] or "",
            c["tarih"] or "",
            _tutar(c.get("matrah")),
            _tutar(c.get("kdv")),
            c.get("unvan") or "",
            " / ".join(c["notlar"]) if c.get("notlar") else "",
        ])
    tablo_yaz(
        ws8, ["VKN", "Belge No", "Tarih", "Matrah", "KDV", "Ünvan", "Notlar"],
        satirlar,
        genislikler=[14, 22, 12, 14, 14, 40, 40],
        sayi_kolonlari={4, 5},
    )

    # ---------- 9) KDV Dağılımı ----------
    ws9 = wb.create_sheet("KDVDagilimi")
    muavin_dagilim = kdv_dagilim_muavin(cetvel_kayitlari)
    fatura_dagilim = kdv_dagilim_fatura(faturalar)

    def _dagilim_tablosu(sayfa, basla, baslik, basliklar, satirlar, genislikler, sayi_kolonlari):
        sayfa.cell(row=basla, column=1, value=baslik).font = Font(bold=True, size=12)
        basla += 1
        for j, b in enumerate(basliklar, start=1):
            hucre = sayfa.cell(row=basla, column=j, value=b)
            hucre.font = BASLIK_FONT
            hucre.fill = BASLIK_DOLGU
            hucre.alignment = Alignment(horizontal="center", vertical="center")
            hucre.border = INCE_KENAR
        for i, satir in enumerate(satirlar, start=basla + 1):
            for j, deger in enumerate(satir, start=1):
                hucre = sayfa.cell(row=i, column=j, value=deger)
                hucre.border = INCE_KENAR
                if j in sayi_kolonlari:
                    hucre.number_format = "#,##0.00"
                    hucre.alignment = Alignment(horizontal="right")
                if isinstance(deger, str) and deger == "TOPLAM":
                    hucre.font = Font(bold=True)
                    hucre.fill = TOPLAM_DOLGU
        for j, g in enumerate(genislikler, start=1):
            sayfa.column_dimensions[get_column_letter(j)].width = g
        return basla + len(satirlar) + 2

    # Muavin 191 dağılımı
    satirlar = []
    toplam_muavin = 0
    for hesap, g in sorted(muavin_dagilim.items(), key=lambda x: -x[1]["kdv"]):
        satirlar.append([hesap, g["adet"], g["kdv"]])
        toplam_muavin += g["kdv"]
    if satirlar:
        satirlar.append(["TOPLAM", sum(x[1] for x in satirlar), toplam_muavin])
    satir = _dagilim_tablosu(
        ws9, 1, "MUAVİN 191 HESAP KDV DAĞILIMI",
        ["Hesap", "Kayıt Adedi", "KDV Toplamı"], satirlar,
        [40, 12, 16], {2, 3},
    )

    # XML/PDF fatura KDV dağılımı
    satirlar = []
    toplam_fatura = 0
    for oran, g in sorted(fatura_dagilim.items(), key=lambda x: x[0] if isinstance(x[0], int) else 999):
        oran_ad = f"%{oran}%" if isinstance(oran, int) else (oran or "Bilinmiyor")
        satirlar.append([oran_ad, g["adet"], g["matrah"], g["kdv"]])
        toplam_fatura += g["kdv"]
    if satirlar:
        satirlar.append(["TOPLAM", sum(x[1] for x in satirlar), sum(x[2] for x in satirlar), toplam_fatura])
    _dagilim_tablosu(
        ws9, satir, "XML/PDF FATURA KDV DAĞILIMI",
        ["Oran", "Fatura Adedi", "Matrah Toplamı", "KDV Toplamı"], satirlar,
        [40, 14, 16, 16], {2, 3, 4},
    )

    # ---------- 10) Ba Formu ----------
    ws10 = wb.create_sheet("BaFormu")
    satirlar = []
    for s in ba_formu(faturalar):
        satirlar.append([
            s["vkn"] or "", s["unvan"] or "", s["adet"], s["matrah"], s["kdv"],
        ])
    toplam_kdv = sum(x[4] for x in satirlar)
    toplam_matrah = sum(x[3] for x in satirlar)
    if satirlar:
        satirlar.append(["", "TOPLAM", sum(x[2] for x in satirlar), toplam_matrah, toplam_kdv])
    tablo_yaz(
        ws10, ["VKN", "Satıcı Ünvanı", "Fatura Adedi", "Matrah Toplamı", "KDV Toplamı"],
        satirlar,
        genislikler=[14, 40, 14, 16, 16],
        sayi_kolonlari={3, 4, 5},
    )
    if satirlar:
        for j in range(1, 6):
            ws10.cell(row=len(satirlar) + 1, column=j).fill = TOPLAM_DOLGU

    # ---------- 11) Grafik ----------
    ws11 = wb.create_sheet("Grafik")
    durum_satirlar = [["Durum", "Adet"]]
    for d in DURUM_ADLARI:
        durum_satirlar.append([DURUM_ADLARI[d], toplamlar[d]["adet"]])
    for i, satir in enumerate(durum_satirlar, start=1):
        for j, deger in enumerate(satir, start=1):
            hucre = ws11.cell(row=i, column=j, value=deger)
            if i == 1:
                hucre.font = BASLIK_FONT
                hucre.fill = BASLIK_DOLGU
    grafik1 = BarChart()
    grafik1.type = "col"
    grafik1.title = "Kontrol Sonuçları Dağılımı"
    veri1 = Reference(ws11, min_col=2, min_row=1, max_row=len(durum_satirlar))
    kategoriler1 = Reference(ws11, min_col=1, min_row=2, max_row=len(durum_satirlar))
    grafik1.add_data(veri1, titles_from_data=True)
    grafik1.set_categories(kategoriler1)
    grafik1.width = 24
    grafik1.height = 12
    ws11.add_chart(grafik1, "E2")

    eksik_satici_g = {}
    for r in sonuc_satirlari:
        if r["durum"] == DURUM_CETVELDE_YOK and r["kdv"] is not None and r["kdv"]:
            anahtar = (r["vkn"], r["unvan"])
            g = eksik_satici_g.setdefault(anahtar, {"adet": 0, "kdv": 0})
            g["adet"] += 1
            g["kdv"] += r["kdv"]
    top_8 = sorted(eksik_satici_g.items(), key=lambda x: -x[1]["kdv"])[:8]
    if top_8:
        grafik_basla = 1
        ws11.cell(row=grafik_basla, column=4, value="Satıcı").font = BASLIK_FONT
        ws11.cell(row=grafik_basla, column=4).fill = BASLIK_DOLGU
        ws11.cell(row=grafik_basla, column=5, value="Eksik KDV").font = BASLIK_FONT
        ws11.cell(row=grafik_basla, column=5).fill = BASLIK_DOLGU
        for i, ((vkn, unvan), g) in enumerate(top_8, start=2):
            ad = (unvan or vkn or "Bilinmeyen")[:45]
            ws11.cell(row=grafik_basla + i - 1, column=4, value=ad)
            ws11.cell(row=grafik_basla + i - 1, column=5, value=g["kdv"]).number_format = "#,##0.00"
        grafik2 = BarChart()
        grafik2.type = "bar"
        grafik2.title = "Muavinde Olmayan KDV - İlk 8 Satıcı"
        veri2 = Reference(ws11, min_col=5, min_row=1, max_row=len(top_8) + 1)
        kategoriler2 = Reference(ws11, min_col=4, min_row=2, max_row=len(top_8) + 1)
        grafik2.add_data(veri2, titles_from_data=True)
        grafik2.set_categories(kategoriler2)
        grafik2.width = 24
        grafik2.height = 12
        ws11.add_chart(grafik2, "E20")

    # ---------- 12) KDV oran kontrolü ----------
    oran_kontrol_sayfasi_ekle(wb, faturalar)

    # ---------- 13) Veri (düz tablo - pivot/filtre için) ----------
    veri = wb.create_sheet("Veri")
    veri_basliklar = ["Tarih", "Durum", "Kaynak", "VKN", "Unvan", "Belge No",
                      "Matrah", "KDV", "Toplam", "Oranlar", "Detay"]
    veri.append(veri_basliklar)
    for c in range(1, len(veri_basliklar) + 1):
        hucre = veri.cell(row=1, column=c)
        hucre.font = BASLIK_FONT
        hucre.fill = BASLIK_DOLGU
        hucre.border = INCE_KENAR
    for r in sonuc_satirlari:
        oranlar = r.get("oranlar") or []
        veri.append([
            r.get("tarih") or "",
            r.get("durum") or "",
            r.get("kaynak") or "",
            r.get("vkn") or "",
            r.get("unvan") or "",
            r.get("belge_no") or "",
            float(r["matrah"]) if r.get("matrah") is not None else None,
            float(r["kdv"]) if r.get("kdv") is not None else None,
            float(r["toplam"]) if r.get("toplam") is not None else None,
            ", ".join(str(o) for o in oranlar),
            r.get("detay") or "",
        ])
    for kolon, genislik in zip("ABCDEFGHIJK",
                               (11, 17, 13, 13, 32, 20, 12, 12, 12, 10, 70)):
        veri.column_dimensions[kolon].width = genislik
    for satir in veri.iter_rows(min_row=2):
        for h in (7, 8, 9):
            satir[h - 1].number_format = "#,##0.00"
        durum_hucre = satir[1]
        dolgu = DURUM_RENK.get(durum_hucre.value)
        if dolgu is not None:
            durum_hucre.fill = dolgu
    veri.freeze_panes = "A2"
    veri.auto_filter.ref = f"A1:K{veri.max_row}"

    wb.save(hedef_yol)
    return hedef_yol


# ============================================================================
# ORAN DOĞRULAMA SAYFASI
# ============================================================================

def oran_kontrol_sayfasi_ekle(wb, faturalar):
    """Matrah × oran = KDV doğrulaması raporu (tüm faturalar, uyumsuzlar önce)."""
    if "KDV Oran Kontrolü" in wb.sheetnames:
        return
    from oran_kontrol import oran_dogrula

    kontroller = []
    for f in faturalar:
        if not f.get("belge_no") or f.get("matrah") is None or f.get("kdv") is None:
            continue
        k = oran_dogrula(f)
        k["_f"] = f
        kontroller.append(k)
    if not kontroller:
        return

    # Belirsizler (fark None) ortada, uyumsuzlar en üstte
    kontroller.sort(key=lambda k: (k["uyumlu"], k["fark"] is None))

    ws = wb.create_sheet("KDV Oran Kontrolü")
    INCE = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    BASLIK = Font(bold=True, color="FFFFFF", size=11)
    DOLGU = PatternFill("solid", fgColor="4472C4")

    uyumlu_adet = sum(1 for k in kontroller if k["uyumlu"])
    belirsiz_adet = sum(1 for k in kontroller if not k["uyumlu"] and k["fark"] is None)
    uyumsuz_adet = len(kontroller) - uyumlu_adet - belirsiz_adet
    ws.cell(row=1, column=1, value=f"Toplam {len(kontroller)} fatura | "
            f"Uyumlu: {uyumlu_adet} | Uyumsuz: {uyumsuz_adet} | Belirsiz: {belirsiz_adet}").font = Font(bold=True, size=11)

    basliklar = ["Belge No", "VKN", "Tarih", "Ünvan", "Matrah", "KDV",
                 "Oranlar", "Beklenen KDV", "Fark", "Uyum", "Mesaj"]
    for j, b in enumerate(basliklar, 1):
        h = ws.cell(row=3, column=j, value=b)
        h.font = BASLIK
        h.fill = DOLGU
        h.border = INCE

    satir_no = 4
    for k in kontroller:
        f = k["_f"]
        if k["uyumlu"]:
            isaret, renk = "✓", PatternFill("solid", fgColor="C6EFCE")
        elif k["fark"] is None:
            isaret, renk = "?", PatternFill("solid", fgColor="FFEB9C")
        else:
            isaret, renk = "✗", PatternFill("solid", fgColor="FFC7CE")

        hucreler = [
            f.get("belge_no") or "",
            f.get("satici_vkn") or "",
            f.get("tarih") or "",
            f.get("unvan") or "",
            float(f.get("matrah") or 0),
            float(f.get("kdv") or 0),
            ", ".join(f"%{o}" for o in f.get("oranlar") or []),
            float(k["beklenen_kdv"]) if k["beklenen_kdv"] is not None else "",
            float(k["fark"]) if k["fark"] is not None else "",
            isaret,
            k["mesaj"],
        ]
        for j, deger in enumerate(hucreler, 1):
            c = ws.cell(row=satir_no, column=j, value=deger)
            c.border = INCE
            if j in (5, 6, 8, 9):
                c.number_format = "#,##0.00"
            if j == 10:
                c.fill = renk
                c.alignment = Alignment(horizontal="center")
        satir_no += 1

    for j, w in enumerate([20, 13, 11, 34, 14, 14, 12, 14, 12, 7, 40], 1):
        ws.column_dimensions[get_column_letter(j)].width = w


def satici_ozet_sayfasi_ekle(wb, sonuc_satirlari, faturalar):
    """Satıcı bazında özet: fatura toplamları + kontrol sonucu kırılımı.

    Sorunlu satıcılar (tutar farkı / muavinde yok / faturalarda yok) en üstte.
    """
    if "SaticiOzeti" in wb.sheetnames:
        return
    ws = wb.create_sheet("SaticiOzeti")

    # Fatura bazlı toplamlar
    gruplar = {}
    for f in faturalar:
        anahtar = (str(f.get("satici_vkn") or ""), str(f.get("satici_unvan") or ""))
        g = gruplar.setdefault(anahtar, {
            "adet": 0, "matrah": 0.0, "kdv": 0.0,
            DURUM_OK: 0, DURUM_TUTAR_FARKI: 0, DURUM_CETVELDE_YOK: 0,
            DURUM_FATURADA_YOK: 0, "diger": 0, "fark_toplam": 0.0,
        })
        g["adet"] += 1
        g["matrah"] += float(f.get("matrah") or 0)
        g["kdv"] += float(f.get("kdv") or 0)

    # Kontrol sonucu kırılımı
    for r in sonuc_satirlari:
        anahtar = (str(r.get("vkn") or ""), str(r.get("unvan") or ""))
        g = gruplar.get(anahtar)
        if g is None:
            g = gruplar[anahtar] = {
                "adet": 0, "matrah": 0.0, "kdv": 0.0,
                DURUM_OK: 0, DURUM_TUTAR_FARKI: 0, DURUM_CETVELDE_YOK: 0,
                DURUM_FATURADA_YOK: 0, "diger": 0, "fark_toplam": 0.0,
            }
        d = r.get("durum")
        if d in g and isinstance(g[d], int):
            g[d] += 1
        else:
            g["diger"] += 1
        if d in (DURUM_TUTAR_FARKI, DURUM_VKN_FARKI) and r.get("fark"):
            try:
                g["fark_toplam"] += abs(float(r["fark"]))
            except (TypeError, ValueError):
                pass

    def _sirala(item):
        g = item[1]
        sorun = g[DURUM_TUTAR_FARKI] + g[DURUM_CETVELDE_YOK] + g[DURUM_FATURADA_YOK]
        return (-sorun, -g["adet"])

    basliklar = ["VKN", "Satıcı Ünvanı", "Fatura Adedi", "Eşleşti",
                 "Tutar Farkı", "Muavinde Yok", "Faturalarda Yok", "Diğer",
                 "Toplam Matrah", "Toplam KDV", "Fark Tutarı"]
    tablo_yaz(
        ws, basliklar, [],
        genislikler=[14, 42, 11, 9, 11, 13, 14, 8, 15, 15, 13],
        sayi_kolonlari={9, 10, 11},
    )

    INCE = INCE_KENAR
    satir_no = 2
    for (vkn, unvan), g in sorted(gruplar.items(), key=_sirala):
        if not vkn and not unvan:
            continue
        sorun = g[DURUM_TUTAR_FARKI] + g[DURUM_CETVELDE_YOK] + g[DURUM_FATURADA_YOK]
        hucreler = [vkn, unvan, g["adet"], g[DURUM_OK], g[DURUM_TUTAR_FARKI],
                    g[DURUM_CETVELDE_YOK], g[DURUM_FATURADA_YOK], g["diger"],
                    g["matrah"], g["kdv"], g["fark_toplam"] or ""]
        for j, deger in enumerate(hucreler, 1):
            c = ws.cell(row=satir_no, column=j, value=deger)
            c.border = INCE
            if j in (9, 10, 11):
                c.number_format = "#,##0.00"
            if sorun:
                if j <= 2:
                    c.font = Font(bold=True, color="9C0006")
                if j == 2:
                    c.fill = SORUN_DOLGU
        satir_no += 1

    toplam_sorunlu = sum(
        min(1, g[DURUM_TUTAR_FARKI] + g[DURUM_CETVELDE_YOK] + g[DURUM_FATURADA_YOK])
        for g in gruplar.values())
    ws.cell(row=satir_no + 1, column=2,
            value=f"Toplam {len(gruplar)} satıcı, {toplam_sorunlu} tanesinde sorun").font = Font(bold=True)


# ============================================================================
# ORAN DOĞRULAMA SAYFASI
# ============================================================================
