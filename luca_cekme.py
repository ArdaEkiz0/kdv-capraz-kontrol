"""Luca (Türmob) web uygulamasından 191/391 muavin dökümü ve e-Belge
(e-Fatura / e-Arşiv, alış + satış) dosyalarının çekimi.

https://agiris.luca.com.tr/SSO/giris.erp ortak giriş sayfasına Üye Numarası +
Kullanıcı Adı + Parola ile girilir (sanal klavye zorunlu değildir). Ardından
uygulama içinden Muavin Defteri raporu bulunup seçilen dönem ve hesap kodları
(191 indirilecek KDV, 391 hesaplanan KDV) için Excel dökümü indirilir.

Not: Luca uygulamasının oturum sonrası ekranları müşteri yapılandırmasına göre
değişebilir; gezinme metin eşleştirmesiyle yapılır ve başarısızlıkta hata ayıkla-
ma için ekran görüntüsü %%TEMP%% altına kaydedilir.
"""
import hashlib
import html as html_cevir
import io
import json
import os
import re
import shutil
import tempfile
import time
import zipfile
from contextlib import contextmanager
from datetime import date, datetime
from functools import wraps
from pathlib import Path

# Optional deps (graceful fallback)
try:
    from tenacity import retry, stop_after_attempt, wait_exponential, retry_if_exception_type
except Exception:  # pragma: no cover
    def retry(*a, **k):
        def deco(f): return f
        return deco
    stop_after_attempt = wait_exponential = retry_if_exception_type = None

try:
    from lxml import etree as ET_LXML
except Exception:  # pragma: no cover
    ET_LXML = None

try:
    from babel.dates import format_date as babel_format_date
except Exception:  # pragma: no cover
    babel_format_date = None


def tr_tarih(tarih):
    """Tarihi TR locale (gg.aa.yyyy) formatinda dondurur; babel yoksa fallback."""
    if babel_format_date:
        try:
            return babel_format_date(tarih, "dd.MM.yyyy", locale="tr_TR")
        except Exception:
            pass
    return tarih.strftime("%d.%m.%Y")


# ==================== PREVENTIVE UTILITIES ====================

def _zip_dogrula(zip_yol, min_boyut=100):
    """ZIP dosyasının bozuk/boş olmadığını doğrular."""
    try:
        if not os.path.exists(zip_yol):
            return False, "dosya yok"
        boyut = os.path.getsize(zip_yol)
        if boyut < min_boyut:
            return False, f"boyut cok kucuk: {boyut} bayt"
        with zipfile.ZipFile(zip_yol, 'r') as zf:
            # En az bir XML/Excel dosyası olmalı
            uyeler = [u for u in zf.namelist() if _zip_uyesi_guvenli_mi(u)]
            if not uyeler:
                return False, "icerik yok (guvenli uye yok)"
            # CRC kontrolü
            corrupt = zf.testzip()
            if corrupt:
                return False, f"CRC hatali: {corrupt}"
        return True, "OK"
    except zipfile.BadZipFile:
        return False, "gecersiz ZIP"
    except Exception as e:
        return False, f"hata: {e}"


def _frame_saglikli_mi(cerceve):
    """Frame'in detached/crashed olmadığını kontrol eder."""
    try:
        # Basit bir evaluate ile frame canlı mı?
        cerceve.evaluate("() => document.readyState")
        return True
    except Exception:
        return False


def _sayfa_saglikli_mi(sayfa):
    """Sayfa/context'in canlı olduğunu kontrol eder."""
    try:
        sayfa.evaluate("() => document.readyState")
        return True
    except Exception:
        return False


@contextmanager
def _gecici_klasor(on_ek="luca_"):
    """Otomatik temizlenen geçici klasör (hata durumunda bile)."""
    tmp = Path(tempfile.mkdtemp(prefix=on_ek))
    try:
        yield tmp
    finally:
        try:
            shutil.rmtree(tmp, ignore_errors=True)
        except Exception:
            pass


def _dosya_kilidi_kontrol(yol, timeout=10):
    """Dosya başka bir süreçce kilitli mi? (Windows antivirus vb.)"""
    import time as _time
    basla = _time.time()
    while _time.time() - basla < timeout:
        try:
            with open(yol, 'r+b'):
                return False  # kilitli değil
        except (OSError, PermissionError):
            _time.sleep(0.5)
    return True  # hala kilitli


def _indirme_tamamlandi_mi(yol, onceki_boyut=-1, bekle=2):
    """Dosya boyutu belli süre değişmezse indirme bitmiş sayılır."""
    import time as _time
    for _ in range(3):
        try:
            boyut = os.path.getsize(yol)
            if boyut == onceki_boyut and boyut > 0:
                return True
            onceki_boyut = boyut
        except OSError:
            return False
        _time.sleep(bekle)
    return False


def _etag_hesapla(icerik):
    """İçerik hash'i (idempotency için)."""
    return hashlib.sha256(icerik).hexdigest()[:16]


def _log_yapilandir(log_klasor=None):
    """Structured JSON log dosyası (CI/CD için)."""
    import logging
    import logging.handlers
    if log_klasor is None:
        log_klasor = Path(os.environ.get("TEMP", ".")) / "luca_logs"
    log_klasor.mkdir(parents=True, exist_ok=True)
    log_dosya = log_klasor / f"luca_{datetime.now():%Y%m%d_%H%M%S}.jsonl"
    logger = logging.getLogger("luca")
    logger.setLevel(logging.DEBUG)
    handler = logging.FileHandler(log_dosya, encoding="utf-8")
    handler.setFormatter(logging.Formatter('%(message)s'))
    logger.addHandler(handler)
    return logger, log_dosya


# ZIP icinden asla cikarilmamasi gereken dosya/klasor adlari (kucuk harfe
# cevrilerek karsilastirilir). Beklenen icerik Excel/XML fatura dokumu
# oldugundan bunlarin cikmasi her zaman supheli bir durumdur.
_ZIP_YASAKLI_ADLAR = {
    "config.py", "ayar.json", "ayarlar.py", "gecmis.json", ".env",
    "__pycache__", ".git",
}


def _zip_uyesi_guvenli_mi(ic_ad):
    """Uye adinin cikarilmaya uygun olup olmadigini (dosya adi bazinda) kontrol eder."""
    taban = os.path.basename(ic_ad).lower()
    if not taban:
        return False
    if taban in _ZIP_YASAKLI_ADLAR:
        return False
    if taban.startswith("."):
        return False
    if taban.endswith((".py", ".pyc", ".pyo", ".dll", ".exe", ".bat", ".sh")):
        return False
    return True


def _guvenli_cikar(zipp, klasor):
    """ZIP icerigini zip-slip'e ve supheli dosya adlarina karsi denetleyerek cikarir.

    Uye adlari '..' icermez ve mutlak yol olamaz; surunen hedef daima
    klasor icinde kalir. Ayrica uygulamanin kendi yapilandirma/kod
    dosyalariyla ayni adi tasiyan veya calistirilabilir turden uyeler de
    atlanir (ornegin zip icine gizlenmis bir config.py). Guvensiz uye
    atlanir, sayisi dondurulur.
    """
    atlanan = 0
    kok = os.path.realpath(klasor)
    for ic_ad in zipp.namelist():
        if ic_ad.endswith("/"):
            continue
        if not _zip_uyesi_guvenli_mi(ic_ad):
            atlanan += 1
            continue
        hedef = os.path.realpath(os.path.join(klasor, ic_ad))
        if not (hedef == kok or hedef.startswith(kok + os.sep)):
            atlanan += 1
            continue
        zipp.extract(ic_ad, klasor)
    return atlanan


def _zipten_ozet(zip_yol, hedef_klasor=None):
    """ZIP icinden UBL XML okuyup matrah/KDV/toplam ozeti cikarir.

    hedef_klasor verilirse ZIP icerigi guvenli sekilde o klasore cikarilir.
    Zengin hata durumunda bos sozluk dondurur (cagri noktasi hatayi
    kendi bildirir).
    """
    o = {}
    try:
        with zipfile.ZipFile(zip_yol) as zipp:
            for ic_ad in zipp.namelist():
                icerik = zipp.read(ic_ad)
                if ic_ad.lower().endswith(".xml"):
                    o = _ubl_ozet(icerik)
            if hedef_klasor:
                _guvenli_cikar(zipp, hedef_klasor)
    except Exception:
        pass
    return o


LUCA_GIRIS_ADRESI = "https://agiris.luca.com.tr/LUCASSO/giris.erp"
LUCA_GIRIS_ADRESLERI = (
    "https://agiris.luca.com.tr/LUCASSO/giris.erp",
    "https://agiris.luca.com.tr/SSO/giris.erp",
)


class LucaHata(Exception):
    """Kullanıcıya gösterilebilir Luca çekim hatası."""


def _bildir_fonksiyonu(ilerleme):
    return ilerleme if ilerleme else (lambda metin: None)


_BRAVE_YOLU = (r"C:\Program Files\BraveSoftware\Brave-Browser"
               r"\Application\brave.exe")


def _tarayici_ac(playwright, gorunur=True):
    """Luca oturumu icin tarayici acar.

    Captcha kullanici tarafindan pencerede elle girildigi icin gorunur
    tarayici varsayilandir. Luca sunucusu otomasyon bayragini
    (--enable-automation / navigator.webdriver) tespit edip SSO'da 500
    dondurdugu icin bu bayraklar ozellikle devre disi birakilir.
    """
    ortak = {
        "headless": not gorunur,
        "ignore_default_args": ["--enable-automation"],
        "args": ["--disable-blink-features=AutomationControlled",
                 "--start-maximized"],
    }
    tarayici = None
    if os.path.exists(_BRAVE_YOLU):
        try:
            tarayici = playwright.chromium.launch(
                executable_path=_BRAVE_YOLU, **ortak)
        except Exception:
            tarayici = None
    if tarayici is None:
        for kanal in ("brave", "msedge", "chrome"):
            try:
                tarayici = playwright.chromium.launch(channel=kanal,
                                                      **ortak)
                break
            except Exception:
                continue
    if tarayici is None:
        try:
            tarayici = playwright.chromium.launch(**ortak)
        except Exception as hata:
            raise LucaHata(
                "Tarayıcı başlatılamadı. Edge veya Chrome kurulu olmalı "
                "ya da 'playwright install chromium' çalıştırılmalı.")
    try:
        tarayici._luca_gorunur = gorunur
    except Exception:
        pass
    return tarayici


def _luca_oturum_ac(tarayici, accept_downloads=True, storage_state=None):
    """Luca icin tarayici oturumu (context) acar.

    Headless modda UA'daki 'HeadlessChrome' ibaresi sunucuda SSO 500'e
    yol actigi icin temizlenir. Sabit viewport da SSO'yu bozdugu icin
    no_viewport kullanilir.

    storage_state: onceki oturumun cookie/localStorage'sini yüklemek için
    JSON dosya yolu (context.storage_state ile kaydedilmiş).
    """
    user_agent = None
    if not getattr(tarayici, "_luca_gorunur", False):
        try:
            olcum = tarayici.new_context()
            sayfa = olcum.new_page()
            sayfa.goto("about:blank", wait_until="domcontentloaded")
            user_agent = sayfa.evaluate("navigator.userAgent")
            olcum.close()
        except Exception:
            user_agent = None
        if user_agent and "Headless" in user_agent:
            user_agent = user_agent.replace("HeadlessChrome", "Chrome")
        else:
            user_agent = None
    kwargs = {"no_viewport": True}
    if accept_downloads:
        kwargs["accept_downloads"] = True
    if user_agent:
        kwargs["user_agent"] = user_agent
    if storage_state and os.path.exists(storage_state):
        kwargs["storage_state"] = storage_state
    return tarayici.new_context(**kwargs)


def _oturum_kaydet(oturum, dosya_yol):
    """Oturum cookie/localStorage'ını JSON olarak kaydet (session persistence)."""
    try:
        oturum.storage_state(path=dosya_yol)
        return True
    except Exception:
        return False


def _hata_ekrani_kaydet(sayfa, etiket):
    try:
        klasor = os.path.join(os.environ.get("TEMP", "."), "opencode")
        os.makedirs(klasor, exist_ok=True)
        yol = os.path.join(klasor, f"luca_{etiket}.png")
        sayfa.screenshot(path=yol, full_page=True)
        with open(os.path.join(klasor, f"luca_{etiket}_url.txt"), "w",
                  encoding="utf-8") as akis:
            akis.write(f"{sayfa.url}\n\n{sayfa.inner_text('body')[:2000]}")
        return yol
    except Exception:
        return ""


def _govde_metni(sayfa):
    try:
        return sayfa.inner_text("body")
    except Exception:
        return ""


def _giris_yapildi_mi(sayfa):
    adres = sayfa.url.lower()
    if "giris.erp" not in adres:
        return True
    return False


def _hata_iletisini_ayikla(metin):
    """Sayfadaki 'HATA ...' bloğundaki mesajı döndürür; yoksa '' döner."""
    satirlar = [s.strip() for s in metin.splitlines()]
    for sira, satir in enumerate(satirlar):
        if satir.upper() == "HATA" and sira + 1 < len(satirlar):
            return satirlar[sira + 1][:120]
    return ""


def _luca_captcha_src(sayfa):
    """Sayfadaki captcha görselinin kaynağını döndürür (değişim takibi)."""
    try:
        return sayfa.eval_on_selector("#captcha", "e => e.src") or ""
    except Exception:
        return ""


def _luca_swal_kapat(sayfa):
    """SweetAlert2 uyarı penceresi açıkssa onay düğmesiyle kapatır."""
    try:
        swal = sayfa.query_selector(".swal2-container")
        if swal is None or not swal.is_visible():
            return
        for secici in (".swal2-confirm", ".swal2-cancel", ".swal2-deny"):
            dugme = sayfa.query_selector(secici)
            if dugme:
                dugme.click(timeout=2000)
                sayfa.wait_for_timeout(600)
                return
        sayfa.keyboard.press("Escape")
        sayfa.wait_for_timeout(500)
    except Exception:
        pass


def _luca_captcha_as(sayfa, bildir, uye_no, kullanici, parola,
                     manuel_bekleme=180):
    """Luca captcha'sini kullanicinin elle girmesini bekler.

    Tarayici gorunur olmalidir; kullanici #captcha-input alanina kodu
    yazip Tamam'a basar. Yanlis girişte çıkan HATA penceresi otomatik
    kapatılır, kullanıcı aynı ekranda yeniden deneyebilir. Oturum
    sıfırlanıp giriş formu geri gelirse kimlikler yeniden gönderilir.
    Sure dolarsa False doner.
    """
    try:
        ilk = sayfa.query_selector("#captcha-input")
    except Exception:
        # Sayfa henuz yukleniyor/geziyor olabilir; bir kez daha bak.
        time.sleep(2)
        try:
            ilk = sayfa.query_selector("#captcha-input")
        except Exception:
            ilk = None
    if ilk is None:
        return True
    bekleme = manuel_bekleme if manuel_bekleme > 0 else 180
    try:
        sayfa.bring_to_front()
    except Exception:
        pass
    bildir(f"Captcha isteniyor: tarayıcı penceresindeki alana görüntüdeki "
           f"kodu elle girip Tamam'a basın ({bekleme} sn)...")
    for _ in range(bekleme):
        sayfa.wait_for_timeout(1000)
        try:
            captcha_kaldi = sayfa.query_selector("#captcha-input") is not None
        except Exception:
            # Kullanicinin girisiyle sayfa gezindi; yoklama bu esnada
            # olusebilir ("execution context destroyed"). Gezinti,
            # basarili giris anlamina gelir.
            captcha_kaldi = False
        if not captcha_kaldi:
            bildir("Captcha girildi; devam ediliyor.")
            return True
        _luca_swal_kapat(sayfa)
        # Oturum sıfırlanıp temiz giriş formu geldiyse tekrar gönder
        try:
            if sayfa.query_selector("#musteriNo") is not None:
                sayfa.fill("#musteriNo", str(uye_no))
                sayfa.fill("#kullaniciAdi", str(kullanici))
                sayfa.fill("#parola", str(parola))
                sayfa.click("input[type=button][value='GİRİŞ']", timeout=4000)
                sayfa.wait_for_timeout(1500)
        except Exception:
            pass
    bildir("Captcha süresi doldu.")
    return False


def giris_yap(sayfa, uye_no, kullanici, parola, bildir, manuel_bekleme=180):
    """Luca ortak giriş sayfasından oturum açar; sayfayı döndürür."""
    for i, adres in enumerate(LUCA_GIRIS_ADRESLERI):
        try:
            sayfa.goto(adres, wait_until="domcontentloaded")
            sayfa.wait_for_timeout(1500)
            if sayfa.query_selector("#musteriNo") is not None or \
                    _giris_yapildi_mi(sayfa):
                break
        except Exception:
            if i == len(LUCA_GIRIS_ADRESLERI) - 1:
                raise
            continue
    if sayfa.query_selector("#musteriNo") is None:
        if not _giris_yapildi_mi(sayfa):
            raise LucaHata(
                "Luca giriş sayfası açılamadı (beklenen alan bulunamadı). "
                f"Adres: {sayfa.url[:100]}")
        bildir("Mevcut Luca oturumu kullanılıyor.")
        return sayfa
    sayfa.fill("#musteriNo", str(uye_no))
    sayfa.fill("#kullaniciAdi", str(kullanici))
    sayfa.fill("#parola", str(parola))
    tiklandi = False
    for secici in ("input[type=button][value='GİRİŞ']",
                   "input[type=submit][value='GİRİŞ']", "text=GİRİŞ"):
        try:
            sayfa.click(secici, timeout=4000)
            tiklandi = True
            break
        except Exception:
            continue
    if not tiklandi:
        raise LucaHata("Luca giriş düğmesi bulunamadı.")
    sayfa.wait_for_timeout(1500)
    if sayfa.query_selector("#captcha-input") is not None:
        bildir("Captcha doğrulaması isteniyor...")
        if not _luca_captcha_as(sayfa, bildir, uye_no, kullanici, parola,
                                manuel_bekleme=manuel_bekleme):
            ekran = _hata_ekrani_kaydet(sayfa, "captcha")
            detay = f" Ekran görüntüsü: {ekran}" if ekran else ""
            raise LucaHata(
                "Captcha süresi içinde girilmedi. Tekrar deneyin ve "
                "tarayıcı penceresindeki captcha'yı zamanında girin." + detay)
    for _ in range(12):
        time.sleep(2)
        metin = _govde_metni(sayfa)
        hata_mesaji = _hata_iletisini_ayikla(metin)
        if hata_mesaji:
            try:
                sayfa.click("text=TAMAM", timeout=2000)
            except Exception:
                pass
            raise LucaHata("Luca girişi reddedildi: " + hata_mesaji)
        if "hatalı" in metin.lower() and "parola" in metin.lower():
            raise LucaHata(
                "Luca girişi reddedildi: Üye numarası, kullanıcı adı veya "
                "parola hatalı.")
        if _giris_yapildi_mi(sayfa):
            bildir("Luca'ya giriş başarılı.")
            return sayfa
    ekran = _hata_ekrani_kaydet(sayfa, "giris")
    detay = f" Ekran görüntüsü: {ekran}" if ekran else ""
    raise LucaHata(
        "Luca girişi doğrulanamadı (sayfa giriş ekranında kaldı). İki aşamalı "
        f"doğrulama isteniyor olabilir.{detay}")


def _aktif_sayfa(oturum, mevcut):
    """Yeni sekme/pop-up açıldıysa ona geçer; aksi halde mevcudu döndürür."""
    try:
        sayfalar = [s for s in oturum.pages if not s.is_closed()]
        if sayfalar and sayfalar[-1] is not mevcut:
            return sayfalar[-1]
    except Exception:
        pass
    return mevcut


def _menu_elemani_tikla(sayfa, desen, aciklama, bildir, zaman_asimi=6):
    """Metni desene uyan ilk görünür öğeye tıklar; başarıda True döner."""
    derleme = re.compile(desen, re.IGNORECASE)
    bitis = time.time() + zaman_asimi
    while time.time() < bitis:
        try:
            ogeler = sayfa.query_selector_all(
                "a, span, td, th, div[onclick], label, li, b")
            for oge in ogeler:
                try:
                    if not oge.is_visible():
                        continue
                    metin = (oge.inner_text() or "").strip()
                    if metin and derleme.search(metin) and len(metin) < 80:
                        bildir(f"{aciklama}: '{metin[:50]}' tıklanıyor...")
                        oge.click()
                        sayfa.wait_for_timeout(2000)
                        return True
                except Exception:
                    continue
        except Exception:
            pass
        time.sleep(1)
    return False


def _gorunur_esles(sayfa, secici):
    """Secicinin gorunur eslesmesini dondurur; yoksa None.

    Luca sayfalari quirks modunda oldugu icin id secicileri gizli
    buyuk harf kopyalara da carpiabilir (HESAP_KODU_ILK gibi); ilk
    eslesme gizli dugum olabilir ve yazim bosluga gider.
    """
    try:
        for oge in sayfa.query_selector_all(secici):
            try:
                if oge.is_visible():
                    return oge
            except Exception:
                continue
    except Exception:
        pass
    return None


def _luca_metin_gir(oge, metin):
    """Luca'nın maskeli alanlarına klavye vuruşuyla yazar; değeri
    özellikten geri okuyup döndürür.

    Bu alanların özel JS araçları (takvim, hesap planı) olduğu için
    click() yerine odak JS ile verilir (tık takvim penceresi açar);
    yazım sonrası Escape/Tab ile araç kapatılıp değer işlenir. Tarih
    aracı ayraç karakterini değiştirebildiği için karşılaştırma
    ayraç-bağımsız yapılır ('/' -> '.').
    """
    try:
        oge.evaluate("el => el.focus && el.focus()")
    except Exception:
        pass
    try:
        oge.evaluate("el => el.select && el.select()")
    except Exception:
        pass
    try:
        oge.type(metin, delay=45)
        oge.press("Enter")
    except Exception:
        try:
            oge.fill(metin, force=True)
        except Exception:
            pass
    try:
        return (oge.input_value() or "").strip().replace("/", ".")
    except Exception:
        return None


def _tarih_alanlarini_doldur(sayfa, bas_tarih, bit_tarih, bildir):
    """Rapor ekranındaki tarih alanlarını bulup doldurmaya çalışır.

    Birincil: gib530 ekranındaki #tarih1/#tarih2 (e-Belge tarih aralığı).
    İkincil: Muavin ekranındaki #tarih_ilk/#tarih_son.
    Üçüncül: Genel seçicilerle tarama.
    """
    # Hem dd.MM.yyyy hem dd/MM/yyyy dene (Luca ekranları farklı format bekleyebilir)
    bas_metin_v1 = tr_tarih(bas_tarih)       # dd.MM.yyyy
    bit_metin_v1 = tr_tarih(bit_tarih)
    bas_metin_v2 = bas_metin_v1.replace(".", "/")  # dd/MM/yyyy
    bit_metin_v2 = bit_metin_v1.replace(".", "/")
    doldurulan = 0
    # Birincil: gib530 e-Belge ekranındaki tarih alanları
    for secici, metin_v1, metin_v2 in (("#tarih1", bas_metin_v1, bas_metin_v2),
                                        ("#tarih2", bit_metin_v1, bit_metin_v2)):
        try:
            oge = sayfa.query_selector(secici)
            if oge is None or not oge.is_visible():
                continue
            # Önce v1 dene, olmazsa v2
            if _luca_metin_gir(oge, metin_v1) == metin_v1:
                doldurulan += 1
            elif _luca_metin_gir(oge, metin_v2) == metin_v2:
                doldurulan += 1
        except Exception:
            continue
    if doldurulan < 2:
        # İkincil: Muavin ekranındaki tarih alanları
        for secici, metin_v1, metin_v2 in (("#tarih_ilk", bas_metin_v1, bas_metin_v2),
                                            ("#tarih_son", bit_metin_v1, bit_metin_v2)):
            try:
                oge = sayfa.query_selector(secici)
                if oge is None or not oge.is_visible():
                    continue
                if _luca_metin_gir(oge, metin_v1) == metin_v1:
                    doldurulan += 1
                elif _luca_metin_gir(oge, metin_v2) == metin_v2:
                    doldurulan += 1
            except Exception:
                continue
    if doldurulan < 2:
        for secici in ("input[name*='Tarih' i]", "input[id*='Tarih' i]",
                       "input[name*='tarih']", "input[id*='tarih']"):
            try:
                ogeler = sayfa.query_selector_all(secici)
            except Exception:
                continue
            for oge in ogeler:
                try:
                    if not oge.is_visible():
                        continue
                    deger = (oge.input_value() or "").strip()
                    if deger and re.match(r"^\d{2}\.\d{2}\.\d{4}$", deger):
                        continue
                    hedef = bas_metin if doldurulan == 0 else bit_metin
                    if _luca_metin_gir(oge, hedef) == hedef:
                        doldurulan += 1
                    if doldurulan >= 2:
                        break
                except Exception:
                    continue
            if doldurulan >= 2:
                break
    if doldurulan < 2:
        bildir("UYARI: Tarih alanları otomatik doldurulamadı; sayfanın kendi "
               "varsayılan dönemi kullanılacak.")
    else:
        bildir(f"Tarih aralığı girildi: {bas_metin} - {bit_metin}")
    return doldurulan


def _indir_butonu_tikla(sayfa, desenler, zaman_asimi=8):
    """Desenlere uyan ilk görünür düğmeye tıklar; True/False döner."""
    derlemeler = [re.compile(d, re.IGNORECASE) for d in desenler]
    bitis = time.time() + zaman_asimi
    while time.time() < bitis:
        try:
            ogeler = sayfa.query_selector_all(
                "input[type=button], input[type=submit], button, a")
            for oge in ogeler:
                try:
                    if not oge.is_visible():
                        continue
                    metin = ((oge.get_attribute("value") or "")
                             + " " + (oge.inner_text() or "")).strip()
                    if any(d.search(metin) for d in derlemeler):
                        oge.click()
                        return True
                except Exception:
                    continue
        except Exception:
            pass
        time.sleep(1)
    return False


def _rapor_seceneklerini_duzelt(cerceve, bildir=None):
    """Muavin rapor formunda bakiyesiz hesaplari gizleyen filtreyi acar.

    'Bakiyesi ve Calismayan Hesaplar' select'i 'Bakiyesiz Hesaplari
    Gosterme' seciliyse, donem icinde hareketi olmayan bitis hesabi
    (orn. 192) dokumden tamamen elenir. Tumunu gosteren degere
    cevrilir; bulunamazsa sessiz gecilir.
    """
    try:
        for secenek in cerceve.query_selector_all("select option"):
            try:
                metin = (secenek.inner_text() or "").strip()
            except Exception:
                continue
            kucuk = metin.lower()
            # 'Tumunu Goster' / 'Tumu' iceren secenek: bakiyesizleri de
            # listeye alir. Turkce karakterler degisik yazilabildigi
            # icin esnek eslestirme kullanilir.
            if "t" + chr(252) + "m" in kucuk and ("goster" in kucuk
                                                  or "g" in kucuk):
                deger = secenek.get_attribute("value")
                secenek.evaluate(
                    "o => o.parentElement.value = arguments[0]", deger)
                try:
                    cerceve.select_option(
                        f"select:has(option[value='{deger}'])", value=deger)
                except Exception:
                    pass
                if bildir:
                    bildir(f"Rapor filtresi '{metin}' olarak ayarlandi.")
                return True
    except Exception:
        pass
    return False


def _hesap_alanlarini_doldur(sayfa, hesap_kodu, bildir=None):
    """Başlangıç ve Bitiş hesap kodu alanlarını aralık olarak doldurur.

    Luca aralığı dizgesel karşılaştırmayla uygular; bitişe aynı kodu
    yazmak (191 -> 191) alt hesapları (191.01.003 gibi) dışarıda
    bırakır. Bitiş kodu bir sonraki hesap olmalıdır (191 -> 192) ki
    tüm alt hesaplar aralığa dahil olsun. Hesap boyu alanlarına
    (hesap_boyu_ilk/son) kesinlikle dokunulmaz.

    Sayfada ayni kimligin gizli buyuk harf kopyalari da vardir
    (HESAP_KODU_ILK); quirks mod seciciyi gizli dugume kilitleyebilir.
    Bu yuzden yalniz GORUNUR dugume gercek tiklama + Ctrl+A + klavye
    ile yazilir ve deger ayni dugumden okunarak dogrulanir.
    """
    try:
        son_kod = str(int(hesap_kodu) + 1)
    except ValueError:
        son_kod = hesap_kodu
    klavye = sayfa.page.keyboard

    def _deger_uyuyor(okunan, beklenen):
        # Maskeli alan '192.' veya '192/...' biciminde dondurabilir;
        # ayraclar temizlenip on ek olarak karsilastirilir.
        if not okunan:
            return False
        temiz = okunan.replace("/", ".").rstrip(".")
        return temiz == beklenen or temiz.startswith(beklenen + ".")

    def _dugume_yaz(oge, metin):
        """Alana tikla, mevcut degeri tamamen sil, metni yaz ve ENTER'a bas.

        Luca maskeli alanlari yazilan degeri Enter ile isler; Tab
        yetmez. Deger ozellikten geri okunarak dogrulanir.
        """
        for deneme in range(3):
            try:
                oge.click(timeout=4000)
            except Exception:
                pass
            time.sleep(0.3)
            # Onceki hesaptan kalan degeri tumuyle temizle:
            klavye.press("Control+a")
            klavye.press("Delete")
            klavye.press("Backspace")
            time.sleep(0.2)
            klavye.type(metin, delay=80)
            time.sleep(0.2)
            # Luca bu alanda Enter bekler; onay olmadan deger islenmez.
            klavye.press("Enter")
            time.sleep(0.5)
            try:
                okunan = (oge.evaluate("el => el.value")
                          or "").strip().replace("/", ".")
                if _deger_uyuyor(okunan, metin):
                    return metin
                if deneme < 2:
                    bildir(f"UYARI: Alan '{metin}' yazımı doğrulanamadı "
                           f"(okunan: '{okunan}'), tekrar deneniyor.")
            except Exception as hata:
                if deneme < 2:
                    bildir(f"UYARI: Alan okunamadı ({str(hata)[:60]}), "
                           "tekrar deneniyor.")
        return None

    def _cift_yaz():
        ilk = _gorunur_esles(sayfa, "#hesap_kodu_ilk")
        if ilk is None:
            bildir("UYARI: Başlangıç hesap kodu alanı bulunamadı.")
            return False
        if _dugume_yaz(ilk, hesap_kodu) is None:
            return False
        son = _gorunur_esles(sayfa, "#hesap_kodu_son")
        if son is None:
            bildir("UYARI: Bitiş hesap kodu alanı bulunamadı; döküm yalnız "
                   f"{hesap_kodu} ile sınırlı kalabilir.")
            return False
        try:
            onceki = (son.evaluate("el => el.value") or "").strip()
        except Exception:
            onceki = ""
        if _dugume_yaz(son, son_kod) is None:
            bildir(f"UYARI: Bitiş hesap koduna '{son_kod}' yazılamadı "
                   f"(alan önce şunuydu: '{onceki}'); elle kontrol edin.")
            return False
        bildir(f"Hesap aralığı yazıldı: {hesap_kodu} -> {son_kod} "
               f"(bitiş alanının önceki değeri: {onceki or 'boş'})")
        return True

    # Bilinen kimlikler: muavin ekranındaki hesap kodu alanları
    try:
        if _gorunur_esles(sayfa, "#hesap_kodu_ilk") is not None:
            if _cift_yaz():
                return True
    except Exception:
        pass

    # Genel yol: bos kalan ilk iki 'Hesap' girdisini bul; hesap boyu
    # alanlarini atla, birine baslangic digerine bitis kodunu yaz.
    doldurulan = 0
    for secici in ("input[name*='Hesap' i]", "input[id*='Hesap' i]",
                   "input[name*='hesap']", "input[id*='hesap']"):
        try:
            ogeler = sayfa.query_selector_all(secici)
        except Exception:
            continue
        for oge in ogeler:
            try:
                kimlik = ((oge.get_attribute("name") or "")
                          + (oge.get_attribute("id") or "")).lower()
                if "boyu" in kimlik:
                    continue
                if oge.is_visible() and not (oge.input_value() or "").strip():
                    hedef = hesap_kodu if doldurulan == 0 else son_kod
                    _luca_metin_gir(oge, hedef)
                    # Luca düğüm değiştirdiğinde geri okuma yanıltır;
                    # yazim istisna vermediyse basarili say. Asil
                    # denetim indirilen dosyada yapilir.
                    doldurulan += 1
                    if doldurulan >= 2:
                        return True
            except Exception:
                continue
    return doldurulan > 0


def cek_muavin(uye_no, kullanici, parola, bas_tarih, bit_tarih, hedef_klasor,
               hesap_kodlari=("191", "391"), firma_adi="", ilerleme=None):
    """Luca'dan muavin dökümünü Excel olarak indirir.

    Dönen değer: indirilen dosya yolları listesi.
    """
    bildir = _bildir_fonksiyonu(ilerleme)
    os.makedirs(hedef_klasor, exist_ok=True)

    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        raise LucaHata("Playwright kurulu değil. Kurulum: pip install playwright")

    dosyalar = []
    with sync_playwright() as p:
        # Captcha kullanicinin elle girdigi icin gorunur tarayici sart;
        # varsayilan da gorunur ama cagirida acikca belirtiyoruz.
        tarayici = _tarayici_ac(p, gorunur=True)
        oturum = _luca_oturum_ac(tarayici)
        sayfa = oturum.new_page()
        try:
            sayfa = giris_yap(sayfa, uye_no, kullanici, parola, bildir)
            sayfa.wait_for_timeout(2500)

            # Menuler iframe icinde render edildigi icin metin
            # eslestirmesiyle gezinme guvenilir degil; ERP penceresine
            # gecip rapor ekranini dogrudan adresiyle yukluyoruz.
            erp = _erp_penceresi(oturum, sayfa, bildir)
            _firma_donem_sec(erp, firma_adi, bas_tarih, bildir)
            cerceve = _muavin_frame(erp, uye_no, bildir)
            if cerceve is None:
                ekran = _hata_ekrani_kaydet(erp, "muavin_menu")
                detay = f" Ekran görüntüsü: {ekran}" if ekran else ""
                raise LucaHata(
                    "Muavin rapor ekranı (raporMizanDetayHazirla) "
                    "yüklenemedi." + detay)

            for hesap in hesap_kodlari:
                donem_etiketi = bas_tarih.strftime("%Y%m")
                hedef = os.path.join(
                    hedef_klasor,
                    f"luca_muavin_{hesap}_{donem_etiketi}.xlsx")
                try:
                    bildir(f"Hesap {hesap} muavini sorgulanıyor...")
                    # Onceki rapordan kalmis bayat form yerine ekranı
                    # her hesap icin tazele (alanlar temiz baslar).
                    yeni = _muavin_frame(erp, uye_no, bildir)
                    if yeni is not None:
                        cerceve = yeni
                    _tarih_alanlarini_doldur(cerceve, bas_tarih, bit_tarih,
                                             bildir)
                    # Bakiyesiz hesaplari gizleyen filtre acilir; yoksa
                    # donem ici hareketi olmayan bitis hesabi (192/392)
                    # dokumden elenir.
                    _rapor_seceneklerini_duzelt(cerceve, bildir)
                    _hesap_alanlarini_doldur(cerceve, hesap, bildir)
                    # Rapor Türü seçimi: varsa Excel'i işaretle
                    excel_secildi = _indir_butonu_tikla(
                        cerceve, (r"^excel$",), zaman_asimi=2)
                    if excel_secildi:
                        bildir("Rapor türü Excel olarak seçildi.")
                    raporda_indi = False
                    try:
                        with erp.expect_download(timeout=25000) as indirme:
                            _indir_butonu_tikla(
                                cerceve,
                                (r"^rapor$", r"^liste$", r"listele",
                                 r"sorgula", r"getir"),
                                zaman_asimi=6)
                        dosya = indirme.value
                        dosya.save_as(hedef)
                        dosyalar.append(hedef)
                        raporda_indi = True
                        bildir(f"İndirildi: {os.path.basename(hedef)}")
                        _muavin_dosya_denetle(hedef, hesap, bildir)
                    except Exception:
                        pass
                    if not raporda_indi:
                        # Rapor ekranda açıldı; ayrı Excel/döküm düğmesi ara
                        try:
                            with erp.expect_download(
                                    timeout=20000) as indirme:
                                if not _indir_butonu_tikla(
                                        cerceve,
                                        (r"excel", r"\bxls\b",
                                         r"aktar", r"döküm\s*al")):
                                    raise RuntimeError(
                                        "Excel/döküm düğmesi bulunamadı")
                            dosya = indirme.value
                            dosya.save_as(hedef)
                            dosyalar.append(hedef)
                            raporda_indi = True
                            bildir(f"İndirildi: {os.path.basename(hedef)}")
                            _muavin_dosya_denetle(hedef, hesap, bildir)
                        except Exception as hata:
                            bildir(f"Hesap {hesap}: döküm alınamadı "
                                   f"({str(hata)[:70]}).")
                        if not raporda_indi:
                            _hata_ekrani_kaydet(erp, f"rapor_{hesap}")
                except LucaHata:
                    raise
                except Exception as hata:
                    bildir(f"Hesap {hesap} hatası: {str(hata)[:90]}")
        finally:
            tarayici.close()
    if not dosyalar:
        raise LucaHata(
            "Luca'dan muavin dökümü indirilemedi. Hesap ekranları müşteriye "
            "özgü olabilir; muavin dosyalarını elle seçerek devam edebilirsiniz.")
    return dosyalar


def _firma_adaylari(sayfa):
    """SMM uye girisi sonrasi ekrandaki firma secim adaylarini toplar.

    Tablo satirlari ve select opsiyonlarindan VKN/TCKN desenli kayitlari
    ayiklar; yapi musteriye gore degisebildigi icin best-effort calisir.
    """
    adaylar = []
    gorulen = set()
    for tr in sayfa.query_selector_all("table tr, [role='row']"):
        try:
            metin = " ".join((tr.inner_text() or "").split())
        except Exception:
            continue
        m = re.search(r"\b(\d{10,11})\b", metin)
        if m and len(metin) >= 8 and metin not in gorulen:
            gorulen.add(metin)
            adaylar.append({"vkn": m.group(1), "metin": metin[:140],
                            "tur": "satir"})
    for opt in sayfa.query_selector_all("select option"):
        try:
            metin = " ".join((opt.inner_text() or "").split())
        except Exception:
            continue
        m = re.search(r"\b(\d{10,11})\b", metin)
        if m and metin not in gorulen:
            gorulen.add(metin)
            adaylar.append({"vkn": m.group(1), "metin": metin[:140],
                            "tur": "opsiyon",
                            "deger": opt.get_attribute("value")})
    return adaylar


def firma_sec(sayfa, vkn, bildir=None):
    """VKN'a uyan firmayi secim ekraninda secer; bulunamazsa LucaHata."""
    bildir = bildir or (lambda s: None)
    adaylar = _firma_adaylari(sayfa)
    hedef = None
    for a in adaylar:
        if a["vkn"] == str(vkn).strip():
            hedef = a
            break
    if hedef is None:
        _hata_ekrani_kaydet(sayfa, "firma_bulunamadi")
        raise LucaHata(f"VKN {vkn} için firma seçiminde bulunamadı "
                       f"({len(adaylar)} aday görüldü).")
    try:
        if hedef["tur"] == "opsiyon":
            sayfa.select_option("select",
                                value=hedef.get("deger") or hedef["metin"])
        else:
            sayfa.click(f"text={hedef['metin'][:60]}", timeout=8000)
    except Exception:
        try:
            sayfa.click(f"text={hedef['vkn']}", timeout=8000)
        except Exception as hata:
            raise LucaHata(f"Firma seçilemedi: {str(hata)[:80]}")
    bildir(f"Firma seçildi: {hedef['metin'][:60]}")
    return hedef


def kesif_raporu(oturum, sayfa, klasor):
    """Aktif ekranin yapisini haritalar: linkler, dugmeler, iframe'ler,
    menu ogeleri + tam ekran goruntusu + HTML dokumu.

    Donen deger: bulunan oge ozetleri sozlugu.
    """
    os.makedirs(klasor, exist_ok=True)
    sayfa = _aktif_sayfa(oturum, sayfa)

    def _liste(secici, nitelik="textContent", sinir=200):
        cikti = []
        for e in sayfa.query_selector_all(secici)[:sinir]:
            try:
                t = " ".join((e.inner_text() or "").split())[:90]
                h = e.get_attribute("href") or ""
                if t or h:
                    cikti.append({"metin": t, "href": h[:160]})
            except Exception:
                pass
        return cikti

    rapor = {
        "url": sayfa.url,
        "baslik": sayfa.title(),
        "linkler": _liste("a"),
        "dugmeler": _liste("button, input[type='button'], "
                           "input[type='submit']"),
        "iframeler": [],
        "menuler": _liste("[class*='menu'] li, nav a, [role='menuitem']"),
    }
    for f in sayfa.query_selector_all("iframe"):
        try:
            rapor["iframeler"].append({
                "src": (f.get_attribute("src") or "")[:160],
                "ad": f.get_attribute("name") or ""})
        except Exception:
            pass
    try:
        rapor["html_uzunluk"] = len(sayfa.content())
        with open(os.path.join(klasor, "sayfa.html"), "w",
                  encoding="utf-8", errors="replace") as d:
            d.write(sayfa.content())
    except Exception:
        pass
    import json
    with open(os.path.join(klasor, "rapor.json"), "w", encoding="utf-8",
              errors="replace") as d:
        json.dump(rapor, d, ensure_ascii=False, indent=1)
    _hata_ekrani_kaydet(sayfa, "kesif")
    return rapor

# ---------------- e-Belge cekimi (e-Fatura / e-Arsiv, alis+satis) ----------------

LUCA_BELGE_KATEGORILERI = (
    ("efatura_alis", r"e\s*[-\s]?fatura",
     (r"gelen", r"al[ıi]nan", r"inbox", r"al[ıi][şs]")),
    ("efatura_satis", r"e\s*[-\s]?fatura",
     (r"giden", r"g[öo]nderilen", r"outbox", r"d[üu]zenlenen",
      r"sat[ıi][şs]")),
    ("earsiv_alis", r"e\s*[-\s]?ar[şs]iv",
     (r"gelen", r"al[ıi]nan", r"al[ıi][şs]")),
    ("earsiv_satis", r"e\s*[-\s]?ar[şs]iv",
     (r"giden", r"g[öo]nderilen", r"d[üu]zenlenen", r"sat[ıi][şs]")),
)

# Luca ERP'de gib530.do ekranlarinin 'tur' parametre karsiliklari.
LUCA_GIB_TURLER = {
    "efatura_alis": "gib_efatura_alis",
    "efatura_satis": "gib_efatura_satis",
    "earsiv_alis": "gib_ebelge_alis",
    "earsiv_satis": "gib_ebelge_satis",
}

# 4 grup için okunaklı Türkçe etiketler (takip ekranında kullanılır).
_KATEGORI_ETIKET = {
    "earsiv_alis": "e-Arşiv Alış",
    "earsiv_satis": "e-Arşiv Satış",
    "efatura_alis": "e-Fatura Alış",
    "efatura_satis": "e-Fatura Satış",
}


def _kategori_etiketi(kategori):
    return _KATEGORI_ETIKET.get(kategori, kategori)


def _turk_kucult(metin):
    """Turkce buyuk harfleri de kuculterek karsilastirma metni uretir.

    I/ı/i farki da esitlenir; boylece 'KIRAZLAR' ile 'KİRAZLAR' ortak
    ibareye iner.
    """
    return (metin.replace("İ", "i").replace("I", "i").replace("Ş", "ş")
            .replace("Ğ", "ğ").replace("Ü", "ü").replace("Ö", "ö")
            .replace("Ç", "ç")).replace("ı", "i").lower()


_SIRKET_EK_KELIMELERI = {
    "a.s.", "as", "a.s", "ltd", "sti", "şti", "s.t.i.", "tic", "san",
    "ve", "ticaret", "sanayi", "limited", "sirketi", "şirketi",
    "kurumsal", "co", "corp", "inc",
}


def _firma_anahtar(metin):
    """Unvandan sirket eklerini atarak anlamli kelime kumesi uretir.

    'KİRAZLAR FERMANTASYON SAN. VE TİC. LTD. ŞTİ.' -> {'kirazlar',
    'fermantasyon'}; Luca'daki kısa ad 'KIRAZLARLT' olsa da kök
    'kirazlar' ile eslesir.
    """
    temiz = _turk_kucult(metin)
    temiz = re.sub(r"[^\w\s]", " ", temiz)
    return {k for k in temiz.split()
            if len(k) >= 3 and k not in _SIRKET_EK_KELIMELERI}


def _firma_eslesen(secenekler, firma_adi):
    """Sirket combo seceneklerini unvanla eslestirir (esnek).

    1. Tam alt-dize (eski davranis, en kesin)
    2. Anahtar-kelime kesisimi: hedef unvanin anlamli kelimeleri,
       Luca'nin kisaltilmis adinda gecmeli; kestirme 'KIRAZLARLT'
       gibi adlar da yakalanir.
    3. Ilk anlamli kelimeyle baslayan ad ('kirazlar*' gibi).
    """
    ibare = _turk_kucult(firma_adi)
    tam = [s for s in secenekler if ibare in _turk_kucult(s["t"])]
    if tam:
        return tam
    anahtarlar = _firma_anahtar(firma_adi) or set()
    if not anahtarlar:
        return []
    skorlu = []
    # Her anlamli kelime sirayla denenir ('kirazlar' da 'fermantasyon'
    # da); kisaltilmis adlar ('KIRAZLARLT') boyle yakalanir, benzer
    # baslangicli ilgisiz firmalar ('AKIN'/'AKKEC') eslesmez.
    for anahtar in sorted(anahtarlar, key=len, reverse=True):
        secili = {s["v"] for _, s in skorlu}
        for s in secenekler:
            if s["v"] in secili:
                continue
            ad_kucuk = re.sub(r"[^\w\s]", " ", _turk_kucult(s["t"]))
            if re.search(r"\b" + re.escape(anahtar), ad_kucuk):
                skorlu.append((1, s))
        if len(skorlu) == 1:
            return [s for _, s in skorlu]
        if len(skorlu) > 1:
            break
    if len(skorlu) == 1:
        return [s for _, s in skorlu]
    if skorlu:
        # Birden fazla aday: sirali dondur; cagiran taraf hata verir.
        return [s for _, s in skorlu]
    ilk = sorted(anahtarlar)[0] if anahtarlar else ""
    if ilk:
        baslayan = [s for s in secenekler
                    if _turk_kucult(s["t"]).startswith(ilk)]
        if len(baslayan) == 1:
            return baslayan
    return []


def _erp_penceresi(oturum, sayfa, bildir):
    """Portalda gonder('formTarget') tetikler; acilan ERP sekmesini dondurur.

    Yeni sekme auygs.luca.com.tr SSO'sundan gecer; otomasyon bayraklari
    kapali degilse sunucu 500 dondurdugu icin _tarayici_ac zorunludur.
    """
    popuplar = []
    oturum.on("page", lambda y: popuplar.append(y))
    sayfa.evaluate("gonder('formTarget')")
    bildir("Mali Müşavir Paketi penceresi açılıyor...")
    for _ in range(30):
        time.sleep(2)
        if not popuplar:
            continue
        hedef = popuplar[0]
        try:
            if len(hedef.content()) > 3000 or len(hedef.frames) > 1:
                hedef.wait_for_timeout(4000)
                return hedef
        except Exception:
            continue
    raise LucaHata(
        "Luca ERP penceresi açılamadı. Sunucu SSO isteğini reddetti "
        "(oturum çakışması olabilir; birkaç dakika sonra deneyin).")


def _firma_donem_sec(erp, firma_adi, bas_tarih, bildir):
    """ERP ust cercevesinde sirket + donem secip Tamam'a basar."""
    ust = None
    for f in erp.frames:
        try:
            if f.query_selector("#SirketCombo"):
                ust = f
                break
        except Exception:
            continue
    if ust is None:
        raise LucaHata("Firma seçim alanı (SirketCombo) bulunamadı.")
    secenekler = ust.eval_on_selector_all(
        "#SirketCombo option",
        "os => os.map(o => ({v:o.value,t:o.textContent.trim()}))"
        ".filter(x => x.v && x.v !== '0')")
    if not secenekler:
        raise LucaHata("Firma listesi boş geldi.")
    if firma_adi:
        eslesen = _firma_eslesen(secenekler, firma_adi)
        if not eslesen:
            ornekler = ", ".join(s["t"] for s in secenekler[:6])
            raise LucaHata(f"Firma bulunamadı: {firma_adi} "
                           f"(örnek firmalar: {ornekler})")
        hedef = eslesen[0]
        if len(eslesen) > 1:
            adaylar = ", ".join(s["t"] for s in eslesen[:6])
            raise LucaHata(
                f"'{firma_adi}' ile {len(eslesen)} firma eşleşti "
                f"({adaylar}). Hangi firma olduğunu netleştirmek için "
                "mükellef kaydındaki Unvan alanını Luca'daki kısa ada "
                "göre düzenleyin (örn. 'KIRAZLARLT' yazın).")
    else:
        if len(secenekler) != 1:
            raise LucaHata(
                f"firma_adi belirtilmeli (hesapta {len(secenekler)} "
                "firma var).")
        hedef = secenekler[0]
    ust.select_option("#SirketCombo", hedef["v"])
    ust.wait_for_timeout(800)
    bildir(f"Luca firması seçildi: {hedef['t']}")
    donemler = ust.eval_on_selector_all(
        "#DonemCombo option",
        "os => os.map(o => ({v:o.value,t:o.textContent.trim()}))"
        ".filter(x => x.v && x.v !== '0')")
    yil = str(bas_tarih.year)
    donem = next((d for d in donemler if yil in d["t"]), None)
    if donem is None and donemler:
        donem = donemler[0]
    if donem is None:
        raise LucaHata("Bu firmaya ait dönem bulunamadı.")
    ust.select_option("#DonemCombo", donem["v"])
    ust.wait_for_timeout(500)
    ust.click("button:has-text('Tamam')")
    bildir(f"Firma/dönem seçildi: {hedef['t']} / {donem['t']}")
    time.sleep(8)
    return hedef["t"], donem["t"]


def _belge_arama_popup_kapat(cerceve, bas_tarih, bit_tarih, bildir=None):
    """'BELGE ARAMA' popup'ı açıksa tarih girip arama yapar veya kapatır.

    Luca'nın gib530 ekranında gonder('indir') bazı durumlarda belge
    arama popup'ı açar. Popup hem frame içinde hem de ana sayfada olabilir.
    """
    if bildir is None:
        bildir = lambda s: None
    sayfa = cerceve.page

    def _popup_acik_mi(ctx):
        try:
            return ctx.evaluate(
                "() => {"
                " const d = document.getElementById('arama-window-div');"
                " if (!d) return false;"
                " const s = window.getComputedStyle(d);"
                " if (s.display === 'none') return false;"
                " if (s.visibility === 'hidden') return false;"
                " if (s.opacity === '0') return false;"
                " return true; }")
        except Exception:
            return False

    # Önce frame'de bak, sonra ana sayfada
    acik = _popup_acik_mi(cerceve) or _popup_acik_mi(sayfa)
    if not acik:
        bildir("BELGE ARAMA popup'i AÇIK DEĞİL (varsayılan dönem kullanılıyor).")
        return
    bildir("BELGE ARAMA popup'i algilandi, tarih giriliyor...")
    bas_metin = tr_tarih(bas_tarih).replace(".", "/")
    bit_metin = tr_tarih(bit_tarih).replace(".", "/")
    # Tarih alanlarini doldur (frame + sayfa dene) - date picker handler'lari tetikle
    for secici, metin in (("#baslangic", bas_metin), ("#bitis", bit_metin)):
        dolduruldu = False
        for ctx in (cerceve, sayfa):
            try:
                alan = ctx.query_selector(secici)
                if alan is None or not alan.is_visible():
                    continue
                # Luca date picker: focus -> click -> value -> blur zinciri
                alan.evaluate("""(el, val) => {
                    el.focus();
                    if (typeof dateFocus === 'function') dateFocus(el);
                    if (typeof dateClick === 'function') dateClick(el);
                    el.value = val;
                    el.dispatchEvent(new Event('input', {bubbles: true}));
                    el.dispatchEvent(new Event('change', {bubbles: true}));
                    if (typeof dateBlur === 'function') dateBlur(el, '', '', true, 2026);
                }""", metin)
                bildir(f"  {secici} = {metin} dolduruldu ({'frame' if ctx is cerceve else 'page'})")
                dolduruldu = True
                break
            except Exception as e:
                bildir(f"  {secici} doldurma hatasi ({'frame' if ctx is cerceve else 'page'}): {e}")
        if not dolduruldu:
            bildir(f"  {secici} HİÇ BULUNAMADI")
    # 'Belge Ara' butonuna bas — popup kapanana kadar bekle
    for deneme in range(3):
        try:
            ara_btn = None
            for ctx in (cerceve, sayfa):
                ara_btn = ctx.query_selector("#faturalari-ara-btn")
                if ara_btn is not None:
                    break
            if ara_btn is not None:
                ara_btn.click()
                bildir(f"Belge Ara tiklandi (deneme {deneme+1}), sonuclar bekleniyor...")
                # Sonuclarin yuklenmesini bekle (popup kapanana kadar)
                for _ in range(20):  # max 20 saniye
                    time.sleep(1)
                    hala_acik = _popup_acik_mi(cerceve) or _popup_acik_mi(sayfa)
                    if not hala_acik:
                        bildir("BELGE ARAMA popup'i kapandi.")
                        return
                bildir("Popup hala acik, hide_window() ile kapatiliyor...")
                try:
                    cerceve.evaluate("hide_window()")
                    time.sleep(1)
                except Exception:
                    pass
                return
            else:
                bildir("  #faturalari-ara-btn BULUNAMADI")
        except Exception as e:
            bildir(f"  Belge Ara tiklama hatası: {e}")
    # Hiçbir buton calısmadıysa hide_window() ile kapat
    try:
        cerceve.evaluate("hide_window()")
        bildir("BELGE ARAMA popup'i hide_window() ile kapatildi.")
        time.sleep(1)
    except Exception:
        pass


@retry(
    stop=stop_after_attempt(3),
    wait=wait_exponential(multiplier=2, min=3, max=15),
    retry=retry_if_exception_type(Exception),
    reraise=True
)
def _gibten_getir(cerceve, bas_tarih, bit_tarih, bildir=None):
    """Luca e-belge ekranında 'GİB'ten Getir' adımını çalıştırır.

    akış:
    1. gonder('indir') çağrılır → Luca "BELGE ARAMA" popup'ı açabilir.
    2. Popup varsa: tarih alanları doldurulup 'Belge Ara'ya basılır
       ya da X ile kapatılıp araç çubuğundaki 'GİB'den Getir' ile devam
       edilir.
    3. Belge listesi dolana kadar beklenir.
    """
    if bildir is None:
        bildir = lambda s: None
    try:
        sayfa = cerceve.page
        # ADIM: hangi tarihse onu yaz. Belge ekranındaki tarih aralığı
        # alanlarına istenen dönem girilir (aynı muavin çekimindeki gibi).
        try:
            _tarih_alanlarini_doldur(cerceve, bas_tarih, bit_tarih, bildir)
        except Exception as th:
            bildir(f"Tarih alanları doldurulamadı ({str(th)[:50]}); "
                   "varsayılan kullanılacak.")
        buton = None

        # Birincil: Luca'nın güncel arabirimindeki "Belgeleri Getir" butonu
        for secici in ("#faturalari-getir-btn",
                       "button[onclick*=\"gonder('indir')\"]",
                       "button[onclick*='indir']"):
            try:
                oge = cerceve.query_selector(secici)
                if oge is not None:
                    buton = oge
                    bildir("Belgeleri Getir butonu bulundu (doğrudan seçim).")
                    break
            except Exception:
                continue

        # İkincil: Geniş aday toplama (eski sürümler için geriye dönük uyumluluk)
        if buton is None:
            adaylar = []
            for secici in ("button", "input", "a", "span", "div", "img",
                           "li", "td", "b", "i"):
                try:
                    ogeler = cerceve.query_selector_all(secici)
                except Exception:
                    continue
                for oge in ogeler:
                    try:
                        metin = ((oge.get_attribute("value") or "")
                                 + " " + (oge.inner_text() or "")
                                 + " " + (oge.get_attribute("onclick") or "")
                                 + " " + (oge.get_attribute("title") or "")
                                 + " " + (oge.get_attribute("alt") or "")
                                 + " " + (oge.get_attribute("src") or ""))
                        k = (metin or "").lower()
                        if ("getir" in k or "internette" in k
                                or ("gib" in k and ("indir" in k or "cek" in k))
                                or "taşı" in k or "tası" in k
                                or ("gib" in k and "yükle" in k)
                                or "belgeleri getir" in k):
                            if "onceki" not in k and "geri" not in k:
                                adaylar.append(oge)
                    except Exception:
                        continue

            for oge in adaylar:
                tag = ""
                try:
                    tag = (oge.evaluate("el => el.tagName") or "").lower()
                except Exception:
                    pass
                m = ""
                try:
                    m = ((oge.get_attribute("value") or "")
                         + " " + (oge.inner_text() or "")
                         + " " + (oge.get_attribute("title") or "")
                         + " " + (oge.get_attribute("onclick") or "")).lower()
                except Exception:
                    pass
                if tag in ("button", "input") or "getir" in m \
                        or "internette" in m or "belgeleri" in m:
                    buton = oge
                    break
            if buton is None and adaylar:
                buton = adaylar[0]

        if buton is None:
            bildir("GİB'ten getir butonu hiç bulunamadı; listeyi "
                   "sorguyla yenilemeyi deniyorum.")
            try:
                _sorgula_listele_butonu(cerceve, bildir)
            except Exception:
                pass
            return
        bildir("GİB'ten getir tıklanıyor (belgeler çekiliyor)...")
        # ÖNCE: Toolbar'daki "Belge Ara" butonunu tıkla (gonder('arama-window'))
        # bu BELGE ARAMA popup'ını açar; içine tarih girip arama yaparız.
        try:
            arama_btn = cerceve.query_selector("button[onclick*=\"gonder('arama-window')\"]")
            if arama_btn is not None:
                arama_btn.click()
                bildir("Toolbar 'Belge Ara' butonu tıklandı (arama-window açılıyor).")
                time.sleep(1.5)
            else:
                # Fallback: JS ile aç
                cerceve.evaluate("gonder('arama-window')")
                bildir("gonder('arama-window') JS ile çağrıldı.")
                time.sleep(1.5)
        except Exception as e:
            bildir(f"Belge Ara butonu tıklanamadı: {e}")

        # SONRA: Popup'ı doldur ve ara (zaten _belge_arama_popup_kapat yapar)
        # ama önce gonder('indir') de çağır (bazı ekranlar bunu ister).
        try:
            cerceve.evaluate("gonder('indir')")
            bildir("gonder('indir') JS ile çağrıldı.")
        except Exception:
            try:
                buton.scroll_into_view_if_needed()
            except Exception:
                pass
            try:
                buton.click()
            except Exception:
                try:
                    buton.evaluate("e => e.click()")
                except Exception:
                    pass
            bildir("Buton tıklama ile çağrıldı (JS fallback).")
        # Luca bazı ekranlarda 'BELGE ARAMA' popup'ı açar; bu popup
        # tarih aralığı filtresi içerir. Doldurup 'Belge Ara' ya basar.
        time.sleep(2)
        _belge_arama_popup_kapat(cerceve, bas_tarih, bit_tarih, bildir)
        # Ek kontrol: popup 2sn'de açılmamışsa biraz daha bekle ve tekrar dene
        try:
            time.sleep(3)
            _belge_arama_popup_kapat(cerceve, bas_tarih, bit_tarih, bildir)
        except Exception:
            pass
        # İlgili onay/uyarı penceresi çıkabilir (Evet/Tamam/liste).
        for _ in range(3):
            try:
                onay = cerceve.query_selector(
                    "button:has-text('Evet'), button:has-text('Tamam'), "
                    "button:has-text('OK'), input[value*='Evet'], "
                    "input[value*='Tamam']")
                if onay is not None and onay.is_visible():
                    try:
                        onay.click()
                        bildir("Onay penceresi kapatıldı.")
                        time.sleep(1)
                    except Exception:
                        pass
                    break
            except Exception:
                pass
            # SweetAlert / custom popup
            try:
                cerceve.evaluate(
                    "document.querySelectorAll('.swal2-confirm, "
                    ".swal2-styled').forEach(b => b.click())")
            except Exception:
                pass
            time.sleep(0.5)
        # Belgeler listeye gelene kadar bekle (akıllı).
        # GİB sorguları 10-60 sn sürebilir; 60 denemeye kadar bekle.
        try:
            import luca_cekme as _l
            for _i in range(60):
                deneme_html = cerceve.content()
                belge_sayisi = len(_l._satirlari_ayikla(deneme_html))
                if belge_sayisi > 0:
                    bildir(f"Listede {belge_sayisi} belge göründü.")
                    break
                if _i % 5 == 4:
                    bildir(f"Hâlâ bekleniyor... ({_i+1}s)")
                time.sleep(1)
            else:
                bildir("60 sn'de belge gelmedi; mevcut liste kullanılıyor.")
        except Exception:
            time.sleep(3)
        bildir("GİB'ten getir tamamlandı; liste güncellendi.")

        # Güvence: getir sonrası 'Sorgula'/'Listele' butonuna basarak
        # listenin tazelenmesini zorla. Luca bazı ekranlarda getir ile
        # listeyi yenilemez; sorgu butonu gerekir.
        try:
            _sorgula_listele_butonu(cerceve, bildir)
        except Exception:
            pass
    except Exception as hata:
        bildir(f"GİB'ten getir başarısız: {str(hata)[:50]}")


def _sorgula_listele_butonu(cerceve, bildir=None):
    """'Sorgula' / 'Listele' / 'Getir' gibi listeyi yenileyen butona basar.

    GİB'ten getir sonrası listenin dolması için gerekli; bulunamazsa
    sessizce geçer.
    """
    if bildir is None:
        bildir = lambda s: None
    adaylar = []
    for secici in ("button", "input", "a", "span", "div"):
        try:
            ogeler = cerceve.query_selector_all(secici)
        except Exception:
            continue
        for oge in ogeler:
            try:
                metin = ((oge.get_attribute("value") or "")
                         + " " + (oge.inner_text() or "")
                         + " " + (oge.get_attribute("onclick") or "")).lower()
                if any(k in metin for k in ("sorgula", "listele", "liste",
                                            "getir", "ara", "tazele")):
                    adaylar.append(oge)
            except Exception:
                continue
    for oge in adaylar:
        try:
            oge.scroll_into_view_if_needed()
        except Exception:
            pass
        try:
            oge.click()
            if cerceve.page is not None:
                cerceve.page.wait_for_timeout(1300)
            return
        except Exception:
            try:
                oge.evaluate("e => e.click()")
                if cerceve.page is not None:
                    cerceve.page.wait_for_timeout(1300)
                return
            except Exception:
                continue
    bildir("Sorgula/Listele butonu bulunamadı; mevcut liste kullanılıyor.")


def _tani_kaydet(kategori, cerceve):
    """Çekim ekranının HTML + URL'sini %TEMP%\\luca_tani\\ altına yazar.

    Gerçek Luca yapısını çözüp çekimi kökten tasarlarken kullanılır;
    her kategori için ayrı dosya oluşturur.
    """
    import time as _t
    try:
        klasor = os.path.join(os.environ.get("TEMP", "."), "luca_tani")
        os.makedirs(klasor, exist_ok=True)
        damga = _t.strftime("%Y%m%d_%H%M%S")
        yol = os.path.join(klasor, f"{kategori}_{damga}.html")
        icerik = cerceve.content() if cerceve is not None else ""
        with open(yol, "w", encoding="utf-8", errors="replace") as f:
            f.write(icerik)
        try:
            url_yol = os.path.join(klasor, f"{kategori}_{damga}.url.txt")
            with open(url_yol, "w", encoding="utf-8") as f:
                f.write((cerceve.url or "") if cerceve is not None else "")
        except Exception:
            pass
    except Exception:
        pass


def _gib530_frame(erp, tur, uye_no, bildir=None):
    """Ana icerik cercevesini istenen gib530 ekranina goturur ve frame'i
    dondurur; yuklenmezse None doner.

    Firma secimi sonrasi cerceveler yeniden yuklendigi icin ilk
    denemede 'execution context destroyed' hatasi normaldir; gezinme
    araliklarla yeniden denenir.

    Yeni mükellefte 'E-Fatura Satış' turunun ana menüde takılmasını
    önlemek için: tur adresi yalnız frm3'e değil, bulunan her
    gib530 frame'ine uygulanır; ayrıca her denemede tüm frame'ler
    taranır ve yalnız istenen turdaki gib530 döndürülür (eski turdayken
    yanlış frame dönmesin diye URL içinde tur de denetlenir).

    Bazen tur geçişi frm3 yönlendirmesiyle olmaz (Luca'nın kimi
    ekranları farklı frame/sekme kullanır). Bu yüzden yönlendirme
    başarısız olursa, en-dış sayfada doğrudan gib530.do adresine
    gidilir ve yükleme beklenir.
    """
    adres = f"gib530.do?tur={tur}&c_musteri_id={uye_no}"
    # Yalnız frame-içi yönlendirme: ana pencereyi asla değiştirme (goto
    # ana ERP'yi bozup çekimi çökertebiliyor). frm3 yanında, kullanıcı
    # yapılandırmasına göre farklı olabilecek adları da dener.
    frame_adlari = ["frm3", "frm1", "frm2", "main", "icerik", "content",
                    "fatura"]
    for deneme in range(10):
        try:
            if deneme in (0, 2, 4, 6, 8):
                for fn in frame_adlari:
                    erp.evaluate(
                        "p => { const [fn, u] = p;"
                        " const f = top.frames[fn];"
                        " if (f) { f.location.href = u; }"
                        " else { const el = document.querySelector("
                        "   'iframe[name=\"' + fn + '\"],frame[name=\"' + fn + '\"]');"
                        "   if (el) el.src = u; } }",
                        [fn, adres])
        except Exception:
            pass
        time.sleep(1.2)
        for f in erp.frames:
            try:
                url = f.url or ""
                if "gib530" in url and tur in url and len(f.content()) > 5000:
                    return f
            except Exception:
                continue
    # Frm3 yönlendirmesiyle olmadıysa, TÜM çocuk frame'lere tur adresini
    # uygula (hangi frame gib530'u taşıyorsa o yüklensin).
    try:
        for f in erp.frames:
            try:
                if f == erp:
                    continue
                f.evaluate("u => { window.location.href = u; }",
                           adres if adres.startswith("http") else
                           "gib530.do?tur=" + tur + "&c_musteri_id=" +
                           str(uye_no))
                time.sleep(1.5)
                if "gib530" in (f.url or "") and tur in (f.url or ""):
                    if len(f.content()) > 5000:
                        return f
            except Exception:
                continue
    except Exception:
        pass
    if bildir is not None:
        try:
            nerede = erp.evaluate(
                "top.frames['frm3'] "
                "? top.frames['frm3'].location.href : 'frm3 yok'")
            bildir(f"frm3 durumu: {str(nerede)[:100]}")
        except Exception:
            pass
    return None


def _muavin_dosya_denetle(yol, hesap, bildir):
    """Indirilen muavin dosyasinda veri satiri ve dogru aralik var mi
    diye bakar.

    Ilk ~4 satir basliktir (MUAVIN DEFTER, firma, dönem, tarih); ilk
    hesap satirinin kodu istenen aralikta (hesap..hesap+1) olmalidir.
    Bitis kodu bir sonraki hesap oldugu icin 192 hesaplarinin da
    gorunmesi normaldir.
    """
    try:
        import openpyxl
        try:
            son_kod = str(int(hesap) + 1)
        except ValueError:
            son_kod = hesap
        wb = openpyxl.load_workbook(yol)
        ws = wb.active
        satir_sayisi = ws.max_row or 0
        ilk_hesap = None
        hareket_var = False
        ana_hesaplar = set()
        if satir_sayisi > 4:
            for satir in ws.iter_rows(min_row=5, values_only=True):
                ilk = str(satir[0] or "").strip() if satir else ""
                m = re.match(r"^(\d{3})(\.|$)", ilk)
                if m:
                    ana_hesaplar.add(m.group(1))
                    if not ilk_hesap:
                        ilk_hesap = ilk
                # Tarihli hareket satiri (dd.mm.yyyy / dd/mm/yyyy)
                if re.match(r"^\d{2}[./]\d{2}[./]\d{4}", ilk):
                    hareket_var = True
        wb.close()
        if satir_sayisi <= 4:
            bildir(f"UYARI: Hesap {hesap} dökümü boş görünüyor "
                   f"({satir_sayisi} satır).")
        elif son_kod not in ana_hesaplar and hesap in ana_hesaplar:
            bildir(f"UYARI: Hesap {hesap} dökümünde bitiş hesabı "
                   f"{son_kod} YOK (içindekiler: "
                   f"{', '.join(sorted(ana_hesaplar))}). Rapor filtresi "
                   "bakiyesiz hesapları eliyor olabilir; döküm aralığı "
                   "eksik.")
        elif ilk_hesap and not (hesap <= ilk_hesap[:3] <= son_kod):
            bildir(f"UYARI: Hesap {hesap} dökümü beklenen aralıkta değil "
                   f"(ilk satır: {ilk_hesap}).")
        elif ilk_hesap and not hareket_var:
            bildir(f"UYARI: Hesap {hesap} dökümünde DÖNEM İÇİ hareket yok "
                   "(yalnız 'Nakli Yekün' açılış satırları var). Belgeler "
                   "Luca'ya düşmüş ama henüz muhasebe fişine işlenmemiş "
                   "olabilir; bu yüzden faturalarla eşleşme çıkmaz.")
        elif ilk_hesap:
            bildir(f"Hesap {hesap} dökümü doğrulandı ({satir_sayisi} satır, "
                   f"ilk hesap: {ilk_hesap}).")
    except Exception:
        pass


def _frame_saglikli_mi(cerceve):
    """Frame'in detached/crashed olmadığını kontrol eder."""
    try:
        cerceve.evaluate("() => document.readyState")
        return True
    except Exception:
        return False


def _sayfa_saglikli_mi(sayfa):
    """Sayfa/context'in canlı olduğunu kontrol eder."""
    try:
        sayfa.evaluate("() => document.readyState")
        return True
    except Exception:
        return False


def _muavin_frame(erp, uye_no, bildir=None):
    """Icerik cercevesini muavin defter ekranina goturur ve frame'i
    dondurur; yuklenmezse None doner.

    Luca menuleri iframe icinde render edildigi icin metin
    eslestirmesiyle gezinme guvenilir degil; ERP menu dizisindeki
    'Tum Yazicilar -> Muavin Defter' adresi (raporMizanDetayHazirla.do)
    dogrudan yuklenir. Firma secimi sonrasi cerceveler yeniden
    yuklendigi icin gezinme araliklarla yeniden denenir.
    """
    adres = f"raporMizanDetayHazirla.do?c_musteri_id={uye_no}"
    for deneme in range(10):
        try:
            if deneme in (0, 3, 6):
                erp.evaluate(
                    "u => { top.frames['frm3'].location.href = u; }",
                    adres)
        except Exception:
            pass
        time.sleep(2)
        for f in erp.frames:
            try:
                if ("raporMizanDetayHazirla" in f.url
                        and len(f.content()) > 5000
                        and _frame_saglikli_mi(f)):
                    return f
            except Exception:
                continue
    if bildir is not None:
        try:
            nerede = erp.evaluate(
                "top.frames['frm3'] "
                "? top.frames['frm3'].location.href : 'frm3 yok'")
            bildir(f"frm3 durumu: {str(nerede)[:100]}")
        except Exception:
            pass
    return None


_FATURA_JSON_RE = re.compile(
    r"fatura=(?:\"([^\"]*)\"|'([^']*)')")


def _fatura_json_bul(html_metin):
    """HTML icindeki fatura="..." / fatura='...' attribute'larini dondurur."""
    for eslesme in _FATURA_JSON_RE.finditer(html_metin):
        # Hangi grup yakalandiysa onu kullan
        veri = eslesme.group(1) or eslesme.group(2) or ""
        if veri.strip():
            yield veri.strip()
_FATURA_TD = re.compile(r'<(?:td|th)[^>]*>(.*?)</(?:td|th)>',
                         re.DOTALL | re.IGNORECASE)
_FATURA_TR = re.compile(r'<tr[^>]*>(.*?)</tr>', re.DOTALL | re.IGNORECASE)


def _tutar_cevir(metin):
    """HTML hücresinden tutar değerini float'a çevirir.

    Türk (1.234,56) ve uluslararası (1234567.89 / 1,234.56) formatları destekler.
    """
    if metin is None:
        return None
    try:
        metin = str(metin).strip()
        if not metin:
            return None
        virgul = metin.rfind(",")
        nokta = metin.rfind(".")
        if virgul > nokta:
            temiz = metin.replace(".", "").replace(",", ".")
        elif nokta > virgul:
            temiz = metin.replace(",", "")
        else:
            temiz = metin.replace(",", ".")
        temiz = re.sub(r'[^\d.\-]', '', temiz)
        if temiz:
            return float(temiz)
    except Exception:
        pass
    return None


def _tablo_basliklarini_bul(html_metin):
    """HTML tablosundaki başlık satırını bulup sütun indekslerini döndürür.

    Döndürdüğü dict: {"matrah": 5, "kdv": 6, "genel_toplam": 7, ...}
    """
    kolon = {}
    baslik_desen = {
        "matrah": re.compile(r"matrah", re.IGNORECASE),
        "kdv": re.compile(r"\bkdv\b", re.IGNORECASE),
        "genel_toplam": re.compile(r"genel\s*toplam|toplam\s*tutar|ödenecek",
                                   re.IGNORECASE),
        "belge_numarasi": re.compile(r"belge\s*no|fatura\s*no", re.IGNORECASE),
        "belge_tarihi": re.compile(r"tarih", re.IGNORECASE),
    }
    for tr_eslesme in _FATURA_TR.finditer(html_metin):
        tr_icerik = tr_eslesme.group(1)
        td_liste = _FATURA_TD.findall(tr_icerik)
        if len(td_liste) < 3:
            continue
        eslesme_sayisi = 0
        for j, td in enumerate(td_liste):
            import html as _html_mod
            temiz = _html_mod.unescape(re.sub(r'<[^>]+>', '', td)).strip()
            for anahtar, desen in baslik_desen.items():
                if desen.search(temiz) and anahtar not in kolon:
                    kolon[anahtar] = j
                    eslesme_sayisi += 1
        if eslesme_sayisi >= 2:
            break
    return kolon


@retry(
    stop=stop_after_attempt(2),
    wait=wait_exponential(multiplier=1, min=1, max=5),
    retry=retry_if_exception_type(Exception),
    reraise=True
)
def _satirlari_ayikla(html_metin):
    """gib530 listesindeki her satirin 'fatura' JSON ozelligini cozer.

    Donen liste [(satir_sirasi, sozluk), ...]; satir_sirasi DOM sirasidir.
    Ek olarak HTML tablosundan matrah/kdv/toplam degerlerini cikarir.
    """
    kolonlar = _tablo_basliklarini_bul(html_metin)
    # fatura iceren TR'leri bul (baslik satirlarini atla)
    fatura_tr = []
    for tr in _FATURA_TR.finditer(html_metin):
        # fatura attribute'u <tr> taginin icinde, group(0) tum eslesme
        if _FATURA_JSON_RE.search(tr.group(0)):
            fatura_tr.append(tr.group(1))
    satirlar = []
    for sira, ham in enumerate(_fatura_json_bul(html_metin)):
        try:
            veri = json.loads(html_cevir.unescape(ham))
        except Exception:
            continue
        # HTML tablosundan matrah/kdv/toplam cek
        if sira < len(fatura_tr):
            td_liste = _FATURA_TD.findall(fatura_tr[sira])
            for alan, anahtar in (("matrah", "matrah_html"),
                                  ("kdv", "kdv_html"),
                                  ("genel_toplam", "toplam_html")):
                idx = kolonlar.get(alan)
                if idx is not None and idx < len(td_liste):
                    ham_hucre = re.sub(r'<[^>]+>', '', td_liste[idx])
                    hucre = html_cevir.unescape(ham_hucre).strip()
                    deger = _tutar_cevir(hucre)
                    if deger is not None:
                        veri[anahtar] = deger
        satirlar.append((sira, veri))
    return satirlar


def _tarih_araliginda(metin, bas_tarih, bit_tarih):
    if not metin:
        return False
    # Luca HTML'inde tarih formatı değişebilir: dd/MM/yyyy, dd.MM.yyyy, yyyy-MM-dd
    for fmt in ("%d/%m/%Y", "%d.%m.%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            gun = datetime.strptime(metin.strip(), fmt).date()
            return bas_tarih <= gun <= bit_tarih
        except (TypeError, ValueError):
            continue
    return False


@retry(
    stop=stop_after_attempt(3),
    wait=wait_exponential(multiplier=1, min=2, max=10),
    retry=retry_if_exception_type(Exception),
    reraise=True
)
def _zip_tikla_indir(frame, sayfa, satir_sirasi, hedef_yol):
    """Satirdaki ZIP ikonuna tiklar; indigi dosyayi kaydeder.
    
    Farklı sayfalarda ZIP butonu farkli selector'larla olabilir:
    - onclick="zip_indir(...)" 
    - onclick="indir('zip', ...)"
    - class="zip-btn" / title="ZIP"
    - img[src*="zip"] / i.fa-file-archive
    """
    # Çoklu selector dene
    selectors = [
        "[onclick*='zip_indir']",
        "[onclick*=\"zip_indir\"]",
        "[onclick*='indir'][onclick*='zip']",
        "[onclick*='indir'][onclick*='ZIP']",
        "a[title*='ZIP']",
        "button[title*='ZIP']",
        "img[src*='zip']",
        "i.fa-file-archive-o",
        "i.fa-file-zip-o",
        ".zip-btn",
        "[data-tip*='ZIP']",
    ]
    
    js_template = """
        (args) => {
            const n = args.n;
            const selectors = args.selectors;
            for (const sel of selectors) {
                const els = [...document.querySelectorAll(sel)];
                if (els.length > n) {
                    els[n].click();
                    return;
                }
            }
            // Fallback: onclick içinde 'zip' veya 'indir' geçen herhangi bir element
            const all = [...document.querySelectorAll('[onclick]')]
                .filter(x => {
                    const oc = (x.getAttribute('onclick') || '').toLowerCase();
                    return oc.includes('zip') || oc.includes('indir');
                });
            if (all.length > n) {
                all[n].click();
                return;
            }
            throw new Error('ZIP düğmesi yok (denenen: ' + selectors.join(', ') + ')');
        }
    """
    # Playwright evaluate: expression + 1 arg (object). selectors'ı JSON olarak geçir.
    import json
    args = {"n": satir_sirasi, "selectors": selectors}
    with sayfa.expect_download(timeout=30000) as bekle:
        frame.evaluate(js_template, args)
    indirme = bekle.value
    indirme.save_as(hedef_yol)
    return indirme.suggested_filename



def _zip_toplu_indir(frame, sayfa, indirme_planlari, bildir=None,
                     pencere=4):
    """ZIP indirmelerini kademeli paralel tetikler.

    `indirme_planlari`: [(satir_sirasi, hedef_yol), ...] çiftleri.
    Playwright'in download olaylarını toplayıcıyla yakalar; her 'pencere'
    adedinde bir sonraki grubu tetikler. Kritik fark: her satır için hedef
    dosya yolu ÖNCEDEN belli olduğundan, indirilen şey yanlış dosyaya
    yazılmaz (suggested_filename güvenilmez). Başarısız olan satırları
    döndürür (kalıp güvenilirliği için).
    """
    import queue
    hedef_map = {sira: yol for sira, yol in indirme_planlari}
    kuyruk = queue.Queue()
    dinleyici = lambda d: kuyruk.put(d)
    sayfa.on("download", dinleyici)
    basarisiz = []
    try:
        bekleyen = [sira for sira, _ in indirme_planlari]
        # Aktif click sayısını değil, inen dosya sayısını izle; kuyruk
        # uzun süre sessiz kalırsa (bazı click'ler indirme başlatmıyorsa)
        # o siparişleri kalanlara geri koy / başarısız işaretle.
        while bekleyen:
            # Yeni clicking: pencere kadar aktif download bekleyebiliriz.
            while bekleyen:
                sira = bekleyen.pop(0)
                hedef = hedef_map[sira]
                try:
                    frame.evaluate(
                        "n => { const e = [...document.querySelectorAll("
                        "'[onclick]')]"
                        ".filter(x => x.getAttribute('onclick')"
                        ".includes('zip_indir'))"
                        "[n]; if (!e) throw new Error('yok'); e.click(); }",
                        sira)
                except Exception:
                    basarisiz.append(sira)
                    continue
                # Belirli kısa süre içinde en az pencere kadar download
                # gelmeye devam etmeli; aksi halde tıklama bir işe
                # yaramadı, bir sonraki adıma geç.
                try:
                    d = kuyruk.get(timeout=12)
                except Exception:
                    # İndirme başlatılamadı — tekrar dene (bir kez).
                    try:
                        frame.evaluate(
                            "n => { const e = [...document.querySelectorAll("
                            "'[onclick]')]"
                            ".filter(x => x.getAttribute('onclick')"
                            ".includes('zip_indir'))"
                            "[n]; if (!e) throw new Error('yok'); e.click(); }",
                            sira)
                        d = kuyruk.get(timeout=15)
                    except Exception:
                        basarisiz.append(sira)
                        continue
                try:
                    d.save_as(hedef)
                except Exception:
                    basarisiz.append(sira)
                time.sleep(0.05)
        # Kalan kuyruktaki olası download'ları da kaydet.
        while True:
            try:
                d = kuyruk.get_nowait()
            except Exception:
                break
            # En iyi tahmini hedef: hedef_map'te kalanlardan birine eşle.
            for sira, yol in hedef_map.items():
                if not os.path.exists(yol):
                    try:
                        d.save_as(yol)
                        break
                    except Exception:
                        continue
    finally:
        try:
            sayfa.remove_listener("download", dinleyici)
        except Exception:
            pass
    return basarisiz


def _ubl_ozet(xml_bytes):
    """UBL fatura XML'inden tutarlari cikarir (lxml ile namespace-aware).

    Kurallar:
    - Para birimi TRY olan degerler oncelidir; hic TRY yoksa belge
      para birimi kullanilir ve 'para' alanindan anlasilir.
    - Bazi gondericiler TaxTotal bloklarini tekrarlar; dogrudan TaxAmount
      ile kendi TaxSubtotal toplami tutarli olmayan bloklar sayilmaz,
      boylece KDV iki katına cikmaz.
    - oran_kalemleri: her KDV orani icin {"oran", "matrah", "kdv"}
      sozlugu; cok oranli faturalar capraz kontrolde oran basina
      degerlendirilir. (Uygulamanin diger yerlerindeki oranlar=[18, 8]
      duz oran listesinden farkli alandir.)
    """
    if ET_LXML is None:
        # fallback to stdlib
        return _ubl_ozet_stdlib(xml_bytes)

    try:
        parser = ET_LXML.XMLParser(recover=True, huge_tree=True)
        kok = ET_LXML.fromstring(xml_bytes, parser=parser)
    except Exception:
        return _ubl_ozet_stdlib(xml_bytes)

    # UBL namespace map
    NS = {
        "cac": "urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2",
        "cbc": "urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2",
        "ext": "urn:oasis:names:specification:ubl:schema:xsd:CommonExtensionComponents-2",
    }

    def sayi(metin):
        try:
            return float((metin or "0").replace(",", "."))
        except (ValueError, TypeError):
            return None

    # TaxTotal blocks with xpath
    bloklar = []
    for tax_total in kok.xpath("//cac:TaxTotal", namespaces=NS):
        dogrudan = None
        para_b = ""
        alt = []

        # Direct TaxAmount
        tax_amount_el = tax_total.xpath("cbc:TaxAmount", namespaces=NS)
        if tax_amount_el:
            dogrudan = sayi(tax_amount_el[0].text)
            para_b = tax_amount_el[0].get("currencyID") or para_b

        # TaxSubtotal children
        for subtotal in tax_total.xpath("cac:TaxSubtotal", namespaces=NS):
            yuzde = None
            taban_t = None
            alt_kdv = None

            pct = subtotal.xpath("cbc:Percent", namespaces=NS)
            if pct:
                yuzde = sayi(pct[0].text)

            taxable = subtotal.xpath("cbc:TaxableAmount", namespaces=NS)
            if taxable:
                taban_t = sayi(taxable[0].text)

            sub_tax = subtotal.xpath("cbc:TaxAmount", namespaces=NS)
            if sub_tax:
                alt_kdv = sayi(sub_tax[0].text)

            if yuzde is not None and alt_kdv is not None:
                alt.append({"oran": round(yuzde, 2),
                            "matrah": round(taban_t, 2) if taban_t is not None else None,
                            "kdv": round(alt_kdv, 2)})

        tutarli = bool(alt) and dogrudan is not None and \
            abs(sum(a["kdv"] for a in alt) - dogrudan) <= 0.02
        bloklar.append({"para": para_b, "dogrudan": dogrudan,
                        "alt": alt, "tutarli": tutarli})

    secili = [b for b in bloklar if b["tutarli"]]
    if not secili:
        secili = [b for b in bloklar if b["alt"]] or bloklar

    # PayableAmount & TaxExclusiveAmount
    genels = []
    matrahs = []

    for pa in kok.xpath("//cbc:PayableAmount", namespaces=NS):
        genels.append((sayi(pa.text), pa.get("currencyID") or ""))
    for ta in kok.xpath("//cbc:TaxExclusiveAmount", namespaces=NS):
        matrahs.append((sayi(ta.text), ta.get("currencyID") or ""))

    def tercih(ciftler):
        temiz = [(d, p) for d, p in ciftler if d is not None]
        if not temiz:
            return None, ""
        for d, p in temiz:
            if p in ("TRY", "TL"):
                return d, "TRY"
        return temiz[0]

    genel_toplam, genel_para = tercih(genels)
    matrah, matrah_para = tercih(matrahs)

    # Blok secimi: gondericiler hem satir kirilimini hem belge toplamini
    # ayri TaxTotal bloklari olarak yazabiliyor. Tabanlari toplami KDV-
    # haric matraya esit olan blok(lar) gercek vergi bilesenleridir;
    # matraya eslesen tek blok varsa digerleri (kismi/kopya) atlanir.
    if len(secili) > 1 and matrah is not None:
        tam_bloklar = [b for b in secili
                       if abs(sum((a.get("matrah") or 0.0)
                                  for a in b["alt"]) - matrah) <= 0.05]
        if len(tam_bloklar) == 1:
            secili = tam_bloklar

    # Alt kalemleri temizle: belge duzeyinde tekrarlanan ozet satirini
    # (tabani diger satirlarin toplami) ve sent kopyalarini kaldirir.
    tum_alt = []
    for b in secili:
        tum_alt.extend(b["alt"])
    temiz = []
    for i, a in enumerate(tum_alt):
        taban_a = a.get("matrah")
        if taban_a is None:
            continue
        digerler = [x for j, x in enumerate(tum_alt) if j != i]
        if digerler and abs(
                sum((x.get("matrah") or 0.0) for x in digerler)
                - taban_a) <= 0.05:
            continue
        temiz.append(a)

    oranlar = []
    for a in sorted(temiz,
                    key=lambda x: (-x["oran"], -x["kdv"])):
        benzer = [b for b in oranlar
                  if b["oran"] == a["oran"]
                  and abs(b["kdv"] - a["kdv"]) <= 0.05]
        if not benzer:
            oranlar.append(a)

    # KDV toplami: tutarli bloklarin dogrudan degerleri; yakin degerler
    # tek sayilir (tekrarlanan TaxTotal bloklari yaygin).
    kdvs = []
    for b in secili:
        d = b["dogrudan"]
        if d is None:
            continue
        if not any(abs(d - y) <= 0.05 for y in kdvs):
            kdvs.append(round(d, 2))

    para = "TRY" if (genel_para == "TRY" or matrah_para == "TRY"
                     or any(b["para"] in ("TRY", "TL")
                            for b in secili)) else (
        genel_para or (secili[0]["para"] if secili else ""))

    ozet = {}
    if matrahs:
        ozet["matrah"] = round(matrah, 2)
    if kdvs:
        ozet["kdv_toplam"] = round(sum(kdvs), 2)
    if genel_toplam is not None:
        ozet["genel_toplam"] = round(genel_toplam, 2)
    if para:
        ozet["para"] = para
    if oranlar:
        ozet["oran_kalemleri"] = oranlar
    return ozet


def _ubl_ozet_stdlib(xml_bytes):
    """Fallback: stdlib ElementTree ile UBL parse (eski kod)."""
    try:
        import xml.etree.ElementTree as ET
        kok = ET.fromstring(xml_bytes)
    except Exception:
        return {}

    def yerel(etiket):
        return etiket.split("}")[-1]

    def sayi(metin):
        try:
            return float((metin or "0").replace(",", "."))
        except ValueError:
            return None

    bloklar = []
    for eb in kok.iter():
        if yerel(eb.tag) != "TaxTotal":
            continue
        dogrudan = None
        para_b = ""
        alt = []
        for el in eb:
            ad = yerel(el.tag)
            if ad == "TaxAmount":
                dogrudan = sayi(el.text)
                para_b = el.get("currencyID") or para_b
            elif ad == "TaxSubtotal":
                yuzde = None
                taban_t = None
                alt_kdv = None
                for a in el:
                    aa = yerel(a.tag)
                    if aa == "Percent":
                        yuzde = sayi(a.text)
                    elif aa == "TaxableAmount":
                        taban_t = sayi(a.text)
                    elif aa == "TaxAmount":
                        alt_kdv = sayi(a.text)
                if yuzde is not None and alt_kdv is not None:
                    alt.append({"oran": round(yuzde, 2),
                                "matrah": round(taban_t, 2)
                                if taban_t is not None else None,
                                "kdv": round(alt_kdv, 2)})
        tutarli = bool(alt) and dogrudan is not None and \
            abs(sum(a["kdv"] for a in alt) - dogrudan) <= 0.02
        bloklar.append({"para": para_b, "dogrudan": dogrudan,
                        "alt": alt, "tutarli": tutarli})

    secili = [b for b in bloklar if b["tutarli"]]
    if not secili:
        secili = [b for b in bloklar if b["alt"]] or bloklar

    genels = []
    matrahs = []
    for el in kok.iter():
        ad = yerel(el.tag)
        if ad == "PayableAmount":
            genels.append((sayi(el.text),
                           el.get("currencyID") or ""))
        elif ad == "TaxExclusiveAmount":
            matrahs.append((sayi(el.text),
                            el.get("currencyID") or ""))

    def tercih(ciftler):
        temiz = [(d, p) for d, p in ciftler if d is not None]
        if not temiz:
            return None, ""
        for d, p in temiz:
            if p in ("TRY", "TL"):
                return d, "TRY"
        return temiz[0]

    genel_toplam, genel_para = tercih(genels)
    matrah, matrah_para = tercih(matrahs)

    if len(secili) > 1 and matrah is not None:
        tam_bloklar = [b for b in secili
                       if abs(sum((a.get("matrah") or 0.0)
                                  for a in b["alt"]) - matrah) <= 0.05]
        if len(tam_bloklar) == 1:
            secili = tam_bloklar

    tum_alt = []
    for b in secili:
        tum_alt.extend(b["alt"])
    temiz = []
    for i, a in enumerate(tum_alt):
        taban_a = a.get("matrah")
        if taban_a is None:
            continue
        digerler = [x for j, x in enumerate(tum_alt) if j != i]
        if digerler and abs(
                sum((x.get("matrah") or 0.0) for x in digerler)
                - taban_a) <= 0.05:
            continue
        temiz.append(a)

    oranlar = []
    for a in sorted(temiz,
                    key=lambda x: (-x["oran"], -x["kdv"])):
        benzer = [b for b in oranlar
                  if b["oran"] == a["oran"]
                  and abs(b["kdv"] - a["kdv"]) <= 0.05]
        if not benzer:
            oranlar.append(a)

    kdvs = []
    for b in secili:
        d = b["dogrudan"]
        if d is None:
            continue
        if not any(abs(d - y) <= 0.05 for y in kdvs):
            kdvs.append(round(d, 2))

    para = "TRY" if (genel_para == "TRY" or matrah_para == "TRY"
                     or any(b["para"] in ("TRY", "TL")
                            for b in secili)) else (
        genel_para or (secili[0]["para"] if secili else ""))

    ozet = {}
    if matrahs:
        ozet["matrah"] = round(matrah, 2)
    if kdvs:
        ozet["kdv_toplam"] = round(sum(kdvs), 2)
    if genel_toplam is not None:
        ozet["genel_toplam"] = round(genel_toplam, 2)
    if para:
        ozet["para"] = para
    if oranlar:
        ozet["oran_kalemleri"] = oranlar
    return ozet


def _ozet_tablo_yaz(yol, kayitlar):
    """Belge ozetini .xlsx olarak yazar; openpyxl yoksa .csv dener."""
    sutunlar = [("belge_numarasi", "Belge No"),
                ("belge_tarihi", "Tarih"),
                ("belge_turu", "Tür"),
                ("karsi_vkn", "VKN/TCKN"),
                ("unvan", "Unvan"),
                ("onay_durumu", "Durum"),
                ("matrah", "Matrah"),
                ("kdv_toplam", "KDV"),
                ("genel_toplam", "Genel Toplam"),
                ("oranlar_metni", "KDV Oranlar"),
                ("para", "Para"),
                ("ettn", "ETTN"),
                ("dosya", "ZIP")]
    try:
        from openpyxl import Workbook
        kitap = Workbook()
        yaprak = kitap.active
        yaprak.title = "belgeler"
        yaprak.append([baslik for _, baslik in sutunlar])
        for kayit in kayitlar:
            yaprak.append([kayit.get(anahtar, "")
                           for anahtar, _ in sutunlar])
        kitap.save(yol)
        return yol
    except Exception:
        csv_yol = os.path.splitext(yol)[0] + ".csv"
        with open(csv_yol, "w", encoding="utf-8-sig", newline="") as dosya:
            dosya.write(",".join(b for _, b in sutunlar) + "\n")
            for kayit in kayitlar:
                hucreler = [str(kayit.get(a, "")).replace('"', "'")
                            for a, _ in sutunlar]
                dosya.write(",".join('"' + h + '"' for h in hucreler) + "\n")
        return csv_yol


def _belge_sorgula_ve_indir(sayfa, hedef, bildir):
    """Ekranda tarihleri doldurmus varsayarak sorgula + Excel indir.

    Once sorgulama dongusuyle dogrudan indirmeyi, olmazsa ayri Excel/dokum
    dugmesini dener. Basarida True doner.
    """
    raporda_indi = False
    try:
        with sayfa.expect_download(timeout=25000) as indirme:
            _indir_butonu_tikla(
                sayfa, (r"^rapor$", r"^liste$", r"listele", r"sorgula",
                        r"getir"), zaman_asimi=6)
        indirme.value.save_as(hedef)
        raporda_indi = True
    except Exception:
        pass
    if not raporda_indi:
        try:
            with sayfa.expect_download(timeout=20000) as indirme:
                if not _indir_butonu_tikla(
                        sayfa, (r"excel", r"\bxls\b", r"aktar",
                                r"d[öo]k[üu]m\s*al")):
                    raise RuntimeError("Excel/döküm düğmesi bulunamadı")
            indirme.value.save_as(hedef)
            raporda_indi = True
        except Exception as hata:
            bildir(f"Döküm alınamadı ({str(hata)[:70]}).")
    return raporda_indi


def _aralik_parcalara_bol(bas_tarih, bit_tarih, adim_gun):
    """Tarih aralığını en fazla adim_gun günlük parçalara böler."""
    parcalar = []
    b = bas_tarih
    while b <= bit_tarih:
        s = min(b.fromordinal(b.toordinal() + adim_gun - 1), bit_tarih)
        parcalar.append((b, s))
        b = b.fromordinal(b.toordinal() + adim_gun)
    return parcalar


def _dosya_saglam(yol):
    """İndirilen dosya makul boyutta mı (boş/yanlış sayfa indirmesi)."""
    try:
        return os.path.getsize(yol) >= 400
    except OSError:
        return False


def _satir_sayisini_buyut(sayfa, bildir):
    """Sayfalama satır seçicisinde yüksek değer/‘tümü’ seçer (best effort)."""
    try:
        return _indir_butonu_tikla(
            sayfa, (r"^\s*(t[üu]m[üu]|500|1000|2000)\s*$",
                    r"sat[ıi]r\s*say[ıi]s"), zaman_asimi=2)
    except Exception:
        return False


def _sayfa_butonlari(cerceve):
    """Luca belge listesindeki sayfa ilerleme düğmelerini döndürür.

    Geniş tarama: input/button/a + span/div + onclick içeren her öğe.
    Metin ya da onclick'te 'sonraki/ileri/next/»/›/>' veya 'sayfa 2'
    deseni aranır; 'önceki/geri' hariç.
    """
    adaylar = []
    try:
        ogeler = cerceve.query_selector_all(
            "input, button, a, span, div, td, li, img, b, i")
        for oge in ogeler:
            try:
                metin = ((oge.get_attribute("value") or "")
                         + " " + (oge.inner_text() or "")
                         + " " + (oge.get_attribute("onclick") or "")
                         + " " + (oge.get_attribute("title") or "")
                         + " " + (oge.get_attribute("alt") or "")).strip()
                if not metin:
                    continue
                k = metin.lower()
                # 'önceki/geri' hariç; ilerleme desenleri
                if "onceki" in k or "geri" in k:
                    continue
                ilerleme = (
                    "sonraki" in k or "ileri" in k or "next" in k
                    or "»" in k or "›" in k or ">>" in k
                    or k.strip() in (">", "→")
                    or "sayfa 2" in k or "sayfa2" in k
                    or re.search(r"goPage|nextPage|sayfaGec|ileri", k)
                    or re.search(r"pager", k)
                )
                if ilerleme:
                    adaylar.append(oge)
            except Exception:
                continue
    except Exception:
        pass
    return adaylar


def _sonraki_sayfa_var_mi(cerceve):
    """Luca belge listesinde 'Sonraki' / ileri sayfa düğmesi varsa True."""
    return len(_sayfa_butonlari(cerceve)) > 0


def _sonraki_sayfaya_git(cerceve):
    """Listedeki 'Sonraki' / 'İleri' düğmesine tıklar; yönlendirme
    sonrası yeni sayfa içeriği yüklenir. Dönüş: True/False."""
    adaylar = _sayfa_butonlari(cerceve)
    for oge in adaylar:
        try:
            oge.scroll_into_view_if_needed()
        except Exception:
            pass
        try:
            oge.click()
        except Exception:
            continue
        try:
            cerceve.page.wait_for_timeout(1300)
        except Exception:
            pass
        # Tıklama sonrası gerçekten sayfa değişti mi? İlk sayfa ile
        # aynı belgeleri tekrar görüyorsak ilerlememiş olabilir;
        # yine de döngü dedup ile yönetir.
        return True
    return False


def cek_luca_belgeleri(uye_no, kullanici, parola, bas_tarih, bit_tarih,
                       hedef_klasor, kategoriler=None, ilerleme=None,
                       gorunur=True, firma_adi=None, duz_yaz=True,
                       olay=None, onay_callback=None):
    """Luca ERP Akıllı Entegrasyon ekranlarından e-Belgeleri indirir.

    Gerçek akış: giriş → portalda gonder('formTarget') ile MM Paketi
    penceresi → SirketCombo/DonemCombo ile firma+dönem seçimi → her
    kategori için gib530.do?tur=... ekranı → listedeki 'fatura' JSON'ları
    okunur, tarih aralığına göre süzülür → her belgenin ZIP'i satırdaki
    ZIP simgesiyle indirilip açılır (UBL XML + HTML).

    firma_adi: firma unvanının içinde geçen ibare; Türkçe büyük/küçük
    harf duyarsız eşleşir. Tek firmalık hesapta boş bırakılabilir.

    gorunur=True önerilir; captcha kullanıcı tarafından tarayıcı
    penceresinde elle girilir, bu yüzden görünür tarayıcı gerekir.

    Kategori başına çıktılar hedef_klasor altına yazılır:
      {kategori}_{dosya}.zip/.xml/.html   → düz yazım (duz_yaz=True)
      luca_{kategori}_{bas}_{bit}/        → alt klasör (duz_yaz=False)
      luca_{kategori}_{bas}_{bit}.xlsx    → özet tablo

    olay: yapılandırılmış ilerleme takibi için sözlük alan callback.
    Her çağrı şöyledir (UI takip ekranı için):
      {\"kategori\": ..., \"adim\": ..., \"durum\": ..., \"sayi\": ...,
       \"toplam\": ..., \"mesaj\": ...}
    adim: 'basla'|'liste'|'indirme'|'indirildi'|'ozet'|'hata'|'bitti'
    durum: 'calisiyor'|'tamam'|'hata'|'atlandi'

    Dönen değer: {kategori: {"zip": [yollar], "ozet": xlsx_yolu,
    "belge_sayisi": n}}. Hiçbir kategori inmezse LucaHata.
    """
    bildir = _bildir_fonksiyonu(ilerleme)
    if not olay:
        olay = lambda o: None
    if not gorunur:
        raise LucaHata(
            "Captcha kullanıcı tarafından elle girildiği için Luca çekimi "
            "yalnız görünür tarayıcıyla çalışır (gorunur=True kullanın).")
    os.makedirs(hedef_klasor, exist_ok=True)
    hedefler = kategoriler or [k for k in LUCA_GIB_TURLER]
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        raise LucaHata(
            "Playwright kurulu değil. Kurulum: pip install playwright")

    sonuc = {}
    with sync_playwright() as p:
        tarayici = _tarayici_ac(p, gorunur=gorunur)
        oturum = _luca_oturum_ac(tarayici)
        sayfa = oturum.new_page()
        try:
            sayfa = giris_yap(sayfa, uye_no, kullanici, parola, bildir)
            sayfa.wait_for_timeout(1300)
            erp = _erp_penceresi(oturum, sayfa, bildir)
            _firma_donem_sec(erp, firma_adi, bas_tarih, bildir)

            for kategori in hedefler:
                tur = LUCA_GIB_TURLER.get(kategori)
                if tur is None:
                    bildir(f"Bilinmeyen kategori atlandı: {kategori}")
                    continue
                birim = _kategori_etiketi(kategori)
                olay({"kategori": kategori, "adim": "basla",
                      "durum": "calisiyor", "sayi": 0, "toplam": 0,
                      "mesaj": f"{birim} ekranı açılıyor..."})
                bildir(f"{kategori}: ekran açılıyor...")
                try:
                    cerceve = _gib530_frame(erp, tur, uye_no, bildir)
                    if cerceve is None:
                        olay({"kategori": kategori, "adim": "hata",
                              "durum": "hata", "sayi": 0, "toplam": 0,
                              "mesaj": "gib530 ekranı yüklenemedi"})
                        raise RuntimeError("gib530 ekranı yüklenmedi")
                    # TANI: gerçek ekran HTML'ini kaydet (çekim kökten
                    # tasarlanırken kullanılacak).
                    try:
                        _tani_kaydet(kategori, cerceve)
                    except Exception:
                        pass
                    # İKİ ADIMLI AKIŞ: Luca'nın e-belge ekranında belgeler
                    # GİB'ten önce 'çekilir' (GİB'ten Getir / İnternetten
                    # Getir butonu), sonra listelenip indirilir. Bu adım
                    # atlanırsa yeni mükelleflerde belge listesi boş kalır.
                    try:
                        _gibten_getir(cerceve, bas_tarih, bit_tarih, bildir)
                    except Exception as g_hata:
                        bildir(f"{kategori}: GİB'ten getir adımı "
                               f"atlandı ({str(g_hata)[:60]})")
                        _hata_ekrani_kaydet(cerceve, f"getir_{kategori}")
                    # İlk olarak satır sayısını artırmayı dene (500/1000/tümü);
                    # Luca tek sayfada en fazla ~500 fatura listeler.
                    try:
                        bildir(f"{kategori}: satır sayısı artırılıyor...")
                        _satir_sayisini_buyut(cerceve, bildir)
                        cerceve.page.wait_for_timeout(1000)
                        bildir(f"{kategori}: satır sayısı artırıldı.")
                    except Exception as hata:
                        bildir(f"{kategori}: satır sayısı artırılamadı "
                               f"({str(hata)[:40]})")
                    # TEK GEÇİŞ: Her sayfadayken hem topla hem ZIP indir.
                    # Böylece sayfa geçişi sonrası frame durumundan etkilenmez.
                    SAYFA_LIMITI = 500
                    tum_satirlar = []           # (sayfa_no, sira, belge)
                    gorulen_belge = set()
                    for sayfa_sirasi in range(1, 60):
                        try:
                            html_icerik = cerceve.content()
                        except Exception as hata:
                            bildir(f"{kategori}: cerceve.content() hatası: "
                                   f"{str(hata)[:80]}")
                            try:
                                html_icerik = cerceve.page.content()
                                bildir(f"{kategori}: sayfadan HTML alındı "
                                       f"({len(html_icerik)} bayt)")
                            except Exception:
                                break
                        sayfa_satirlari = _satirlari_ayikla(html_icerik)
                        bildir(f"{kategori}: sayfa {sayfa_sirasi} - "
                               f"HTML {len(html_icerik)} bayt, "
                               f"{len(sayfa_satirlari)} belge.")
                        yeni = 0
                        for sira, belge in sayfa_satirlari:
                            anahtar = (str(belge.get("belge_numarasi") or "")
                                       + "|" + str(belge.get("belge_tarihi") or "")
                                       + "|" + str(belge.get("ettn") or ""))
                            if anahtar not in gorulen_belge:
                                gorulen_belge.add(anahtar)
                                tum_satirlar.append((sayfa_sirasi, sira, belge))
                                yeni += 1
                        if not sayfa_satirlari or yeni == 0:
                            break
                        # Sayfada tam limit kadar belge var mı? O zaman
                        # mutlaka devamı vardır.
                        tam_sayfa = len(sayfa_satirlari) >= SAYFA_LIMITI
                        if not tam_sayfa:
                            break
                        try:
                            sonraki_var = _sonraki_sayfa_var_mi(cerceve)
                            bildir(f"{kategori}: sonraki sayfa var mi: "
                                   f"{sonraki_var}")
                        except Exception as hata:
                            bildir(f"{kategori}: sonraki sayfa kontrol hata: "
                                   f"{str(hata)[:60]}")
                            sonraki_var = False
                        if sayfa_sirasi >= 30 and sonraki_var:
                            bildir(f"{kategori}: {sayfa_sirasi}. sayfa "
                                   "toplandı, ilerleniyor...")
                        if not sonraki_var and not tam_sayfa:
                            break
                        if not _sonraki_sayfaya_git(cerceve):
                            if tam_sayfa:
                                try:
                                    cerceve.page.wait_for_timeout(1300)
                                    continue
                                except Exception:
                                    pass
                            break
                    # klasor/on_ek'i ZIP planlaması için ERKEK tanımla
                    if duz_yaz:
                        klasor = hedef_klasor
                    else:
                        klasor = os.path.join(
                            hedef_klasor,
                            f"luca_{kategori}_{bas_tarih:%Y%m%d}_"
                            f"{bit_tarih:%Y%m%d}")
                    on_ek = ("" if duz_yaz
                             else f"luca_{kategori}_")
                    os.makedirs(klasor, exist_ok=True)
                    # Tarih aralığında kalan ve geçerli (red/iptal olmayan)
                    # belgelerin tam listesi (belge no, sayac ile).
                    gorulen_no = {}
                    tum_secili = []
                    atlanan_belge = 0
                    for sayfa_no, sira, belge in tum_satirlar:
                        if not _tarih_araliginda(belge.get("belge_tarihi"),
                                                 bas_tarih, bit_tarih):
                            continue
                        durum_kisa = _turk_kucult(
                            str(belge.get("onay_durumu") or ""))
                        iptal_ibare = str(belge.get("iptal_itiraz") or
                                          belge.get("iptal_itiraz_durumu")
                                          or "").strip()
                        if (iptal_ibare
                                or ("red" in durum_kisa)
                                or ("iptal" in durum_kisa)
                                or (durum_kisa and "onay" not in durum_kisa)):
                            atlanan_belge += 1
                            continue
                        belge_no = (belge.get("belge_numarasi")
                                    or f"belge{sira}").strip()
                        gorulen_no[belge_no] = \
                            gorulen_no.get(belge_no, 0) + 1
                        tum_secili.append((sayfa_no, sira, belge, belge_no,
                                           gorulen_no[belge_no]))
                    # Tarih alanları doldurulamadıysa Luca kendi
                    # dönemini kullanmış olabilir; filtre Hiç belge
                    # tutmadıysa tarihsiz tüm belgeleri al.
                    if not tum_secili and tum_satirlar:
                        bildir(f"{kategori}: tarih filtresi hiçbir belge "
                               "tutmadı; tüm belgeler alınıyor.")
                        gorulen_no = {}
                        tum_secili = []
                        atlanan_belge = 0
                        for sayfa_no, sira, belge in tum_satirlar:
                            durum_kisa = _turk_kucult(
                                str(belge.get("onay_durumu") or ""))
                            iptal_ibare = str(belge.get("iptal_itiraz") or
                                              belge.get("iptal_itiraz_durumu")
                                              or "").strip()
                            if (iptal_ibare
                                    or ("red" in durum_kisa)
                                    or ("iptal" in durum_kisa)
                                    or (durum_kisa
                                        and "onay" not in durum_kisa)):
                                atlanan_belge += 1
                                continue
                            belge_no = (belge.get("belge_numarasi")
                                        or f"belge{sira}").strip()
                            gorulen_no[belge_no] = \
                                gorulen_no.get(belge_no, 0) + 1
                            tum_secili.append((sayfa_no, sira, belge, belge_no,
                                               gorulen_no[belge_no]))
                    secili = [(sayfa_no, sira, belge, belge_no)
                              for sayfa_no, sira, belge, belge_no, _ in tum_secili]
                    bildir(f"{kategori}: listede {len(tum_satirlar)} belge, "
                           f"tarih aralığında {len(secili)} tanesi var.")
                    olay({"kategori": kategori, "adim": "liste",
                          "durum": "calisiyor", "sayi": len(secili),
                          "toplam": len(secili),
                          "mesaj": f"{birim}: {len(secili)} belge bulundu"})
                    zip_yollari = []
                    kayitlar = []
                    bildir(f"{kategori}: {len(secili)} belge indirilecek.")

                    # ZIP indirme: filtrelenmiş belgeler için ikinci geçiş
                    # Her sayfayı tekrar gez, o sayfadaki seçili belgeleri indir.
                    if kategori in ("efatura_alis", "efatura_satis",
                                    "earsiv_alis", "earsiv_satis") and secili:
                        # ÖNCE: Sayfa 1'e dönmek için frame'i yeniden yükle
                        # Pagination loop sonunda son sayfadayız, sayfa 1'e dönmek lazım
                        try:
                            tur = LUCA_GIB_TURLER.get(kategori)
                            if tur:
                                cerceve.evaluate(
                                    "u => { window.location.href = u; }",
                                    f"gib530.do?tur={tur}&c_musteri_id={uye_no}")
                                cerceve.page.wait_for_timeout(2000)
                                bildir(f"{kategori}: sayfa 1'e dönüldü (frame yenilendi)")
                        except Exception as e:
                            bildir(f"{kategori}: sayfa 1'e dönüş hatası: {e}")
                        # Seçili belgeleri sayfa_no'ya göre grupla
                        from collections import defaultdict
                        secili_sayfa = defaultdict(list)
                        for sayfa_no, sira, belge, belge_no in secili:
                            secili_sayfa[sayfa_no].append((sira, belge, belge_no))
                        # Sayfa 1'den başla, her sayfa için ZIP indir
                        for hedef_sayfa in sorted(secili_sayfa.keys()):
                            # Hedef sayfaya git
                            if hedef_sayfa > 1:
                                # Baştan başla ve hedef sayfaya kadar ilerle
                                # (Luca'da doğrudan sayfa atlama yok)
                                bildir(f"{kategori}: sayfa {hedef_sayfa} için ZIP indiriliyor...")
                                for _ in range(hedef_sayfa - 1):
                                    if not _sonraki_sayfaya_git(cerceve):
                                        break
                                    cerceve.page.wait_for_timeout(1000)
                            # Bu sayfadaki seçili belgeleri indir
                            for sira, belge, belge_no in secili_sayfa[hedef_sayfa]:
                                zip_yol = os.path.join(klasor, f"{on_ek}{belge_no}.zip")
                                try:
                                    _zip_tikla_indir(cerceve, cerceve.page, sira, zip_yol)
                                    ubl_ozet = _zipten_ozet(zip_yol, klasor)
                                    if ubl_ozet:
                                        belge["matrah"] = ubl_ozet.get("matrah")
                                        belge["kdv_toplam"] = ubl_ozet.get("kdv_toplam")
                                        belge["genel_toplam"] = ubl_ozet.get("genel_toplam")
                                        belge["para"] = ubl_ozet.get("para", "TRY")
                                        belge["oran_kalemleri"] = ubl_ozet.get("oran_kalemleri", [])
                                except Exception as e:
                                    bildir(f"{kategori}: {belge_no} ZIP indirme hatası: {e}")
                                zip_yollari.append(zip_yol)

                    # Kayıt oluştur: filtrelenmiş `secili` listesinden.
                    bildir(f"{kategori}: {len(secili)} belge "
                           "kayıt oluşturuluyor...")
                    for sayfa_no, sira, belge, belge_no in secili:
                        # fatura JSON'unda olası alan isimleri + HTML tablosu + UBL ZIP
                        matrah = (belge.get("matrah_html")
                                  or belge.get("matrah")
                                  or belge.get("mal_hizmet_tutari")
                                  or belge.get("matrah_tutari"))
                        kdv = (belge.get("kdv_html")
                               or belge.get("kdv_toplam")
                               or belge.get("kdv")
                               or belge.get("kdv_tutari")
                               or belge.get("toplam_kdv"))
                        toplam = (belge.get("toplam_html")
                                  or belge.get("genel_toplam")
                                  or belge.get("toplam")
                                  or belge.get("toplam_tutar")
                                  or belge.get("genel_toplam_tutari"))
                        # İlk belgede tüm alanları logla (debug)
                        if sira == 0 and secili:
                            bildir(f"{kategori}: belge alanları: "
                                   f"{list(belge.keys())}")
                            bildir(f"{kategori}: matrah={matrah} "
                                   f"kdv={kdv} toplam={toplam}")
                        ozet = {
                            "belge_numarasi": belge_no,
                            "belge_tarihi": belge.get("belge_tarihi", ""),
                            "belge_turu": kategori,
                            "ettn": belge.get("ettn", ""),
                            "karsi_vkn": str(belge.get("alici_vkn_tckn", "")),
                            "unvan": belge.get("alici_unvan_ad_soyad", ""),
                            "onay_durumu": belge.get("onay_durumu", ""),
                            "matrah": matrah,
                            "kdv_toplam": kdv,
                            "genel_toplam": toplam,
                            "para": belge.get("para_birimi",
                                              belge.get("para", "TRY")),
                            "oranlar_metni": belge.get("oranlar_metni",
                                                       belge.get("kdv_oran",
                                                                 "")),
                            "dosya": "",
                        }
                        kayitlar.append(ozet)
                    for numara, (sayfa_no, sira, belge, belge_no) in \
                            enumerate(secili, 1):
                        olay({"kategori": kategori, "adim": "indirildi",
                              "durum": "calisiyor",
                              "sayi": numara,
                              "toplam": len(secili),
                              "mesaj": f"{birim}: {numara}/"
                                       f"{len(secili)} "
                                       f"({belge_no})"})
                        bildir(f"{kategori}: {numara}/"
                               f"{len(secili)} "
                               f"belge ({belge_no}).")
                    if not kayitlar:
                        bildir(f"{kategori}: tarih aralığında belge bulunamadı "
                               "(0 belge).")
                        ozet_yol = None
                        # Eski Excel dosyasını temizle
                        eski_xlsx = os.path.join(
                            hedef_klasor,
                            f"luca_{kategori}_{bas_tarih:%Y%m%d}_"
                            f"{bit_tarih:%Y%m%d}.xlsx")
                        if os.path.exists(eski_xlsx):
                            try:
                                os.remove(eski_xlsx)
                            except Exception:
                                pass
                    else:
                        ozet_yol = _ozet_tablo_yaz(
                            os.path.join(
                                hedef_klasor,
                                f"luca_{kategori}_{bas_tarih:%Y%m%d}_"
                                f"{bit_tarih:%Y%m%d}.xlsx"), kayitlar)
                    sonuc[kategori] = {
                        "zip": zip_yollari,
                        "ozet": ozet_yol,
                        "belge_sayisi": len(kayitlar)}
                    bildir(f"{kategori}: TAMAMLANDI — {len(kayitlar)} belge.")
                    olay({"kategori": kategori, "adim": "bitti",
                          "durum": "tamam", "sayi": len(kayitlar),
                          "toplam": len(secili),
                          "mesaj": f"{birim}: {len(kayitlar)} belge tamam"})
                    # Ekran istiflenmesini önle: kategori sonrası fazladan sekmeleri kapat
                    try:
                        for p in oturum.pages:
                            if p != erp and p != sayfa:
                                p.close()
                    except Exception:
                        pass
                except Exception as hata:
                    # Tek kategori inmedi: digerlerini engelleme.
                    bildir(f"{kategori}: HATA — {str(hata)[:80]}")
                    olay({"kategori": kategori, "adim": "hata",
                          "durum": "hata", "sayi": 0, "toplam": 0,
                          "mesaj": f"{birim}: {str(hata)[:80]}"})
                    _hata_ekrani_kaydet(sayfa, f"belge_{kategori}")
                    # Ekran istiflenmesini önle: hata durumunda da fazladan sekmeleri kapat
                    try:
                        for p in oturum.pages:
                            if p != erp and p != sayfa:
                                p.close()
                    except Exception:
                        pass
        finally:
            tarayici.close()
    if not sonuc:
        raise LucaHata(
            "Luca'dan hiçbir e-Belge kategorisi indirilemedi. Ekran "
            "görüntüleri %TEMP% altına kaydedildi.")
    return sonuc
